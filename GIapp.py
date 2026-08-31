# ==========================================
# 非寿险精算对标报告系统 - 完整版
# 包含：登录认证、官网监控、智能页码定位、图表生成、AI分析、PDF导出
# ==========================================

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import plotly.express as px
import openpyxl
import io
import re
import time
import json
import base64
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from step8 import show_step_8_content

# ==========================================
# 0. 页面基础配置（必须在所有 st 命令之前）
# ==========================================
st.set_page_config(
    page_title="财险数智年报平台 | 对标报告系统",
    page_icon="Digi.png",
    layout="wide"
)

# ==========================================
# KPMG 官方色板（源自官方色卡）
# ==========================================
KPMG_COLORS = [
    "#00338D",  # 深蓝
    "#1E49E2",  # 亮蓝
    "#76D2FF",  # 淡蓝
    "#00B8F5",  # 青蓝
    "#ACEAFF",  # 极淡蓝
    "#510DBC",  # 深紫
    "#B497FF",  # 淡紫
    "#7213EA",  # 紫
    "#AB0D82",  # 玫红
    "#FD349C",  # 粉红
    "#FB8E7E",  # 橙红
    "#00C0AE",  # 青绿
]

# 保活脚本
components.html("""
<script>
    setInterval(() => {
        window.parent.document.dispatchEvent(new Event('mousemove'));
    }, 300000);
</script>
""", height=0, width=0)

# 左上角 Logo（已注释，可自行恢复）
# try:
#     st.logo("Digi.png", icon_image="Digi.png")
# except Exception as e:
#     st.error(f"Logo 加载失败，请检查文件名。错误: {e}")

# ==========================================
# 0.1 CSS 全量美化
# ==========================================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@300;400;500;600;700&display=swap');
    html, body, [class*="css"] { font-family: "PingFang SC", "HarmonyOS Sans", "Microsoft YaHei", "Noto Sans SC", sans-serif !important; color: #4A5568; }
    .stApp { background: linear-gradient(160deg, #F8FAFC 0%, #F1F5F9 50%, #E2E8F0 100%); }
    ::-webkit-scrollbar { width: 6px; height: 6px; }
    ::-webkit-scrollbar-track { background: transparent; }
    ::-webkit-scrollbar-thumb { background: #B0BEC5; border-radius: 3px; }
    ::-webkit-scrollbar-thumb:hover { background: #90A4AE; }
    ::selection { background: #B8C5D6; color: #1E293B; }
    [data-testid="stSidebar"] { background: linear-gradient(180deg, #E2E8F0 0%, #CBD5E1 100%) !important; border-right: 1px solid #94A3B8; }
    [data-testid="stSidebar"] .stTextInput label, [data-testid="stSidebar"] .stMarkdown { color: #334155 !important; font-weight: 500; }
    h1 { color: #111827 !important; font-weight: 600 !important; letter-spacing: 2px; border-bottom: none !important; padding-bottom: 10px; background: transparent !important; }
    h3 { color: #1E293B !important; font-size: 17px !important; font-weight: 600 !important; letter-spacing: 1px; background: rgba(255, 255, 255, 0.45) !important; backdrop-filter: blur(10px); -webkit-backdrop-filter: blur(10px); border: 1px solid rgba(255, 255, 255, 0.8); border-radius: 8px !important; padding: 8px 18px !important; box-shadow: 0 4px 12px rgba(0, 0, 0, 0.03), inset 0 1px 0 rgba(255, 255, 255, 0.5) !important; width: fit-content; margin-bottom: 18px !important; margin-top: 10px !important; }
    .stButton > button { font-size: 14px !important; font-weight: 600 !important; letter-spacing: 1.5px; background-color: #94A3B8; color: #FFFFFF; border: 1px solid #64748B; border-radius: 4px; padding: 8px 24px; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1), 0 2px 4px -1px rgba(0,0,0,0.06); transition: all 0.2s ease; }
    .stButton > button:hover { background-color: #64748B; border-color: #475569; color: #F8FAFC; transform: translateY(-1px); box-shadow: 0 10px 15px -3px rgba(0,0,0,0.1); }
    .stButton > button:active { transform: translateY(1px); box-shadow: none; }
    .stButton > button:disabled { background-color: #CBD5E1 !important; color: #94A3B8 !important; border-color: #CBD5E1 !important; box-shadow: none !important; cursor: not-allowed; }
    .stTabs [data-baseweb="tab-list"] { gap: 4px; background-color: transparent; border-bottom: 2px solid #CBD5E1; }
    .stTabs [data-baseweb="tab"] { font-size: 14px; letter-spacing: 0.5px; background-color: #EDF2F7; color: #64748B; border-radius: 6px 6px 0 0; border: 1px solid #CBD5E1; border-bottom: none; padding: 10px 22px; transition: all 0.15s ease; }
    .stTabs [data-baseweb="tab"]:hover { background-color: #E2E8F0; color: #475569; }
    .stTabs [aria-selected="true"] { background-color: #94A3B8 !important; color: #FFFFFF !important; font-weight: 600 !important; border-color: #64748B !important; }
    .stTextInput > div > div > input, .stNumberInput > div > div > input { background-color: #FFFFFF !important; border: 1px solid #CBD5E1 !important; border-radius: 4px !important; color: #1E293B !important; font-weight: 500; transition: border-color 0.2s ease; }
    .stTextInput > div > div > input:focus, .stNumberInput > div > div > input:focus { border-color: #64748B !important; box-shadow: 0 0 0 2px rgba(100,116,139,0.15) !important; }
    .stMultiSelect [data-baseweb="tag"] { background-color: #CBD5E1 !important; color: #1E293B !important; border-radius: 4px !important; font-size: 13px !important; }
    [data-testid="stFileUploader"] section { border: 2px dashed #B0BEC5 !important; border-radius: 8px !important; background-color: #FAFBFC !important; transition: border-color 0.2s ease; }
    [data-testid="stFileUploader"] section:hover { border-color: #78909C !important; background-color: #F5F7FA !important; }
    hr { border: none !important; height: 1px !important; background: linear-gradient(to right, transparent, #B0BEC5, transparent) !important; margin: 20px 0 !important; }
    .info-card { background: #FFFFFF; border: 1px solid #E2E8F0; border-left: 4px solid #94A3B8; border-radius: 6px; padding: 20px 24px; margin: 12px 0; box-shadow: 0 1px 3px rgba(0,0,0,0.04); }
    .info-card h4 { color: #475569; margin: 0 0 8px 0; font-weight: 600; font-size: 15px; }
    .info-card p, .info-card li { color: #64748B; font-size: 14px; line-height: 1.8; }
    .info-card.pink { border-left-color: #D4A5A5; }
    .info-card.green { border-left-color: #A5C4B5; }
    .info-card.blue { border-left-color: #94A3B8; }
    .sidebar-brand { text-align: center; padding: 8px 0 16px 0; border-bottom: 1px solid #B0BEC5; margin-bottom: 20px; }
    .sidebar-brand .logo-text { font-size: 20px; font-weight: 700; color: #475569; letter-spacing: 3px; }
    .sidebar-brand .logo-sub { font-size: 11px; color: #94A3B8; letter-spacing: 4px; margin-top: 4px; }
    .placeholder-section { text-align: center; padding: 40px 20px; color: #94A3B8; }
    .placeholder-section .big-icon { font-size: 48px; margin-bottom: 16px; opacity: 0.6; }
    .placeholder-section p { font-size: 14px; line-height: 2; color: #64748B; }
</style>
""", unsafe_allow_html=True)

# ==========================================
# 0.2 登录认证 + Step0~Step1（财险适配版）
# ==========================================

# ---------- 1. 初始化系统状态变量 ----------
for k, v in {
    'logged_in': False,
    'user_role': None,
    'api_key': "",
    'base_url': "https://api.deepseek.com",
    'model_name': "deepseek-chat"
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ---------- 2. 独立登录界面 ----------
if not st.session_state['logged_in']:
    
    st.markdown("""<style>
    header {visibility: hidden;}
    .stApp {background: linear-gradient(135deg, #040B16 0%, #0A1931 50%, #002266 100%); color: #E2E8F0;}
    [data-testid="column"]:nth-of-type(2) {background: rgba(255,255,255,0.08); backdrop-filter: blur(25px); -webkit-backdrop-filter: blur(25px); border-radius: 20px; border: 1px solid rgba(0,243,255,0.3); box-shadow: 0 15px 35px rgba(0,0,0,0.5), inset 0 0 15px rgba(0,243,255,0.15); padding: 40px 30px; margin-top: 6vh;}
    .title-glow {background: linear-gradient(90deg, #FFFFFF 0%, #76D2FF 50%, #00F3FF 100%); -webkit-background-clip: text; -webkit-text-fill-color: transparent; font-weight: 900; filter: drop-shadow(0 0 15px rgba(0,243,255,0.6));}
    div[data-baseweb="input"]>div, div[data-baseweb="select"]>div {background: rgba(0,0,0,0.4)!important; border: 1px solid rgba(0,243,255,0.3)!important; border-radius: 8px!important;}
    div[data-baseweb="input"]>div:focus-within, div[data-baseweb="select"]>div:focus-within {border-color: #00F3FF!important; box-shadow: 0 0 12px rgba(0,243,255,0.5)!important;}
    div[data-testid="stRadio"] label p {color: #FFFFFF!important; font-weight: bold!important; font-size: 14px!important;}
    label p, .stSelectbox label p, .stTextInput label p {color: #00F3FF!important; letter-spacing: 1px!important;}
    input, .stSelectbox span {color: #FFFFFF!important; font-size: 14px!important;}
    button[kind="primary"] {background: linear-gradient(90deg, #0044CC, #0088FF)!important; border: 1px solid rgba(0,243,255,0.5)!important; box-shadow: 0 0 15px rgba(0,136,255,0.4)!important; color: white!important; font-weight: bold!important; letter-spacing: 2px!important; border-radius: 8px!important; margin-top: 5px!important;}
    button[kind="primary"]:hover {box-shadow: 0 0 25px rgba(0,243,255,0.8)!important; transform: scale(1.02);}
    [data-testid="stPopover"] {display: flex; justify-content: flex-end;}
    [data-testid="stPopover"] button {background: transparent!important; border: none!important; box-shadow: none!important; color: #94A3B8!important; font-size: 12px!important; font-weight: normal!important; padding: 0!important; min-height: 0!important; text-decoration: underline; margin-top: 8px; margin-bottom: 15px;}
    [data-testid="stPopover"] button:hover {color: #00F3FF!important;}
    </style>""", unsafe_allow_html=True)

    _, col2, _ = st.columns([1, 1.5, 1])
    with col2:
        st.markdown("<div style='text-align:center; margin-bottom:30px;'><h1 style='font-size:32px; margin:0;'><span class='title-glow'>财险数智年报平台</span></h1><p style='color:#76D2FF; font-size:11px; letter-spacing:4px; margin-top:8px; font-weight:bold;'>P&C INTELLIGENCE</p></div>", unsafe_allow_html=True)
        
        u_type = st.radio("访问权限", ["普通用户", "项目组成员"], label_visibility="collapsed", horizontal=True)
        sec_code = st.text_input("安全验证", type="password", placeholder="请输入内部安全码")
        
        ai_map = {
            "阿里云百炼 (通义千问)": ("https://dashscope.aliyuncs.com/compatible-mode/v1", "qwen-plus"),
            "DeepSeek (深度求索)": ("https://api.deepseek.com", "deepseek-v4-flash"),
            "月之暗面 (Kimi)": ("https://api.moonshot.cn/v1", "moonshot-v1-8k"),
            "智谱AI (GLM-4)": ("https://open.bigmodel.cn/api/paas/v4", "glm-4"),
            "OpenAI (ChatGPT)": ("https://api.openai.com/v1", "gpt-4o")
        }
        ai_pr = st.selectbox("选择您将使用的AI", list(ai_map.keys()) + ["自定义私有化节点"])
        
        d_url, d_mod = ai_map.get(ai_pr, ("https://api.deepseek.com", "deepseek-chat"))
        if ai_pr == "自定义私有化节点":
            d_url, d_mod = st.text_input("RPC 接口地址", "https://api.deepseek.com"), st.text_input("指定模型版本", "deepseek-chat")

        api_input = st.text_input("请填写您使用AI的API Key", type="password", placeholder=" sk-... (不填则仅开放年报检测、数据合并及可视化本地分析功能)")
        
        if "OpenAI" in ai_pr:
            st.markdown("<p style='font-size:11px; color:#F87171; text-align:right; margin-top:4px; margin-bottom:0;'>⚠️ 严禁向境外节点传输涉密数据</p>", unsafe_allow_html=True)
        
        with st.popover("如何获取API key?"):
            st.markdown("**1.** 前往各大模型官方开放平台注册账号。\n\n**2.** 在控制台生成 `sk-` 开头的密钥。\n\n**3.** 复制填入上方输入框即可体验 AI 功能。")

        if st.button("启 动 系 统", type="primary", use_container_width=True):
            if u_type == "普通用户" and sec_code != "KPMG1234":
                st.error("❌ 拒绝访问：普通用户安全码错误")
            elif u_type == "项目组成员" and sec_code != "KPMG666":
                st.error("❌ 拒绝访问：项目组成员安全码错误")
            else:
                st.session_state.update({
                    'logged_in': True,
                    'user_role': u_type,
                    'api_key': api_input,
                    'base_url': d_url,
                    'model_name': d_mod
                })
                st.rerun()

        st.markdown("<div style='text-align:center; color:#94A3B8; font-size:11px; margin-top:30px; letter-spacing:1px;'>系统版本：v3.0 (Alpha) © 2026<br>Developed by 曾萍Polly@KPMG</div>", unsafe_allow_html=True)

    st.stop()

# ==========================================
# 🆕 在这里添加全局变量初始化（必须放在所有 def 函数之前）
# ==========================================
notes_dict = {}
ordered_modules = []

# ---------- 3. 导入 PDF 处理依赖 ----------
import fitz
import pdfplumber
from openai import OpenAI

# ---------- 4. 辅助函数：单位嗅探 ----------
def get_report_unit(text):
    """从文本中嗅探金额单位"""
    if not text:
        return None
    m1 = re.search(r'(?:金额)?单位[：:\s]*((?:人民币)?[元十百千万亿]+)', text)
    if m1:
        return m1.group(1).strip()
    m2 = re.search(r'(人民币[十百千万亿]+元)', text)
    if m2:
        return m2.group(1).strip()
    if "人民币元" in text:
        return "人民币元"
    return None
def normalize_field(s):
    if not isinstance(s, str):
        return ""
    # 去除常见空白字符
    s = s.replace('\xa0', '').replace('\u3000', '').replace(' ', '').strip()
    # 全角括号转半角
    s = s.replace('（', '(').replace('）', ')')
    return s
# ---------- 5. PDF 视觉提取引擎 ----------
def extract_single_page_vision(pdf_bytes, page_num, expected_name, api_key, base_url, model_name):
    """将 PDF 页面转为高清图片，使用视觉大模型提取表格"""
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        if page_num < 1 or page_num > len(doc):
            return None, ""
        page = doc.load_page(page_num - 1)
        
        zoom_matrix = fitz.Matrix(3.0, 3.0)
        pix = page.get_pixmap(matrix=zoom_matrix, alpha=False)
        img_data = pix.tobytes("png")
        base64_image = base64.b64encode(img_data).decode('utf-8')
        doc.close()

        prompt = f"""你是一个四大会计师事务所的资深财险数字化审计专家。
我为你提供了一张【保险公司年报的单页高清截图】，目标是精准提取：【{expected_name}】。
表格中的【数字是由图片构成的】，请发挥强大的视觉识别能力，提取所有文字和表格。
【强制要求】：
1. 所有的科目名、数字金额之间，必须被 "|" 隔开。严禁使用连续空格。
2. 精准对齐多级表头！如果有留白单元格，必须用 "|" 占位补齐！
3. 绝不能漏掉任何一行数据，保留金额里的逗号和括号。
4. 纯文本输出，不要使用 Markdown 代码块。不需要输出 SHEET_NAME 标签，直接排版。"""

        client = OpenAI(api_key=api_key, base_url=base_url)
        response = client.chat.completions.create(
            model=model_name,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{base64_image}"}}
                    ]
                }
            ],
            temperature=0.0
        )
        
        result_text = response.choices[0].message.content.strip()
        result_text = re.sub(r'^```(csv|text)?\n?', '', result_text, flags=re.MULTILINE).replace("```", "").strip()
        
        lines = result_text.split('\n')
        parsed_data = []
        max_cols = 0
        for row in lines:
            clean_row = row.strip()
            if not clean_row:
                continue
            if re.match(r'^[\s\|-]+$', clean_row) and '-' in clean_row:
                continue
            if clean_row.startswith('|'):
                clean_row = clean_row[1:]
            if clean_row.endswith('|'):
                clean_row = clean_row[:-1]
            cols = [col.strip() for col in clean_row.split('|')] if '|' in clean_row else [clean_row]
            parsed_data.append(cols)
            max_cols = max(max_cols, len(cols))
            
        for row in parsed_data:
            if len(row) < max_cols:
                row.extend([''] * (max_cols - len(row)))
                
        return pd.DataFrame(parsed_data), "【提示：该页已使用 OCR 视觉引擎处理，原文本为图片结构】"
        
    except Exception as e:
        return None, f"图像处理失败: {str(e)}"

# ---------- 6. PDF 文本提取引擎 ----------
def extract_single_page(pdf_bytes, page_num, expected_name, api_key, base_url, model_name):
    """使用 pdfplumber 提取文本，并用 LLM 转换为结构化数据"""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            if page_num < 1 or page_num > len(pdf.pages):
                return None, ""
            page = pdf.pages[page_num - 1]
            raw_text = page.extract_text(layout=True)
            
        if not raw_text:
            return None, ""

        prompt = f"""你是一个四大会计师事务所的资深财险数字化审计专家。
当前任务：提取【{expected_name}】。
请将以下纯文本中的【段落文字】原样保留为一行，将【财务表格】转化为用 "|" 严格分隔的标准网格格式。
【强制要求】：
1. 所有的科目名、附注编号、数字金额之间，必须被 "|" 隔开。严禁使用连续空格。
2. 精准对齐多级表头！如果某个单元格是空的，必须用 "|" 占位补齐！
3. 绝不能漏掉任何一行数据，保留金额里的逗号和括号。
4. 纯文本输出，不要使用 Markdown 代码块。不需要输出 SHEET_NAME，直接输出转化后的数据。

以下是原始文本，请开始输出：
{raw_text}"""

        client = OpenAI(api_key=api_key, base_url=base_url)
        response = client.chat.completions.create(
            model=model_name,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0
        )
        
        result_text = response.choices[0].message.content.strip()
        result_text = re.sub(r'^```(csv|text)?\n?', '', result_text, flags=re.MULTILINE)
        result_text = result_text.replace("```", "").strip()
        
        lines = result_text.split('\n')
        parsed_data = []
        max_cols = 0
        for row in lines:
            clean_row = row.strip()
            if not clean_row:
                continue
            if re.match(r'^[\s\|-]+$', clean_row) and '-' in clean_row:
                continue
            if clean_row.startswith('|'):
                clean_row = clean_row[1:]
            if clean_row.endswith('|'):
                clean_row = clean_row[:-1]
            if '|' not in clean_row:
                cols = [clean_row]
            else:
                cols = [col.strip() for col in clean_row.split('|')]
            parsed_data.append(cols)
            max_cols = max(max_cols, len(cols))
            
        for row in parsed_data:
            if len(row) < max_cols:
                row.extend([''] * (max_cols - len(row)))
            
        return pd.DataFrame(parsed_data), raw_text
        
    except Exception as e:
        return None, f"提取失败: {str(e)}"

# ---------- 7. AI 页码定位引擎（财险版） ----------
def ai_find_pages(pdf_bytes, api_key, target_tables, base_url, model_name, company_name=""):
    """混合双引擎：AI负责读目录推算主表，Python雷达负责深潜寻找附注表（财险版）"""
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        
        # 1. 提取前 60 页（用于AI目录识别）
        toc_text = ""
        for i in range(min(60, len(doc))):
            page = doc.load_page(i)
            page_text = page.get_text("text").replace("\n", " ")
            page_text = " ".join(page_text.split())
            toc_text += f"---第{i+1}页---\n{page_text[:1200]}\n"
            
        # ==========================================
        # 财险专属附注雷达特征矩阵（其他表保持不变）
        # ==========================================
        feature_matrix = {
            "保险业务收入表 (附注)": [
                ["保险业务收入", "已赚保费", "签单保费"],
                ["原保险", "再保险", "分入"],
                ["分保费收入", "分保费用", "保费收入"]
            ],
            "综合成本率拆解表 (附注)": [
                ["综合成本率", "COR", "成本率"],
                ["赔付率", "费用率", "综合赔付率", "综合费用率"],
                ["已发生赔款", "手续费及佣金", "业务及管理费"]
            ],
            "准备金评估表 (附注)": [
                ["未到期责任准备金", "未决赔款准备金"],
                ["IBNR", "已发生未报告"],
                ["准备金", "负债", "准备金余额"]
            ],
            "费用结构表 (附注)": [
                ["业务及管理费", "手续费及佣金"],
                ["职工薪酬", "折旧", "租赁", "办公费", "宣传费"],
                ["获取费用", "维持费用"]
            ],
            "投资收益表 (附注)": [
                ["投资收益", "投资资产", "净投资回报"],
                ["公允价值变动", "利息收入", "股息收入"],
                ["债券", "股票", "基金", "长期股权投资"]
            ],
            "偿付能力表 (附注)": [
                ["偿付能力", "综合偿付能力充足率", "核心偿付能力充足率"],
                ["实际资本", "最低资本"],
                ["风险", "资本"]
            ],
            "业务及管理费 (附注)": [
                ["业务及管理费", "管理费用"],
                ["职工薪酬", "折旧", "租赁", "办公费", "宣传费"]
            ],
            "主要经营指标": [
                ["偿付能力充足率", "核心偿付能力", "综合偿付能力"],
                ["实际资本", "最低资本", "风险资本"],
                ["保险业务收入", "净利润", "净资产"]
            ]
        }
        # ==========================================
        # 保险合同负债及资产 专属识别规则
        # ==========================================
        insurance_rules = {
            "平安": {
                "title_words": ["保险合同负债/资产", "保险合同负债／资产", "保险合同负债及资产"],
                "title_require": [],
                "include": ["保险合同负债", "保险合同资产", "合同服务边际"],
                "exclude": ["再保险合同", "持有的再保险", "分出的再保险"],
                "require_any": []
            },
            "阳光": {
                "title_words": ["保险合同负债及资产"],
                "title_require": [],
                "include": ["保险合同负债", "保险合同资产", "合同服务边际"],
                "exclude": ["再保险合同", "持有的再保险"],
                "require_any": []
            },
            "众安": {
                "title_words": ["保险合同资产和负债", "保险合同负债和资产", "保险合同资产及负债", "保险合同负债及资产"],
                "title_require": [],
                "include": ["签发的保险合同", "保费分配法"],
                "exclude": ["持有的再保险合同", "再保险合同"],
                "require_any": ["保费分配法"]
            },
            "人保": {
                "title_words": ["保险合同"],
                "title_require": [
                    "本公司签发的保险合同采用保费分配法计量",
                    "本集团和本公司签发的保险合同采用通用模型计量"
                ],
                "include": [
                    "本公司签发的保险合同采用保费分配法计量",
                    "本集团和本公司签发的保险合同采用通用模型计量",
                    "通用模型计量"
                ],
                "exclude": [],
                "require_any": []
            },
            "太保": {
                "title_words": ["保险合同负债/资产", "保险合同负债／资产", "保险合同负债及资产"],
                "title_require": [],
                "include": ["保险合同负债", "保险合同资产", "合同服务边际"],
                "exclude": ["再保险合同", "持有的再保险"],
                "require_any": []
            },
            "太平": {
                "title_words": ["保险合同资产及负债", "保险合同负债及资产"],
                "title_require": [],
                "include": ["保险合同资产及负债", "合同服务边际", "保险合同负债"],
                "exclude": ["再保险合同", "持有的再保险"],
                "require_any": []
            }
        }

        # 未知/自动识别公司时的兜底规则
        default_insurance_rule = {
            "title_words": [
                "保险合同负债及资产", "保险合同资产及负债", "保险合同负债/资产",
                "保险合同负债／资产", "保险合同资产和负债", "保险合同负债和资产"
            ],
            "title_require": [],
            "include": ["保险合同负债", "保险合同资产", "合同服务边际"],
            "exclude": ["再保险合同", "持有的再保险"],
            "require_any": []
        }
        
        simple_title_OR = {
            "利润表 (附注)": ["利润表", "损益表", "利润总额"],
            "资产负债表 (附注)": ["资产负债表", "资产", "负债", "所有者权益"],
            "现金流量表 (附注)": ["现金流量表", "现金流", "经营活动"],
            "股东/所有者权益变动表 (附注)": ["股东权益变动表", "权益变动", "所有者权益变动表"],
            "公司利润表": ["利润表", "损益表", "利润总额", "公司利润表"],
            "公司资产负债表": ["资产负债表", "公司资产负债表"],
            "业务及管理费": ["业务及管理费", "管理费用"],
            "主要经营指标": ["主要经营指标", "经营指标", "偿付能力指标"]
        }
        
        context_anchors = ["项目注释", "财务报表附注", "报表附注", "项目附注"]
        
        found_hints = {table: [] for table in target_tables}   # 为所有目标表都初始化空列表
        # ====== 新增：初始化 page_type_map ======
        page_type_map = {}

        # ==========================================================
        # 从第0页开始扫描
        # ==========================================================
        for i in range(0, len(doc)):
            page = doc.load_page(i)
            text_raw = page.get_text("text")
            text_clean = text_raw.replace("\n", "").replace(" ", "")
            header_text = text_clean[:800]
            
            # 检测表格
            try:
                tables = page.find_tables()
                has_actual_table = len(tables.tables) > 0 if tables else False
            except:
                has_actual_table = False
            is_continued_table = "(续)" in text_clean or "（续）" in text_clean
            is_table_page = is_continued_table or has_actual_table
            
            is_notes_page = any(anchor in text_clean for anchor in context_anchors)
            
            # ---------- 1. 匹配 feature_matrix（其他表） ----------
            if is_notes_page and is_table_page:
                for table, condition_groups in feature_matrix.items():
                    if table in target_tables:
                        matched = all(any(kw in text_clean for kw in group) for group in condition_groups)
                        if matched:
                            page_num = i + 1
                            if page_num not in found_hints[table]:
                                found_hints[table].append(page_num)
            
            # ---------- 2. 匹配 simple_title_OR（其他表） ----------
            for table, any_of_words in simple_title_OR.items():
                if table in target_tables:
                    if any(kw in header_text for kw in any_of_words):
                        page_num = i + 1
                        if "合并" in header_text:
                            page_type_map[page_num] = 'group'
                        else:
                            page_type_map[page_num] = 'company'
                        if page_num not in found_hints[table]:
                            found_hints[table].append(page_num)
            
            # ---------- 3. 特殊处理“保险合同负债及资产” ----------
            
            if "保险合同负债及资产" in target_tables:
                # 3.0 公司归一化匹配（支持"平安产险"/"太保财险"等全称命中短键）
                company_key = ""
                if company_name:
                    for _key in insurance_rules:
                        if _key in company_name:
                            company_key = _key
                            break
                rule = insurance_rules.get(company_key, default_insurance_rule)

                # 3.1 数据页判断（动态年份，支持2023等历史报告；续表页豁免年份要求）
                if not is_continued_table and not re.search(r'20\d{2}', text_clean):
                    continue

                # 3.2 排除资产负债表误命中（需"资产负债表"标题 或 ≥2个强信号；
                #     注意："负债合计"是"保险合同负债合计"的子串，单独出现不能判为资产负债表）
                if "资产负债表" in text_clean:
                    continue
                bs_strong = ["资产总计", "所有者权益合计", "流动资产", "非流动资产", "流动负债", "非流动负债"]
                if sum(1 for w in bs_strong if w in text_clean) >= 2:
                    continue

                # 3.3 表名称匹配（按公司差异化表名）
                score = 0   
                title_words = rule.get("title_words", default_insurance_rule["title_words"])
                title_match = any(word in text_clean for word in title_words)
                # 表名较泛的公司（如人保仅"保险合同"），额外要求命中具体子表头
                title_require = rule.get("title_require", [])
                if title_match and title_require:
                    if not any(word in text_clean for word in title_require):
                        title_match = False
                if not title_match:
                    continue
                score += 10

                # 3.4 硬性必含字段（如众安仅取"保费分配法"部分）
                require_any = rule.get("require_any", [])
                if require_any:
                    if not any(word in text_clean for word in require_any):
                        continue

                # 3.5 IFRS17核心字段加分（新增"通用模型"）
                indicators = ["合同服务边际","未来现金流量","非金融风险调整","履约现金流",
                              "亏损部分","非亏损部分","未到期责任负债","已发生赔款负债",
                              "保费分配法","通用模型"]
                for word in indicators:
                    if word in text_clean:
                        score += 1

                # 3.6 公司差异规则（用归一化后的 rule，不再用 company_name 直接取）
                for word in rule.get("include", []):
                    if word in text_clean:
                        score += 3
                for word in rule.get("exclude", []):
                    if word in text_clean:
                        score -= 8

                # 3.7 续表识别 / 3.8 最终判断（不变）
                if "(续)" in text_clean or "（续）" in text_clean:
                    score += 3
                if score >= 10:
                    page_num = i + 1
                    if page_num not in found_hints["保险合同负债及资产"]:
                        found_hints["保险合同负债及资产"].append(page_num)

            # ---------- 4. 特殊处理“保险产品经营信息” ----------
            if "保险产品经营信息" in target_tables:
                # 使用更精准的标题关键词（只保留最独特的标识）
                if any(kw in header_text for kw in ["保险产品经营信息", "原保险保费收入居前5位的险种"]):
                    # 必须包含实际表格（find_tables 检测到表格，或为续表）
                    if has_actual_table or is_continued_table:
                        # 排除利润表标题（避免误匹配）
                        if not any(kw in header_text for kw in ["利润表", "损益表", "利润总额", "公司利润表"]):
                            page_num = i + 1
                            if page_num not in found_hints.get("保险产品经营信息", []):
                                found_hints.setdefault("保险产品经营信息", []).append(page_num)

            # ---------- 5. 文本类披露：折现率、非金融风险调整 ----------
            if "折现率披露" in target_tables:
                if "折现率" in text_clean or "折现率假设" in text_clean:
                    page_num = i + 1
                    if page_num not in found_hints["折现率披露"]:
                        found_hints["折现率披露"].append(page_num)
            
            if "非金融风险调整披露" in target_tables:
                if "非金融风险调整" in text_clean:
                    page_num = i + 1
                    if page_num not in found_hints["非金融风险调整披露"]:
                        found_hints["非金融风险调整披露"].append(page_num)

        # ==========================================================
        # 公司口径过滤
        # ==========================================================
        for table in ["公司利润表", "公司资产负债表"]:
            if table in found_hints and found_hints[table]:
                company_pages = [p for p in found_hints[table] if page_type_map.get(p) == 'company']
                # 如果没有公司口径，则保留所有（包括合并）
                found_hints[table] = company_pages if company_pages else found_hints[table]
                if company_pages:
                    found_hints[table] = company_pages

        doc.close()
        
        # 构建提示文本（供AI参考）
        hint_text = "\n\n【⚡ Python程序附注雷达线索】\n"
        for table, pages in found_hints.items():
            if pages:
                pages = sorted(list(set(pages)))
                hint_text += f"- {table} 真实表格物理页码位于: {pages[:4]}\n"

        # AI 目录驱动提示（保持原样）
        prompt = f"""你是一个顶级的财险审计专家，任务是定位 PDF 中的核心财务报表页码。

【关键定位逻辑（双模式自适应）】
模式 A：目录驱动（适用于有目录的报表）
- 如果前几页存在"目录"或"Contents"，请识别目标报表对应的页码。
- 计算物理偏移量（封面/说明页占用的页数），得出真实的物理页码。

模式 B：标题驱动（适用于无目录或目录不全的报表）
- 如果没有明确目录，请直接扫描文本中 `---第N页---` 标记下方的页眉或正文大标题。
- 例如：若在 `---第8页---` 下方看到"合并利润表"，则物理页码即为 8。

【通用执行准则】
1. 🎯 如果需求表单名称包含"（合并）"，优先寻找：
   合并资产负债表
   合并利润表
   合并现金流量表
   合并股东权益变动表

2. 🎯 如果需求表单名称包含"（公司）"，优先寻找：
   公司资产负债表
   公司利润表
   公司现金流量表
   公司股东权益变动表

3. 🎯 如果年报中没有区分"合并"和"公司"，仅出现：
   资产负债表
   利润表
   现金流量表
   股东权益变动表
   则同一页码同时返回给对应的（合并）和（公司）

4. 🎯 跨页逻辑：由于财务报表通常包含"(续)"，请务必返回所有相关的物理页码数组。

5. 🎯 附注表线索（最高优先级）：对于带有"(附注)"字样的复杂底表，请【无脑完全信任】下方提供的【Python程序附注雷达线索】，直接返回线索中的页码。

需求表单列表：
{json.dumps(target_tables, ensure_ascii=False)}

【待扫描文本内容（前 60 页）】：
{toc_text}

【⚡ Python程序附注雷达辅助线索】：
{hint_text}

【强制输出格式】
只输出合法 JSON，键为表单名，值为物理页码数组。找不到填 [0]。
格式示例：{{"合并利润表": [8, 9], "资产负债表 (合并优先)": [2, 3, 4]}}"""

        client = OpenAI(api_key=api_key, base_url=base_url)
        response = client.chat.completions.create(
            model=model_name,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1
        )
        
        result_text = response.choices[0].message.content.strip()
        result_text = result_text.replace("```json", "").replace("```", "").strip()
        return json.loads(result_text)
        
    except Exception as e:
        return {"error": str(e)}
# ==========================================
# 以下为原有财险对标报告系统（第1~10段）—— 所有函数定义移至主逻辑之前
# ==========================================
# 0.添加公司边框和标题
def add_company_borders(fig, companies, x_labels, top_margin=60, row=None, col=None):
    """
    为单图或子图添加公司边框和标题。
    如果 row 和 col 不为空，则视为子图模式，直接覆盖整个子图区域。
    """
    # ===== 子图模式 =====
    if row is not None and col is not None:
        if len(companies) == 1:
            co = companies[0]
            # 使用 x domain / y domain 限定在子图内部
            fig.add_shape(
                type="rect",
                xref="x domain", yref="y domain",
                x0=0, x1=1,
                y0=0, y1=1.0,          # 覆盖整个子图区域（从底部到顶部）
                fillcolor="rgba(200,200,200,0.05)",
                line=dict(color="#CCCCCC", width=1.2),
                layer="below",
                row=row, col=col
            )
            # 公司名称放在子图顶部内部（y=0.98）
            fig.add_annotation(
                x=0.5,
                y=0.95,
                text=f"<b>{co}</b>",
                showarrow=False,
                font=dict(size=12, color="#888888"),
                xanchor="center",
                yanchor="bottom",
                xref="x domain",
                yref="y domain",
                row=row, col=col
            )
        return fig
        
    # ===== 单图模式（原有逻辑） =====
    company_ranges = {}
    for co in companies:
        indices = [i for i, label in enumerate(x_labels) if co in label]
        if indices:
            company_ranges[co] = (min(indices), max(indices))
    if not company_ranges:
        for co in companies:
            if co in x_labels:
                idx = x_labels.index(co)
                company_ranges[co] = (idx, idx)
    if not company_ranges:
        return fig

    for co, (start, end) in company_ranges.items():
        width_extend = 0.55 if (end - start) > 0 else 0.45
        x0 = start - width_extend
        x1 = end + width_extend
        fig.add_shape(
            type="rect",
            xref="x", yref="paper",
            x0=x0, x1=x1,
            y0=0, y1=1.03,
            fillcolor="rgba(200,200,200,0.05)",
            line=dict(color="#CCCCCC", width=1.2),
            layer="below"
        )
        fig.add_annotation(
            x=(start + end) / 2,
            y=0.985,
            text=f"<b>{co}</b>",
            showarrow=False,
            font=dict(size=12, color="#888888"),
            xanchor="center",
            yanchor="bottom",
            xref="x",
            yref="paper"
        )
    fig.update_layout(margin=dict(t=top_margin))
    return fig
    
# 1.全局颜色工具
def get_color_map(all_cos):
    current_selection_key = tuple(sorted(all_cos))
    if 'company_color_map' not in st.session_state or st.session_state.get('_last_color_selection') != current_selection_key:
        PRESET_COLORS = ["#C00000", "#0865EE", "#FEAED7", "#92D050", "#7030A0", "#EF9867", "#61CBF4", "#C7A0F7"]
        st.session_state['company_color_map'] = {co: PRESET_COLORS[i % len(PRESET_COLORS)] for i, co in enumerate(all_cos)}
        st.session_state['_last_color_selection'] = current_selection_key
    return st.session_state['company_color_map']

# 2.辅助显示函数（AI引擎、笔记、图表渲染等）
@st.cache_data(show_spinner=False)
def _call_llm_api_cached(prompt, field_name, latest_year, unit_str, user_api_key, api_base, api_model):
    try:
        client = OpenAI(api_key=user_api_key, base_url=api_base)
        res = client.chat.completions.create(
            model=api_model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3
        )
        return f"<b><span style='color:#00338D'></span></b> {res.choices[0].message.content}"
    except Exception as e:
        return f"<span style='color:#C00000; font-size:12px;'>⚠️ AI报错: {str(e)}</span>"

def generate_ai_insight(df, field_name, is_pct=False, template=""):
    """
    生成 AI 分析文本
    - template: 参考话术模板（来自“分析内容-默认”列）
    """
    enable_ai = st.session_state.get('enable_ai', False)
    latest_year = st.session_state.get('latest_year', 2025)
    selected_cos = st.session_state.get('selected_cos_cache', [])
    divisor = st.session_state.get('divisor', 1)
    unit_label = st.session_state.get('unit_label', '百万元')

    if not enable_ai or df is None or df.empty or not field_name:
        return ""

    try:
        col_field = '字段名' if '字段名' in df.columns else '指标名称'
        # 提取最新年份数据
        d_sub = df[(df[col_field].astype(str).str.contains(field_name, na=False)) &
                   (df['报告年份'].astype(str) == str(latest_year)) &
                   (df['公司'].isin(selected_cos))].copy()
        if d_sub.empty:
            return ""

        val_col = "(百万)人民币" if "(百万)人民币" in d_sub.columns else d_sub.columns[-1]
        # 上年数据
        prev_year = st.session_state.get('prev_year', 2024)
        d_prev = df[(df[col_field].astype(str).str.contains(field_name, na=False)) &
                    (df['报告年份'].astype(str) == str(prev_year)) &
                    (df['公司'].isin(selected_cos))].copy()

        # ---- 构建数据事实 ----
        data_dict = {}
        for _, r in d_sub.iterrows():
            co = r['公司']
            val = r[val_col]
            data_dict[co] = val

        # 格式化数据
        data_summary = []
        for co, val in data_dict.items():
            if is_pct:
                data_summary.append(f"{co}: {val:.1%}")
            else:
                data_summary.append(f"{co}: {val/divisor:.1f}{unit_label}")

        # 排名和变化
        sorted_items = sorted(data_dict.items(), key=lambda x: x[1], reverse=True)
        top_co, top_val = sorted_items[0]
        bottom_co, bottom_val = sorted_items[-1]

        changes = []
        for co in selected_cos:
            curr_val = data_dict.get(co)
            prev_vals = d_prev[d_prev['公司'] == co][val_col]
            prev_val = prev_vals.iloc[0] if not prev_vals.empty else None
            if curr_val is not None and prev_val is not None and prev_val != 0:
                pct_change = (curr_val - prev_val) / abs(prev_val)
                changes.append((co, pct_change))
        changes.sort(key=lambda x: x[1], reverse=True)
        top_gain = changes[0] if changes else None
        top_loss = changes[-1] if changes else None

        # 构建事实文本
        facts = f"最新年（{latest_year}）数据：{', '.join(data_summary)}。"
        facts += f"最高为{top_co}（{top_val:.2f}），最低为{bottom_co}（{bottom_val:.2f}）。"
        if top_gain:
            facts += f"同比增幅最大为{top_gain[0]}（{top_gain[1]:.1%}），降幅最大为{top_loss[0]}（{top_loss[1]:.1%}）。"

        # ---- 构建提示词 ----
        if template and template.strip():
            prompt = f"""
你是一位精算分析专家。请根据以下数据事实，参考给定的“话术模板”风格，生成一段简洁通顺的分析文字（不超过80字）。
数据事实：{facts}
话术模板：{template}
要求：仿照模板的表述方式，但必须基于数据事实，不要出现模板中的固定公司名或无关内容。
输出直接是分析文字，不要前缀。
"""
        else:
            prompt = f"你是一位精算分析专家。请根据以下数据事实，用简洁通顺的中文总结一段分析（不超过60字）：{facts}"

        # ---- 调用AI ----
        api_key = st.session_state.get('api_key', "").strip()
        if api_key:
            return _call_llm_api_cached(prompt, field_name, latest_year, unit_label, api_key, st.session_state.get('base_url'), st.session_state.get('model_name'))
        else:
            # 无API时返回数据事实
            return f"<b>📊 {facts}</b>"

    except Exception as e:
        return f"<span style='color:#C00000;'>⚠️ AI分析出错: {e}</span>"

def display_notes(module_id, ai_df=None, ai_field=None, is_pct=False):
    global notes_dict
    md = notes_dict.get(module_id, {})
    an_default = md.get('analysis_default', "")   # 话术模板
    an_custom  = md.get('analysis_custom', "")
    nt         = md.get('note', "")
    ai_txt     = generate_ai_insight(ai_df, ai_field, is_pct, template=an_default)

    # 如果启用AI，优先使用AI生成的文本，否则使用预设
    if ai_txt and st.session_state.get('enable_ai', False):
        display_text = ai_txt
    else:
        display_text = an_default
        if an_custom:
            display_text += "<br>" + an_custom if display_text else an_custom
        if ai_txt and not st.session_state.get('enable_ai', False):
            display_text += "<br>" + ai_txt if display_text else ai_txt

    if display_text:
        html = '<div style="text-align:left; background:#F0F4FA; border-left:4px solid #00338D; padding:3px 10px; margin-bottom:6px; border-radius:4px; font-family:Microsoft YaHei, 微软雅黑, sans-serif;">'
        html += f'<p style="margin:2px 0; color:#0A1F5C; font-size:14px; line-height:1.4;">{display_text}</p>'
        st.markdown(html + "</div>", unsafe_allow_html=True)

    return (an_default or an_custom), nt

def display_bottom_note(nt_text):
    if nt_text and str(nt_text).lower() != 'nan':
        st.markdown(f'<div style="text-align: left;margin-top: 1px; margin-bottom: 0px; padding-left: 5px;text-align: left;"><p style="margin: 0; color: #888; font-size: 12px; font-style: italic;">* 注释：{nt_text}</p></div>', unsafe_allow_html=True)

def show_chart(fig, p_mode, m_id=None):
    if not fig:
        return
    if p_mode:
        # 打印模式：自动适应宽度，通过 flex 居中
        fig.update_layout(
            autosize=True,
            height=500,
            margin=dict(t=30, b=30, l=50, r=20)
        )
        st.markdown('<div style="display: flex; justify-content: center; width: 100%;">', unsafe_allow_html=True)
        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        # 非打印模式保持不变
        if m_id in ["claim_ratio_trend", "expense_ratio_trend"]:
            fig.update_layout(
                autosize=False,
                width=900,
                height=550,
                margin=dict(t=50, b=100, l=60, r=40)
            )
            st.plotly_chart(fig, use_container_width=False, config={"displayModeBar": False})
        else:
            fig.update_layout(autosize=True)
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})
            
# 3.文本披露卡片
def display_textual_disclosures(df, cos, cy):
    st.markdown("#### 📄 关键会计政策与精算假设披露")
    for co in cos:
        df_co = df[df['公司'] == co]
        def get_text(y, kw):
            s = df_co[(df_co['报告年份'].astype(str) == str(y)) & 
                      (df_co['字段名'].astype(str).str.contains(kw, na=False))]['值']
            if s.empty:
                return "未披露"
            val = s.iloc[0]
            return str(val) if pd.notna(val) and str(val).strip() != "" else "未披露"
        discount_rate = get_text(cy, '折现率')
        risk_margin = get_text(cy, '风险边际') or get_text(cy, '非金融风险调整')
        policy_method = get_text(cy, '计量方法') or get_text(cy, '会计政策')
        if all(v == "未披露" for v in [discount_rate, risk_margin, policy_method]):
            continue
        with st.expander(f"📌 {co} - 精算假设与政策", expanded=False):
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("计量方法（政策）", policy_method)
            with col2:
                st.metric("折现率假设", discount_rate)
            with col3:
                st.metric("风险边际（RA）", risk_margin)


# 4.综合成本率拆解（多因子分组柱状图）
def create_cor_breakdown_chart(df, cos, year, divisor, highlight_co="无"):
    factors = [
        "当期发生赔款及理赔费用",
        "已发生赔款负债履约现金流变动",
        "亏损合同损益",
        "承保财务损益",
        "再保净成本",
        "提取保费准备金"
    ]
    color_2024 = "#1E49E2"
    color_2025 = "#00338D"
    
    prev_year = st.session_state.get('prev_year', 2024)
    latest_year = st.session_state.get('latest_year', 2025)
    years = [str(prev_year), str(latest_year)]
    
    raw_df = st.session_state.get('integrated_data', None)
    if raw_df is None:
        st.error("❌ 未找到集成数据，请先完成 Step 5 数据集成。")
        return [go.Figure() for _ in factors]
    
    raw_df = df.copy()
    raw_df.columns = raw_df.columns.str.strip()
    raw_df['公司'] = raw_df['公司'].astype(str).str.strip()
    raw_df['字段名'] = raw_df['字段名'].astype(str).str.strip()
    raw_df['报告年份'] = raw_df['报告年份'].astype(str).str.replace('.0', '', regex=False).str.strip()
    raw_df = raw_df[raw_df['报告年份'] != '']
    raw_df = raw_df[~raw_df['报告年份'].str.lower().isin(['nan', 'none'])]
    
    companies = [c.strip() for c in cos]
    
    service_revenue = {}
    for co in companies:
        for yr in years:
            mask = (raw_df['公司'] == co) & (raw_df['报告年份'] == yr) & (raw_df['字段名'] == '保险服务收入')
            rev_series = raw_df.loc[mask, '(百万)人民币']
            rev = rev_series.sum() if not rev_series.empty else 0
            if pd.isna(rev) or rev == 0:
                mask2 = (raw_df['公司'] == co) & (raw_df['报告年份'] == yr) & (raw_df['字段名'] == '保险业务收入')
                rev_series2 = raw_df.loc[mask2, '(百万)人民币']
                rev = rev_series2.sum() if not rev_series2.empty else 0
            if rev == 0:
                st.warning(f"⚠️ 公司 {co} 在 {yr} 年未找到保险服务收入或保险业务收入，分母设为 1")
                rev = 1
            service_revenue[(co, yr)] = rev
    
    data = {}
    for f in factors:
        data[f] = {}
        for co in companies:
            df_co = raw_df[raw_df['公司'] == co]
            for yr in years:
                mask_f = (df_co['报告年份'] == yr) & (df_co['字段名'] == f)
                v_series = df_co.loc[mask_f, '(百万)人民币']
                v = v_series.sum() if not v_series.empty else 0
                denom = service_revenue.get((co, yr), 1)
                ratio = v / denom if denom != 0 else 0
                data[f][(co, yr)] = ratio if pd.notna(ratio) else 0
    
    figures = []
    for f in factors:
        y_2024 = [data[f].get((co, years[0]), 0) for co in companies]
        y_2025 = [data[f].get((co, years[1]), 0) for co in companies]
        all_vals = y_2024 + y_2025
        
        if all_vals:
            max_val = max(all_vals)
            min_val = min(all_vals)
            abs_max = max(abs(max_val), abs(min_val))
            
            if abs_max < 0.001:
                y_range = [-0.001, 0.001]
            else:
                range_min = min_val * 1.2 if min_val < 0 else 0
                range_max = max_val * 1.3 if max_val > 0 else 0
                if range_max - range_min < 0.002:
                    margin = (0.002 - (range_max - range_min)) / 2
                    range_min -= margin
                    range_max += margin
                y_range = [range_min, range_max]
        else:
            y_range = [-0.1, 0.1]
        
        fig = go.Figure()
        fig.add_trace(go.Bar(
            name=f"{years[0]}YE",
            x=companies,
            y=y_2024,
            marker_color=color_2024,
            width=0.35,
            offset=-0.2,
            text=[f"{v:.2%}" for v in y_2024],  # 显示两位小数百分比
            textposition='outside',
        ))
        fig.add_trace(go.Bar(
            name=f"{years[1]}YE",
            x=companies,
            y=y_2025,
            marker_color=color_2025,
            width=0.35,
            offset=0.2,
            text=[f"{v:.2%}" for v in y_2025],
            textposition='outside',
        ))
        
        fig.update_layout(
            barmode='group',
            title=f"<b>{f}</b>",
            height=350,
            margin=dict(t=50, b=40, l=50, r=20),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            yaxis=dict(
                tickformat=".2%",
                title="占保险服务收入比例",
                range=y_range,
                zeroline=True,
                zerolinecolor='gray',
                zerolinewidth=1,
            ),
            xaxis=dict(
                tickangle=-30,
                tickfont=dict(size=10),
            )
        )
        figures.append(fig)
    
    return figures
    
# 5.通用多分类构成分析引擎
def create_kpmg_multi_composition_chart(df, field_map, color_map, title_prefix, show_labels, 
                                        label_size, bar_width, co_font_size, highlight_co="无",
                                        add_zero_line=False, years_to_show=None):
    # ===== 从 session_state 读取配置 =====
    selected_cos = st.session_state.get('selected_cos_cache', [])
    HL_BOX_FILL = st.session_state.get('HL_BOX_FILL', "rgba(0,51,141,0.03)")
    HL_BOX_LINE = st.session_state.get('HL_BOX_LINE', "rgba(0,51,141,0.35)")
    fields = list(field_map.keys())
    
    # 如果指定了年份过滤，则先过滤 df
    if years_to_show is not None:
        df = df[df['报告年份'].astype(str).isin([str(y) for y in years_to_show])].copy()
    
    d = df[df['公司'].isin(selected_cos)].drop_duplicates(subset=['公司', '报告年份', '字段名'], keep='first').copy()
    d['报告年份'] = d['报告年份'].astype(str).str.replace(".0", "", regex=False)
    
    d_p = d[d['字段名'].isin(fields)].pivot_table(
        index=['公司', '报告年份'], columns='字段名', values='(百万)人民币', aggfunc='first'
    ).fillna(0) if not d.empty else pd.DataFrame()
    
    valid_fields = [f for f in fields if f in d_p.columns]
    if valid_fields:
        d_p['Total'] = d_p[valid_fields].abs().sum(axis=1).replace(0, 1)
    else:
        d_p['Total'] = 1.0
    
    for f in fields:
        if f in d_p.columns:
            d_p[field_map[f]] = d_p[f] / d_p['Total'] * 100
        else:
            d_p[field_map[f]] = 0
    
    all_yrs = sorted(d['报告年份'].unique())
    av_cos = [co for co in selected_cos if co in d['公司'].unique()]
    if not av_cos:
        fig = go.Figure()
        fig.add_annotation(
            text="⚠️ 无数据：请检查公司或字段",
            x=0.5, y=0.5, showarrow=False,
            font=dict(size=18, color="red")
        )
        fig.update_layout(
            height=300,
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            xaxis=dict(showticklabels=False),
            yaxis=dict(showticklabels=False)
        )
        return fig, pd.DataFrame()
    
    titles = [f"<b>{co}</b>" for co in av_cos]
    fig = make_subplots(rows=1, cols=len(av_cos), shared_yaxes=True, 
                        horizontal_spacing=0.015, column_titles=titles)
    
    for i, co in enumerate(av_cos):
        d_co = d[d['公司']==co].pivot(index='报告年份', columns='字段名', values='(百万)人民币').reindex(all_yrs).fillna(0)
        valid_co_fields = [f for f in fields if f in d_co.columns]
        raw_total = d_co[valid_co_fields].abs().sum(axis=1) if valid_co_fields else pd.Series(0, index=d_co.index)
        d_co['Total_raw'] = raw_total
        
        for f, d_n in field_map.items():
            if f in d_co.columns:
                val = d_co[f] / d_co['Total_raw'].replace(0, 1) * 100
            else:
                val = pd.Series(0, index=d_co.index)
            
            clr = color_map.get(f, "#CCCCCC")
            is_dark = any(x in clr for x in ["00338D", "510DBC", "1E49E2", "0A2B5E", "FD349C"])
            txt_c = "white" if is_dark else "black"
            
            fig.add_trace(go.Bar(
                x=[f"{y}YE" for y in d_co.index], 
                y=val,
                name=d_n if i==0 else None,
                marker_color=clr,
                text=[f"{v:.0f}%" if abs(v) >= 1 else "" for v in val] if show_labels else None,
                textangle=0, textposition='inside', insidetextanchor='middle',
                textfont=dict(size=label_size, color=txt_c),
                constraintext='none', cliponaxis=False,
                width=bar_width, showlegend=(i==0), legendgroup=f
            ), row=1, col=i+1)
        
        missing_y = [100 if v == 0 else 0 for v in raw_total]
        missing_t = ["未披露" if v == 0 else "" for v in raw_total]
        fig.add_trace(go.Bar(
            x=[f"{y}YE" for y in d_co.index],
            y=missing_y,
            marker_color="#CDCDCD",
            text=missing_t,
            textangle=0, textposition='inside', insidetextanchor='middle',
            textfont=dict(size=12, color="white"),
            constraintext='none', cliponaxis=False,
            width=bar_width, showlegend=False, hoverinfo='skip'
        ), row=1, col=i+1)
        
        is_hl = (str(co).strip() == str(highlight_co).strip())
        bg_fill = HL_BOX_FILL if is_hl else "rgba(0,0,0,0)"
        line_dict = dict(color=HL_BOX_LINE, width=1.5) if is_hl else dict(color="rgba(0,0,0,0)", width=0)
        fig.add_shape(
            type="rect", xref="x domain", yref="y domain",
            x0=-0.06, x1=1.06, y0=-0.1, y1=1.08,
            fillcolor=bg_fill, line=line_dict, layer="above", row=1, col=i+1
        )
    
    df_avg = d_p[[field_map[f] for f in fields if f in d_p.columns]].groupby('报告年份').mean()
    df_avg = df_avg.reindex(all_yrs[-2:] if len(all_yrs)>=2 else all_yrs)
    df_avg.index = [f"{y}YE" for y in df_avg.index]
    
    fig.update_layout(
        barmode='relative',
        height=400,
        margin=dict(t=50, b=80, l=20, r=20),
        legend=dict(orientation="h", yanchor="top", y=-0.17, xanchor="center", x=0.5, font=dict(size=10)),
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)'
    )
    
    if add_zero_line:
        fig.add_hline(y=0, line_color="orange", line_width=1.5, layer="below")
    
    for ann in fig.layout.annotations:
        if "<b>" in str(ann.text):
            ann.update(y=1, font=dict(size=co_font_size, color="#00338D"))
    
    return fig, df_avg

#5.0 概览表格
def create_overview_table(df, cos, latest_year, prev_year, unit_label="百万元"):
    """
    生成概览表格：按保险服务收入排序，展示各公司关键指标同比变化（%）
    返回 Plotly 表格，表头单元格包含三行：指标名称、%变化、25YE/24YE-1
    """
    # 提取所需指标（内部字段名）
    metrics = ["保险服务收入", "承保利润", "净利润"]
    # 对应的显示名称（将“承保利润”显示为“保险服务业绩”）
    display_names = ["保险服务收入", "保险服务业绩", "净利润"]
    
    # 获取最近两年数据
    df_latest = df[df['报告年份'].astype(str) == str(latest_year)]
    df_prev = df[df['报告年份'].astype(str) == str(prev_year)]
    
    # 计算各公司保险服务收入（用于排序）
    premium_latest = {}
    for co in cos:
        val = df_latest[(df_latest['公司'] == co) & (df_latest['字段名'] == '保险服务收入')]['(百万)人民币']
        premium_latest[co] = val.iloc[0] if not val.empty else 0
    
    # 按保险服务收入降序排序
    sorted_cos = sorted(cos, key=lambda x: premium_latest.get(x, 0), reverse=True)
    
    data = {'公司': sorted_cos}
    for metric in metrics:
        pct_changes = []
        for co in sorted_cos:
            curr_val = df_latest[(df_latest['公司'] == co) & (df_latest['字段名'] == metric)]['(百万)人民币']
            prev_val = df_prev[(df_prev['公司'] == co) & (df_prev['字段名'] == metric)]['(百万)人民币']
            curr = curr_val.iloc[0] if not curr_val.empty else None
            prev = prev_val.iloc[0] if not prev_val.empty else None
            if curr is not None and prev is not None and prev != 0:
                pct = (curr - prev) / abs(prev)
            else:
                pct = None
            pct_changes.append(pct)
        data[metric] = pct_changes
    
    df_table = pd.DataFrame(data)
    # 格式化百分比
    def fmt_pct(x):
        if pd.isna(x):
            return "N/A"
        return f"{x:.1%}"
    for m in metrics:
        df_table[m] = df_table[m].apply(fmt_pct)
    
    # 构建表头：第一列"公司"也使用三行（补两个空行），其他列显示三行
    header_values = ['公司<br><br>']
    for display_name in display_names:
        header_values.append(f"{display_name}<br>%变化<br>25YE/24YE-1")
    
    # 使用 Plotly 绘制表格
    fig = go.Figure(data=[go.Table(
        header=dict(
            values=header_values,
            fill_color='#00338D',
            align='center',
            font=dict(color='white', size=13)
        ),
        cells=dict(
            values=[df_table['公司']] + [df_table[m] for m in metrics],
            fill_color=[['#F8F9FA', 'white'] * (len(df_table)//2 + 1)],
            align='center',
            font=dict(size=12)
        )
    )])
    fig.update_layout(
        title="关键指标同比变化（%）",
        margin=dict(l=10, r=10, t=50, b=10),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)'
    )
    return fig

# 5.1 通用单指标柱状图
def create_kpmg_chart(df, field_name, title_prefix, show_labels, pct_font_size, global_gap,
                      sort_by_value=False, is_percentage=False):
    # ===== 从 session_state 读取配置 =====
    selected_cos = st.session_state.get('selected_cos_cache', [])
    divisor = st.session_state.get('divisor', 1)
    prev_year = st.session_state.get('prev_year', 2024)
    latest_year = st.session_state.get('latest_year', 2025)
    highlight_co = st.session_state.get('highlight_co', "无")
    unit_label = st.session_state.get('unit_label', '百万元')
    HL_BOX_FILL = st.session_state.get('HL_BOX_FILL', "rgba(0,51,141,0.03)")
    HL_BOX_LINE = st.session_state.get('HL_BOX_LINE', "rgba(0,51,141,0.35)")

    # 清洗字段名和公司名
    df_clean = df.copy()
    df_clean['字段名'] = df_clean['字段名'].astype(str).str.strip()
    df_clean['公司'] = df_clean['公司'].astype(str).str.strip()
    field_name_clean = field_name.strip()
    selected_cos = [c.strip() for c in selected_cos]

    # 过滤数据
    d = df_clean[df_clean['公司'].isin(selected_cos)].copy()
    d['报告年份'] = d['报告年份'].astype(str).str.replace('.0', '', regex=False).str.strip()
    val_col = '(百万)人民币' if '(百万)人民币' in d.columns else d.columns[-1]

    # 模糊匹配字段名
    norm_field = normalize_field(field_name_clean)
    df_plot = d[d['字段名'].apply(lambda x: normalize_field(x) == norm_field)].copy()

    if df_plot.empty:
        fig = go.Figure()
        fig.add_annotation(text="⚠️ 无数据：请检查字段名或数据源", x=0.5, y=0.5, showarrow=False, font=dict(size=18, color="red"))
        fig.update_layout(height=700, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                          xaxis=dict(showticklabels=False), yaxis=dict(showticklabels=False))
        return fig

    # 数值缩放
    df_plot['value'] = df_plot[val_col]
    if not is_percentage:
        df_plot['value'] = df_plot['value'] / divisor

    y_old, y_new = str(prev_year), str(latest_year)

    # 排序
    if sort_by_value:
        df_new = df_plot[df_plot['报告年份'] == y_new].set_index('公司')['value']
        sorted_cos = sorted(selected_cos, key=lambda co: df_new.get(co, -float('inf')), reverse=True)
    else:
        sorted_cos = selected_cos

    fig = go.Figure()
    color_map = {y_old: "#1E49E2", y_new: "#00338D"}

    valid_vals = df_plot['value'].dropna()
    all_max = valid_vals.max() if not valid_vals.empty else 100
    abs_max = valid_vals.abs().max() if not valid_vals.empty else 100
    fixed_offset = all_max * 0.2
    placeholder_h = (abs_max * 0.15 if abs_max > 0 else 10) / divisor

    def fmt(val):
        if is_percentage:
            return f"{val:.1%}" if pd.notna(val) else ""
        else:
            if pd.notna(val):
                if divisor >= 1000:
                    return f"{val:,.2f}"
                elif divisor >= 100:
                    return f"{val:,.2f}"
                else:
                    return f"{val:,.0f}"
            return ""

    # 绘制柱状图
    for yr in [y_old, y_new]:
        df_yr = df_plot[df_plot['报告年份'] == yr].set_index('公司').reindex(sorted_cos).reset_index()
        y_vals, m_colors, t_texts, t_colors, t_pos = [], [], [], [], []
        for v in df_yr['value']:
            if pd.isna(v):
                y_vals.append(placeholder_h)
                m_colors.append("#CDCDCD")
                t_texts.append("未披露")
                t_colors.append("white")
                t_pos.append("inside")
            else:
                y_vals.append(v)
                m_colors.append(color_map[yr])
                t_texts.append(fmt(v) if show_labels and v != 0 else "")
                t_colors.append("#333333")
                t_pos.append("outside")
        fig.add_trace(go.Bar(
            name=f"{yr}YE",
            x=df_yr['公司'],
            y=y_vals,
            marker_color=m_colors,
            text=t_texts,
            textposition=t_pos,
            textfont=dict(size=12, color=t_colors),
            cliponaxis=False,
            textangle=0,
            constraintext='none'
        ))

    # 标注增长率
    df_old_series = df_plot[df_plot['报告年份'] == y_old].set_index('公司')['value']
    df_new_series = df_plot[df_plot['报告年份'] == y_new].set_index('公司')['value']
    for co in sorted_cos:
        v_old = df_old_series.get(co, np.nan)
        v_new = df_new_series.get(co, np.nan)
        if pd.notna(v_old) and pd.notna(v_new) and v_old != 0:
            if is_percentage:
                diff = (v_new - v_old) * 100
                label = f"↗ +{diff:.1f}pp" if diff >= 0 else f"↘ {diff:.1f}pp"
                color = "#FD349C" if diff >= 0 else "#269924"
            else:
                pct = (v_new - v_old) / abs(v_old)
                label = f"↗ {pct:.1%}" if pct >= 0 else f"↘ {pct:.1%}"
                color = "#FD349C" if pct >= 0 else "#269924"
            fig.add_annotation(
                x=co,
                y=max(v_old, v_new) + fixed_offset,
                text=f"<b>{label}</b>",
                showarrow=False,
                font=dict(color=color, size=pct_font_size),
                xshift=15
            )

    # 高亮框（特定追踪公司）
    if highlight_co in sorted_cos:
        idx = sorted_cos.index(highlight_co)
        fig.add_shape(
            type="rect",
            xref="x", yref="paper",
            x0=idx - 0.45, x1=idx + 0.45,
            y0=-0.08, y1=1,
            fillcolor=HL_BOX_FILL,
            line=dict(color=HL_BOX_LINE, width=1.5),
            layer="below"
        )

    # ===== 🆕 添加灰色公司边框 =====
    fig = add_company_borders(fig, sorted_cos, sorted_cos, top_margin=70)
    # 隐藏 x 轴下方的公司名称（避免重复）
    fig.update_xaxes(showticklabels=False)

    # ===== 布局设置 =====
    fig.update_layout(
        barmode='group',
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        bargroupgap=0.0,
        bargap=global_gap,
        legend=dict(
            orientation="v",        # 改为垂直
            yanchor="middle",
            y=0.5,
            xanchor="right",
            x=-0.1,                # 图例紧贴图表左侧
            font=dict(size=12)
        ),
        margin=dict(t=70, b=40, l=20, r=20),
        height=700
    )
    fig.update_xaxes(showgrid=False, zeroline=False, showline=False)

    # ===== 设置纵轴 =====
    y_vals = df_plot['value'].dropna()
    if not y_vals.empty:
        y_min = y_vals.min()
        y_max = y_vals.max()
        padding = (y_max - y_min) * 0.2 if y_max != y_min else abs(y_max) * 0.3
        y_range = [y_min - padding if y_min < 0 else 0,
                   y_max + padding if y_max > 0 else 0]
        if y_range[0] == 0 and y_range[1] == 0:
            y_range = [0, 1]

        all_max_abs = y_vals.abs().max()
        fixed_offset = all_max_abs * 0.2
        max_label_y = y_max + fixed_offset if y_max > 0 else y_max
        y_range[1] = max(y_range[1], max_label_y) * 1.05

        data_range = y_range[1] - y_range[0]
        if data_range > 0:
            raw_step = data_range / 6
            magnitude = 10 ** (len(str(int(raw_step))) - 1) if raw_step >= 1 else 10 ** (len(str(int(raw_step * 10))) - 2)
            for base in [1, 2, 5]:
                candidate = base * magnitude
                if candidate >= raw_step / 2:
                    step = candidate
                    break
            else:
                step = magnitude
        else:
            step = None
    else:
        y_range = [0, 1]
        step = 1

    fig.update_yaxes(
        showgrid=False,
        zeroline=True,
        zerolinecolor="#E0E0E0",
        zerolinewidth=1,
        showline=False,
        range=y_range,
        dtick=step,
        tickformat=',.0f' if divisor <= 1 else ',.2f',
        title_text="百分比" if is_percentage else f"{unit_label}人民币",
        title_font=dict(size=12, color="#333333"),
        exponentformat='none',
    )

    return fig
                          
# 5.2 保险业务构成（两种方法占比堆叠图）
def create_method_composition_chart(df, cos, year, divisor=1, unit_label="百万元"):
    """
    绘制每家公司保险服务收入采用保费分配法与未采用保费分配法的占比堆叠图
    横轴为公司，每个公司显示两个年份（去年和今年）的堆叠柱，纵轴为百分比。
    """
    # 筛选字段（原始名称，用于匹配）
    field_未采用_raw = "保险业务构成(未采用保费分配法)"
    field_采用_raw = "保险业务构成（采用保费分配法）"
    
    # 获取年份列表（去年和今年）
    prev_year = st.session_state.get('prev_year', 2024)
    latest_year = st.session_state.get('latest_year', 2025)
    years = [str(prev_year), str(latest_year)]
    
    # 提取数据，清洗年份（去除 .0 和可能的 YE 后缀）
    df_year = df.copy()
    df_year['报告年份'] = df_year['报告年份'].astype(str).str.replace('.0', '', regex=False).str.strip()
    df_year['报告年份'] = df_year['报告年份'].str.replace('YE', '', regex=False).str.strip()
    df_year = df_year[df_year['报告年份'].isin(years)]
    df_year = df_year[df_year['公司'].isin(cos)]
    
    # 检查数据是否存在（模糊匹配）
    available_fields = df_year['字段名'].unique()
    norm_available = [normalize_field(f) for f in available_fields]
    target_未 = normalize_field(field_未采用_raw)
    target_采 = normalize_field(field_采用_raw)
    
    actual_未 = None
    actual_采 = None
    for f in available_fields:
        if normalize_field(f) == target_未:
            actual_未 = f
        if normalize_field(f) == target_采:
            actual_采 = f
    
    if actual_未 is None or actual_采 is None:
        fig = go.Figure()
        fig.add_annotation(
            text="未找到保费分配法构成数据",
            x=0.5, y=0.5, showarrow=False,
            font=dict(size=16, color="red")
        )
        fig.update_layout(height=300, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
        return fig
    
    # 构建数据结构：{(公司, 年份): {未采用占比, 采用占比}}
    data = {}
    for co in cos:
        for yr in years:
            # 获取未采用占比
            val_un = df_year[(df_year['公司'] == co) & (df_year['报告年份'] == yr) & (df_year['字段名'] == actual_未)]['(百万)人民币']
            val_un = val_un.iloc[0] if not val_un.empty else 0
            # 获取采用占比
            val_ad = df_year[(df_year['公司'] == co) & (df_year['报告年份'] == yr) & (df_year['字段名'] == actual_采)]['(百万)人民币']
            val_ad = val_ad.iloc[0] if not val_ad.empty else 0
            # 如果两个都是0，跳过
            if val_un == 0 and val_ad == 0:
                data[(co, yr)] = None
            else:
                # 转换为百分比（小数转百分数）
                data[(co, yr)] = {
                    "未采用": val_un * 100,
                    "采用": val_ad * 100
                }
    
    # 如果没有有效数据
    if not any(v is not None for v in data.values()):
        fig = go.Figure()
        fig.add_annotation(
            text="所有公司均无有效构成数据",
            x=0.5, y=0.5, showarrow=False,
            font=dict(size=16, color="red")
        )
        fig.update_layout(height=300, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
        return fig
    
    # 创建子图布局：每个公司一列，共享y轴
    num_companies = len(cos)
    fig = make_subplots(
        rows=1, cols=num_companies,
        shared_yaxes=True,
        horizontal_spacing=0.02
    )
    
    # 颜色
    colors = {"未采用": "#00338D", "采用": "#0865EE"}
    
    for i, co in enumerate(cos):
        col_idx = i + 1
        x_vals = []
        y_un = []
        y_ad = []
        # 按年份顺序（去年、今年）
        for yr in years:
            x_vals.append(f"{yr}YE")
            d = data.get((co, yr))
            if d is None:
                y_un.append(0)
                y_ad.append(0)
            else:
                y_un.append(d["未采用"])
                y_ad.append(d["采用"])
        
        # 未采用 trace
        fig.add_trace(
            go.Bar(
                x=x_vals,
                y=y_un,
                name="未采用保费分配法",
                legendgroup="未采用保费分配法",
                marker_color=colors["未采用"],
                text=[f"{v:.1f}%" if v > 0 else "" for v in y_un],
                textposition='inside',
                insidetextanchor='middle',
                textfont=dict(color="white", size=11),
                width=0.6,
                showlegend=(i == 0)      # 只在第一个子图显示图例
            ),
            row=1, col=col_idx
        )
        # 采用 trace
        fig.add_trace(
            go.Bar(
                x=x_vals,
                y=y_ad,
                name="采用保费分配法",
                legendgroup="采用保费分配法",
                marker_color=colors["采用"],
                text=[f"{v:.1f}%" if v > 0 else "" for v in y_ad],
                textposition='inside',
                insidetextanchor='middle',
                textfont=dict(color="white", size=11),
                width=0.6,
                showlegend=(i == 0)
            ),
            row=1, col=col_idx
        )
        
        # ===== 🆕 为当前子图添加公司边框 =====
        # x_vals 是该子图的横轴标签（年份）
        fig = add_company_borders(fig, [co], x_vals, top_margin=40, row=1, col=col_idx)
    
    fig.update_layout(
        barmode='relative',  # 堆叠模式
        height=400,
        margin=dict(t=40, b=40, l=40, r=20),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="center",
            x=0.5
        ),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        yaxis=dict(
            title="占比（%）",
            tickformat=".0f",
            range=[0, 105]
        )
    )
    
    # 更新每个子图的x轴和y轴
    for i in range(1, num_companies + 1):
        fig.update_xaxes(
            tickangle=0,
            showline=False,
            showgrid=False,
            zeroline=False,
            row=1, col=i
        )
        fig.update_yaxes(
            showgrid=False,
            zeroline=False,
            showline=False,
            row=1, col=i
        )
    
    return fig

# 5.3 利润构成图（含公司边框）
def create_profit_composition_chart(df, cos, year, divisor=1, unit_label="百万元"):
    """
    绘制每家公司承保利润与投资利润的分组条形图（实际金额），并标注占比。
    添加灰色公司边框，隐藏x轴公司名称。
    """
    field_uw = "承保利润"
    field_inv = "投资利润"
    prev_year = st.session_state.get('prev_year', 2024)
    latest_year = st.session_state.get('latest_year', 2025)
    years = [str(prev_year), str(latest_year)]

    df_year = df.copy()
    df_year['报告年份'] = df_year['报告年份'].astype(str).str.replace('.0', '', regex=False).str.strip()
    df_year['报告年份'] = df_year['报告年份'].str.replace('YE', '', regex=False).str.strip()
    df_year = df_year[df_year['报告年份'].isin(years)]
    df_year = df_year[df_year['公司'].isin(cos)]

    available_fields = df_year['字段名'].unique()
    norm_available = [normalize_field(f) for f in available_fields]
    target_uw = normalize_field(field_uw)
    target_inv = normalize_field(field_inv)

    actual_uw = None
    actual_inv = None
    for f in available_fields:
        if normalize_field(f) == target_uw:
            actual_uw = f
        if normalize_field(f) == target_inv:
            actual_inv = f

    if actual_uw is None or actual_inv is None:
        fig = go.Figure()
        fig.add_annotation(text="未找到承保利润或投资利润数据", x=0.5, y=0.5, showarrow=False, font=dict(size=16, color="red"))
        fig.update_layout(height=400, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
        return fig

    # 提取各公司各年份的原始值（已换算为百万元）
    raw_data = {}
    for co in cos:
        for yr in years:
            val_uw = df_year[(df_year['公司'] == co) & (df_year['报告年份'] == yr) & (df_year['字段名'] == actual_uw)]['(百万)人民币']
            val_uw = val_uw.iloc[0] if not val_uw.empty else 0
            val_inv = df_year[(df_year['公司'] == co) & (df_year['报告年份'] == yr) & (df_year['字段名'] == actual_inv)]['(百万)人民币']
            val_inv = val_inv.iloc[0] if not val_inv.empty else 0
            raw_data[(co, yr)] = {"承保利润": val_uw / divisor, "投资利润": val_inv / divisor}

    yr_str = str(year)
    companies = cos
    data = {co: raw_data.get((co, yr_str), {"承保利润": 0, "投资利润": 0}) for co in companies}

    labels = ["承保利润", "投资利润"]
    colors = ["#00338D", "#0865EE"]
    fig = go.Figure()

    # 收集所有值用于确定y轴范围
    all_vals = []
    for co in companies:
        all_vals.append(data[co]["承保利润"])
        all_vals.append(data[co]["投资利润"])

    y_max = max(all_vals) if all_vals else 1
    y_min = min(all_vals) if all_vals else 0
    y_range = [min(0, y_min * 1.2), max(0, y_max * 1.3)]

    # 绘制分组条形
    for i, (label, color) in enumerate(zip(labels, colors)):
        values = [data[co][label] for co in companies]
        ratios = []
        for co in companies:
            total = data[co]["承保利润"] + data[co]["投资利润"]
            if total != 0:
                ratios.append(data[co][label] / total)
            else:
                ratios.append(0)
        texts = []
        for val, rat in zip(values, ratios):
            if val == 0:
                texts.append("")
            else:
                texts.append(f"{val:.1f}\n({rat:.1%})")
        fig.add_trace(go.Bar(
            name=label,
            x=companies,
            y=values,
            marker_color=color,
            text=texts,
            textposition='outside',
            textfont=dict(size=10),
            width=0.35,
            offsetgroup=i,
            cliponaxis=False
        ))

    fig.add_hline(y=0, line_dash="dash", line_color="gray", line_width=1.5, opacity=0.7)

    # ===== 🆕 添加灰色公司边框 =====
    fig = add_company_borders(fig, companies, companies, top_margin=70)
    # 隐藏 x 轴下方的公司名称（避免重复）
    fig.update_xaxes(showticklabels=False)

    fig.update_layout(
        barmode='group',
        title=f"各公司利润构成（最新年 {year}YE）",
        xaxis_title="公司",
        yaxis_title=f"金额（{unit_label}）",
        yaxis=dict(range=y_range, tickformat=",.0f", zeroline=True, zerolinecolor='gray'),
        height=450,
        margin=dict(t=70, b=60, l=40, r=20),  # 顶部边距从 50 改为 70
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        bargap=0.15,
        bargroupgap=0.1,
    )
    return fig

def create_profit_stacked_pct_chart(df, cos, years, divisor=1, unit_label="百万元", highlight_co="无"):
    """
    绘制每家公司承保利润与投资利润的百分比堆叠柱状图（两年对比）
    每个公司一个子图，柱子的正负方向反映各组成部分的实际正负贡献。
    正贡献向上，负贡献向下，两者的绝对值之和为100%（基于|承保|+|投资|）。
    柱顶/底显示总利润金额，内部显示各部分的绝对值百分比。
    图例水平置于底部，纵轴自动适配数据范围。
    """
    field_uw = "承保利润"
    field_inv = "投资利润"

    df_year = df.copy()
    df_year['报告年份'] = df_year['报告年份'].astype(str).str.replace('.0', '', regex=False).str.strip()
    df_year = df_year[df_year['报告年份'].isin([str(y) for y in years])]
    df_year = df_year[df_year['公司'].isin(cos)]

    available_fields = df_year['字段名'].unique()
    norm_available = [normalize_field(f) for f in available_fields]
    target_uw = normalize_field(field_uw)
    target_inv = normalize_field(field_inv)

    actual_uw = None
    actual_inv = None
    for f in available_fields:
        if normalize_field(f) == target_uw:
            actual_uw = f
        if normalize_field(f) == target_inv:
            actual_inv = f

    if actual_uw is None or actual_inv is None:
        fig = go.Figure()
        fig.add_annotation(text="未找到承保利润或投资利润数据", x=0.5, y=0.5, showarrow=False, font=dict(size=16, color="red"))
        fig.update_layout(height=400, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
        return fig

    # 提取数据
    data = {}
    for co in cos:
        data[co] = {}
        for yr in years:
            yr_str = str(yr)
            val_uw = df_year[(df_year['公司'] == co) & (df_year['报告年份'] == yr_str) & (df_year['字段名'] == actual_uw)]['(百万)人民币']
            val_uw = val_uw.iloc[0] if not val_uw.empty else 0.0
            val_inv = df_year[(df_year['公司'] == co) & (df_year['报告年份'] == yr_str) & (df_year['字段名'] == actual_inv)]['(百万)人民币']
            val_inv = val_inv.iloc[0] if not val_inv.empty else 0.0
            data[co][yr_str] = {
                "承保": val_uw / divisor,
                "投资": val_inv / divisor,
                "总计": (val_uw + val_inv) / divisor
            }

    valid_cos = [co for co in cos if any(abs(data[co][str(yr)]["承保"]) + abs(data[co][str(yr)]["投资"]) != 0 for yr in years)]
    if not valid_cos:
        fig = go.Figure()
        fig.add_annotation(text="所有公司承保和投资均为0", x=0.5, y=0.5, showarrow=False)
        return fig

    n = len(valid_cos)
    fig = make_subplots(rows=1, cols=n, shared_yaxes=True,
                        subplot_titles=[f"<b>{co}</b>" for co in valid_cos],
                        horizontal_spacing=0.02 if n > 1 else 0)

    color_uw = "#00338D"
    color_inv = "#1E49E2"

    # 用于收集所有 y 值，以便动态调整纵轴范围
    all_y_values = []

    for col_idx, co in enumerate(valid_cos):
        col = col_idx + 1
        x_labels = [f"{yr}YE" for yr in years]
        uw_pcts = []
        inv_pcts = []
        total_vals = []

        for yr in years:
            yr_str = str(yr)
            total = data[co][yr_str]["总计"]
            total_vals.append(total)
            abs_sum = abs(data[co][yr_str]["承保"]) + abs(data[co][yr_str]["投资"])
            if abs_sum != 0:
                uw_pct = (data[co][yr_str]["承保"] / abs_sum) * 100
                inv_pct = (data[co][yr_str]["投资"] / abs_sum) * 100
            else:
                uw_pct = inv_pct = 0
            uw_pcts.append(uw_pct)
            inv_pcts.append(inv_pct)
            all_y_values.extend([uw_pct, inv_pct])

        # 承保部分
        fig.add_trace(go.Bar(
            x=x_labels,
            y=uw_pcts,
            name="保险服务业绩" if col_idx == 0 else None,
            marker_color=color_uw,
            text=[f"{abs(v):.1f}%" if v != 0 else "" for v in uw_pcts],
            textposition='inside',
            insidetextanchor='middle',
            textfont=dict(color="white", size=11),
            width=0.6,
            showlegend=(col_idx == 0)
        ), row=1, col=col)

        # 投资部分
        fig.add_trace(go.Bar(
            x=x_labels,
            y=inv_pcts,
            name="投资服务业绩" if col_idx == 0 else None,
            marker_color=color_inv,
            text=[f"{abs(v):.1f}%" if v != 0 else "" for v in inv_pcts],
            textposition='inside',
            insidetextanchor='middle',
            textfont=dict(color="white", size=11),
            width=0.6,
            showlegend=(col_idx == 0)
        ), row=1, col=col)

        # 标注总利润金额
        for idx, yr in enumerate(years):
            total = total_vals[idx]
            if total != 0:
                abs_sum = abs(data[co][str(yr)]["承保"]) + abs(data[co][str(yr)]["投资"])
                net_pct = (total / abs_sum) * 100 if abs_sum != 0 else 0
                y_pos = net_pct + (5 if net_pct >= 0 else -5)
                fig.add_annotation(
                    x=x_labels[idx],
                    y=y_pos,
                    text=f"{total:.1f}",
                    showarrow=False,
                    font=dict(size=10, color="#333"),
                    row=1, col=col,
                    xanchor="center",
                    yanchor="bottom" if net_pct >= 0 else "top"
                )

        # 高亮公司边框
        if highlight_co != "无" and co == highlight_co:
            fig.add_shape(
                type="rect",
                xref="x domain", yref="y domain",
                x0=-0.05, x1=1.05, y0=-1.1, y1=1.1,
                fillcolor="rgba(0,51,141,0.05)",
                line=dict(color="rgba(0,51,141,0.8)", width=1.5),
                layer="above",
                row=1, col=col
            )

    # 动态计算 y 轴范围
    if all_y_values:
        min_y = min(all_y_values)
        max_y = max(all_y_values)
        # 留 10% 边距
        padding = (max_y - min_y) * 0.1 if max_y != min_y else 10
        y_range = [min_y - padding, max_y + padding]
        # 如果最小值为正，下限设为0；如果最大值为负，上限设为0（但通常有正有负）
        if min_y >= 0:
            y_range[0] = 0
        if max_y <= 0:
            y_range[1] = 0
        # 确保至少留出标注空间（±5%）
        if y_range[0] > -5:
            y_range[0] = min(y_range[0], -5)
        if y_range[1] < 5:
            y_range[1] = max(y_range[1], 5)
    else:
        y_range = [-10, 10]

    # 布局
    fig.update_layout(
        barmode='relative',
        height=450,
        margin=dict(t=60, b=50, l=40, r=20),   # 底部留空间给图例
        legend=dict(
            orientation="h",                    # 水平排列
            yanchor="top",
            y=-0.15,                           # 放在图表底部外侧
            xanchor="center",
            x=0.5,
            font=dict(size=12)
        ),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        bargap=0.15,
        bargroupgap=0.1
    )

    # 应用 y 轴范围到所有子图
    fig.update_yaxes(range=y_range, tickformat=".0f", title_text="占比（%）", row=1, col=1)
    for i in range(1, n+1):
        fig.update_xaxes(tickangle=0, row=1, col=i)

    # 调整子图标题（公司名称）位置，避免与图例重叠
    for ann in fig.layout.annotations:
        if ann.text and "<b>" in ann.text:
            ann.update(y=1.02)

    return fig
    
# 5.4 综合赔付率拆解堆叠图（自动提取年份，优化配色 + 公司边框 + 年份标注）
def create_cor_breakdown_stacked_chart(df, cos, latest_year, prev_year, divisor=1, unit_label="百万元", highlight_co="无"):
    """
    绘制综合赔付率拆解的堆叠柱状图（各因子占保险服务收入比例）
    如果某公司某年份没有拆解因子，则用一个灰色柱子显示综合成本率。
    """
    factors = [
        "当期发生赔款及理赔费用",
        "已发生赔款负债履约现金流变动",
        "亏损合同损益",
        "承保财务损益",
        "再保净成本",
        "提取保费准备金"
    ]
    
    factor_colors = {
        "当期发生赔款及理赔费用": "#00338D",
        "已发生赔款负债履约现金流变动": "#510DBC",
        "亏损合同损益": "#B0BEC5",
        "承保财务损益": "#76D2FF",
        "再保净成本": "#FD349C",
        "提取保费准备金": "#00C0AE"
    }
    
    raw = df.copy()
    raw['报告年份'] = raw['报告年份'].astype(str).str.replace('.0', '', regex=False)
    raw['公司'] = raw['公司'].astype(str).str.strip()
    raw['字段名'] = raw['字段名'].astype(str).str.strip()
    
    # 提取实际存在的年份
    available_years = sorted(
        [int(y) for y in raw['报告年份'].unique() if y.isdigit()],
        reverse=True
    )
    if len(available_years) >= 2:
        years = [str(available_years[1]), str(available_years[0])]
        year_display = f"{available_years[1]}YE vs {available_years[0]}YE"
    elif len(available_years) == 1:
        years = [str(available_years[0]), str(available_years[0])]
        year_display = f"{available_years[0]}YE"
    else:
        years = [str(prev_year), str(latest_year)]
        year_display = f"{prev_year}YE vs {latest_year}YE"
    
    # 计算分母（保险服务收入）
    service_revenue = {}
    for co in cos:
        for yr in years:
            rev = raw[(raw['公司'] == co) & (raw['报告年份'] == yr) & (raw['字段名'] == '保险服务收入')]['(百万)人民币']
            service_revenue[(co, yr)] = rev.sum() if not rev.empty else 1.0
    
    # 构建数据
    rows = []
    for co in cos:
        for yr in years:
            row = {'公司': co, '年份': yr}
            for f in factors:
                val = raw[(raw['公司'] == co) & (raw['报告年份'] == yr) & (raw['字段名'] == f)]['(百万)人民币']
                val_sum = val.sum() if not val.empty else 0.0
                ratio = val_sum / service_revenue[(co, yr)] * 100
                row[f] = ratio
            cor_val = raw[(raw['公司'] == co) & (raw['报告年份'] == yr) & (raw['字段名'] == '综合成本率')]['(百万)人民币']
            row['ratio_value'] = cor_val.iloc[0] if not cor_val.empty else np.nan
            rows.append(row)
    df_plot = pd.DataFrame(rows)
    
    # 填充 NaN 为 0
    df_plot[factors] = df_plot[factors].fillna(0)
    df_plot['factor_sum'] = df_plot[factors].sum(axis=1)
    
    # 调试输出（可删除）
    # st.write("太平产险数据：", df_plot[df_plot['公司']=='太平产险'])
    
    # 判断无因子但有综合成本率的行
    no_factor_mask = (df_plot['factor_sum'] == 0) & df_plot['ratio_value'].notna() & (df_plot['ratio_value'] != 0)
    
    factor_cols = factors[:]
    if no_factor_mask.any():
        df_plot.loc[no_factor_mask, factors] = 0
        df_plot['综合成本率'] = 0.0
        df_plot.loc[no_factor_mask, '综合成本率'] = df_plot.loc[no_factor_mask, 'ratio_value'] * 100
        factor_cols.append('综合成本率')
        factor_colors['综合成本率'] = '#B0BEC5'
    
    # 构造x轴标签
    df_plot['x_label'] = df_plot['公司'] + '<br>' + df_plot['年份'] + 'YE'
    x_labels = df_plot['x_label'].unique()
    
    fig = go.Figure()
    for f in factor_cols:
        fig.add_trace(go.Bar(
            x=df_plot['x_label'],
            y=df_plot[f],
            name=f,
            marker_color=factor_colors.get(f, '#CCCCCC'),
            legendgroup=f,
            text=[f"{v:.1f}%" if abs(v) > 0.5 else "" for v in df_plot[f]],
            textposition='inside',
            insidetextanchor='middle',
            textfont=dict(size=9, color='white' if f != '综合成本率' else '#333'),
            hovertemplate=f"{f}: %{{y:.1f}}%<extra>%{{x}}</extra>"
        ))
    
    fig.add_hline(y=0, line_dash="dash", line_color="gray", line_width=1, opacity=0.5)
    fig = add_company_borders(fig, cos, x_labels, top_margin=70)
    
    all_labels = df_plot['x_label'].tolist()
    year_labels = [label.split('<br>')[1] if '<br>' in label else label for label in all_labels]
    fig.update_xaxes(tickvals=all_labels, ticktext=year_labels, showticklabels=True, tickangle=0)
    
    if highlight_co != "无":
        highlight_indices = [i for i, label in enumerate(all_labels) if highlight_co in label]
        for idx in highlight_indices:
            fig.add_shape(
                type="rect", xref="x", yref="y",
                x0=idx - 0.48, x1=idx + 0.48, y0=0, y1=1,
                fillcolor="rgba(0,51,141,0.05)",
                line=dict(color="rgba(0,51,141,0.8)", width=1.5), layer="below"
            )
    
    fig.update_layout(
        barmode='relative',
        title=f"综合赔付率拆解（{year_display}）",
        xaxis_title="", yaxis_title="占保险服务收入比例（%）",
        legend=dict(orientation="v", yanchor="middle", y=0.5, xanchor="right", x=-0.15, font=dict(size=11)),
        height=550,
        margin=dict(l=20, r=20, t=70, b=60),
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
        bargap=0.15, bargroupgap=0.1, hovermode='x unified'
    )
    return fig
    
# 5.5 保险服务收入业务构成堆叠图（保费贡献） - 按全局公司顺序
def create_premium_stacked_chart(df, cos, year, divisor=1, unit_label="百万元", highlight_co="无"):
    """
    绘制各公司保险服务收入业务构成堆叠图（按险种）
    """
    field_type = "保险服务收入"
    prefix = f"{field_type}-"
    all_fields = df['字段名'].unique()
    business_fields = [f for f in all_fields if isinstance(f, str) and f.startswith(prefix)]
    business_fields = [f for f in business_fields if not f.endswith("合计")]

    if not business_fields:
        fig = go.Figure()
        fig.add_annotation(text=f"未找到{field_type}业务构成数据", x=0.5, y=0.5, showarrow=False, font=dict(size=16, color="red"))
        fig.update_layout(height=400)
        return fig

    df_year = df[df['报告年份'].astype(str) == str(year)]
    df_year = df_year[df_year['公司'].isin(cos)]

    # 构建数据透视
    pivot_data = []
    for co in cos:
        row = {'公司': co}
        for f in business_fields:
            val_series = df_year[(df_year['公司'] == co) & (df_year['字段名'] == f)]['(百万)人民币']
            val = val_series.iloc[0] if not val_series.empty else 0
            label = f.replace(prefix, "")
            row[label] = val / divisor
        pivot_data.append(row)
    df_plot = pd.DataFrame(pivot_data)
    display_labels = [c for c in df_plot.columns if c != '公司']

    if not df_plot[display_labels].map(lambda x: x != 0).any().any():
        fig = go.Figure()
        fig.add_annotation(text="所有公司业务构成均为零或未披露", x=0.5, y=0.5, showarrow=False, font=dict(size=16, color="red"))
        fig.update_layout(height=400)
        return fig

    # 按传入的 cos 顺序确保 df_plot 的顺序
    df_plot = df_plot.set_index('公司').reindex(cos).reset_index()

    # 计算每个公司的总金额（用于百分比）
    totals = df_plot[display_labels].sum(axis=1)

    # ===== 唯一颜色分配 =====
    import plotly.express as px
    base_colors = KPMG_COLORS
    if len(display_labels) > len(base_colors):
        extra_colors = px.colors.qualitative.Plotly
        extra_unique = [c for c in extra_colors if c not in base_colors]
        all_colors = base_colors + extra_unique
    else:
        all_colors = base_colors
    if len(display_labels) > len(all_colors):
        all_colors = all_colors + px.colors.qualitative.Alphabet
    color_map = {label: all_colors[i % len(all_colors)] for i, label in enumerate(display_labels)}

    fig = go.Figure()

    for i, label in enumerate(display_labels):
        values = df_plot[label].fillna(0).tolist()
        ratios = []
        for val, total in zip(values, totals):
            if total != 0:
                ratios.append(val / total * 100)
            else:
                ratios.append(0)
        texts = []
        for val, rat in zip(values, ratios):
            if val == 0:
                texts.append("")
            else:
                texts.append(f"{rat:.1f}%")
        fig.add_trace(go.Bar(
            name=label,
            x=df_plot['公司'],
            y=ratios,
            marker_color=color_map[label],
            text=texts,
            textposition='inside',
            insidetextanchor='middle',
            textfont=dict(size=10, color='white' if i % 2 == 0 else 'black'),
            hovertemplate=f"{label}: %{{y:.1f}}%<extra>%{{x}}</extra>",
            showlegend=True,
            legendgroup=label,
        ))

    fig.add_hline(y=0, line_dash="dash", line_color="gray", line_width=1, opacity=0.5)

    if highlight_co != "无" and highlight_co in df_plot['公司'].values:
        idx = df_plot['公司'].tolist().index(highlight_co)
        fig.add_shape(
            type="rect",
            xref="x", yref="y",
            x0=idx - 0.45, x1=idx + 0.45,
            y0=0, y1=1,
            fillcolor="rgba(0,51,141,0.05)",
            line=dict(color="rgba(0,51,141,0.8)", width=1.5),
            layer="below"
        )

    fig.update_layout(
        barmode='relative',
        title=f"保险服务收入业务构成（{year}YE）",
        xaxis_title="公司",
        yaxis_title="占比（%）",
        legend=dict(
            orientation="v",
            yanchor="middle",
            y=0.5,
            xanchor="right",
            x=-0.15,
            font=dict(size=11)
        ),
        height=550,
        margin=dict(l=20, r=20, t=50, b=50),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        bargap=0.15,
        bargroupgap=0.1,
        hovermode='x unified'
    )
    return fig

# 5.6 承保利润业务构成堆叠图 - 按全局公司顺序
def create_profit_contribution_stacked_chart(df, cos, year, divisor=1, unit_label="百万元", highlight_co="无"):
    """
    绘制各公司承保利润业务构成堆叠图（按险种，显示占比）
    支持正负值堆叠（正数向上，负数向下）
    """
    field_type = "承保利润"
    prefix = f"{field_type}-"
    all_fields = df['字段名'].unique()
    business_fields = [f for f in all_fields if isinstance(f, str) and f.startswith(prefix)]
    business_fields = [f for f in business_fields if not f.endswith("合计")]

    if not business_fields:
        fig = go.Figure()
        fig.add_annotation(text=f"未找到{field_type}业务构成数据", x=0.5, y=0.5, showarrow=False, font=dict(size=16, color="red"))
        fig.update_layout(height=400)
        return fig

    df_year = df[df['报告年份'].astype(str) == str(year)]
    df_year = df_year[df_year['公司'].isin(cos)]

    pivot_data = []
    for co in cos:
        row = {'公司': co}
        for f in business_fields:
            val_series = df_year[(df_year['公司'] == co) & (df_year['字段名'] == f)]['(百万)人民币']
            val = val_series.iloc[0] if not val_series.empty else 0
            label = f.replace(prefix, "")
            row[label] = val / divisor
        pivot_data.append(row)
    df_plot = pd.DataFrame(pivot_data)
    display_labels = [c for c in df_plot.columns if c != '公司']

    if not df_plot[display_labels].map(lambda x: x != 0).any().any():
        fig = go.Figure()
        fig.add_annotation(text="所有公司业务构成均为零或未披露", x=0.5, y=0.5, showarrow=False, font=dict(size=16, color="red"))
        fig.update_layout(height=400)
        return fig

    # 按传入的 cos 顺序确保 df_plot 的顺序
    df_plot = df_plot.set_index('公司').reindex(cos).reset_index()

    # 计算每个公司的绝对值总和（用于归一化）
    abs_sums = {}
    for co in cos:
        vals = [df_plot[df_plot['公司'] == co][label].iloc[0] for label in display_labels]
        total = sum(abs(v) for v in vals)
        abs_sums[co] = total if total != 0 else 1.0

    # 归一化为百分比（保留符号）
    for co in cos:
        for label in display_labels:
            val = df_plot[df_plot['公司'] == co][label].iloc[0]
            pct = val / abs_sums[co] * 100
            df_plot.loc[df_plot['公司'] == co, label] = pct

    # ===== 唯一颜色分配 =====
    import plotly.express as px
    base_colors = KPMG_COLORS
    if len(display_labels) > len(base_colors):
        extra_colors = px.colors.qualitative.Plotly
        extra_unique = [c for c in extra_colors if c not in base_colors]
        all_colors = base_colors + extra_unique
    else:
        all_colors = base_colors
    if len(display_labels) > len(all_colors):
        all_colors = all_colors + px.colors.qualitative.Alphabet
    color_map = {label: all_colors[i % len(all_colors)] for i, label in enumerate(display_labels)}

    fig = go.Figure()

    for i, label in enumerate(display_labels):
        values = []
        for co in cos:
            val = df_plot[df_plot['公司'] == co][label].iloc[0]
            values.append(val)
        texts = []
        for idx, co in enumerate(cos):
            val = values[idx]
            if val == 0:
                texts.append("")
            else:
                texts.append(f"{val:.1f}%")
        fig.add_trace(go.Bar(
            name=label,
            x=cos,
            y=values,
            marker_color=color_map[label],
            text=texts,
            textposition='inside',
            insidetextanchor='middle',
            textfont=dict(size=10, color='white' if i % 2 == 0 else 'black'),
            hovertemplate=f"{label}: %{{y:.1f}}%<extra>%{{x}}</extra>",
            showlegend=True,
            legendgroup=label,
        ))

    fig.add_hline(y=0, line_dash="dash", line_color="gray", line_width=1, opacity=0.5)

    if highlight_co != "无" and highlight_co in cos:
        idx = cos.index(highlight_co)
        fig.add_shape(
            type="rect",
            xref="x", yref="y",
            x0=idx - 0.45, x1=idx + 0.45,
            y0=-100, y1=100,
            fillcolor="rgba(0,51,141,0.05)",
            line=dict(color="rgba(0,51,141,0.8)", width=1.5),
            layer="below"
        )

    fig.update_layout(
        barmode='relative',
        title=f"承保利润业务构成（{year}YE）",
        xaxis_title="公司",
        yaxis_title="占比（%）",
        legend=dict(
            orientation="v",
            yanchor="middle",
            y=0.5,
            xanchor="right",
            x=-0.15,
            font=dict(size=11)
        ),
        height=550,
        margin=dict(l=20, r=20, t=50, b=50),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        bargap=0.15,
        bargroupgap=0.1,
        hovermode='x unified'
    )
    return fig

#5.7 保险业务收入柱状图
def create_premium_old_chart(df, selected_cos, years, divisor=1, unit_label="百万元", highlight_co="无"):
    """
    绘制多家公司多年保险业务收入（旧准则）的分组柱状图
    横轴为公司，每个公司显示三年的柱子（年份分组）
    支持公司边框、高亮追踪、数值标签
    """
    field_name = "保险业务收入"
    df_clean = df.copy()
    df_clean['字段名'] = df_clean['字段名'].astype(str).str.strip()
    df_clean['公司'] = df_clean['公司'].astype(str).str.strip()
    df_clean['报告年份'] = df_clean['报告年份'].astype(str).str.replace('.0', '', regex=False).str.strip()
    df_clean = df_clean[df_clean['公司'].isin(selected_cos)]
    # 筛选类别为“保费收入”
    if '类别' in df_clean.columns:
        df_clean = df_clean[df_clean['类别'] == '保费收入']
    # 筛选目标字段（模糊匹配保险业务收入）
    norm_field = normalize_field(field_name)
    df_plot = df_clean[df_clean['字段名'].apply(lambda x: normalize_field(x) == norm_field)].copy()
    if df_plot.empty:
        fig = go.Figure()
        fig.add_annotation(text="未找到保险业务收入数据", x=0.5, y=0.5, showarrow=False, font=dict(size=16, color="red"))
        fig.update_layout(height=500, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
        return fig
    
    # 按年份过滤
    available_years = sorted([int(y) for y in df_plot['报告年份'].unique() if y.isdigit()])
    years = [y for y in years if y in available_years]
    if not years:
        fig = go.Figure()
        fig.add_annotation(text="指定年份无数据", x=0.5, y=0.5, showarrow=False)
        return fig
    
    val_col = '(百万)人民币' if '(百万)人民币' in df_plot.columns else df_plot.columns[-1]
    pivot_df = df_plot.pivot_table(index='公司', columns='报告年份', values=val_col, aggfunc='first').reindex(selected_cos)
    # 按最新年份排序（降序）
    latest_yr = str(max(years))
    if latest_yr in pivot_df.columns:
        pivot_df['_sort'] = pivot_df[latest_yr]
        pivot_df = pivot_df.sort_values('_sort', ascending=False).drop('_sort', axis=1)
    sorted_cos = pivot_df.index.tolist()
    
    # 颜色映射
    color_map = {str(y): KPMG_COLORS[i % len(KPMG_COLORS)] for i, y in enumerate(years)}
    if len(years) > len(KPMG_COLORS):
        import plotly.express as px
        extra = px.colors.qualitative.Plotly
        for i, y in enumerate(years):
            if y not in color_map:
                color_map[str(y)] = extra[i % len(extra)]
    
    fig = go.Figure()
    # 先计算所有数值，用于确定 y 轴范围
    all_values = []
    for yr in years:
        yr_str = str(yr)
        if yr_str in pivot_df.columns:
            vals = pivot_df[yr_str].fillna(0).tolist()
            vals = [v / divisor for v in vals]
            all_values.extend(vals)
        else:
            all_values.extend([0]*len(sorted_cos))
    
    # 确定 y 轴上限：最大值的 1.15 倍，但至少为 1（避免全为零）
    y_max = max(all_values) if all_values else 1
    y_upper = max(y_max * 1.15, 1)  # 留 15% 空间给数值标签
    
    for yr in years:
        yr_str = str(yr)
        if yr_str in pivot_df.columns:
            values = pivot_df[yr_str].fillna(0).tolist()
            values = [v / divisor for v in values]
        else:
            values = [0]*len(sorted_cos)
        texts = [f"{v:.1f}" if v != 0 else "" for v in values]
        fig.add_trace(go.Bar(
            name=f"{yr}YE",
            x=sorted_cos,
            y=values,
            marker_color=color_map[yr_str],
            text=texts,
            textposition='outside',
            textfont=dict(size=10),
            width=0.25,
            offsetgroup=yr_str
        ))
    
    # 高亮框
    if highlight_co != "无" and highlight_co in sorted_cos:
        idx = sorted_cos.index(highlight_co)
        fig.add_shape(
            type="rect",
            xref="x", yref="paper",
            x0=idx - 0.45, x1=idx + 0.45,
            y0=-0.08, y1=1,
            fillcolor="rgba(0,51,141,0.05)",
            line=dict(color="rgba(0,51,141,0.8)", width=1.5),
            layer="below"
        )
    
    # 公司边框
    fig = add_company_borders(fig, sorted_cos, sorted_cos, top_margin=70)
    fig.update_xaxes(showticklabels=False)
    
    # 纵轴范围，从 0 到 y_upper
    fig.update_layout(
        barmode='group',
        height=500,
        margin=dict(t=70, b=40, l=50, r=20),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        yaxis=dict(
            range=[0, y_upper],
            tickformat=",.0f",
            title_text=f"金额（{unit_label}）",
            zeroline=True,
            zerolinecolor='gray'
        ),
        xaxis=dict(showgrid=False, zeroline=False, showline=False),
        bargap=0.15,
        bargroupgap=0.1
    )
    return fig

#5.8保险服务收入、新旧准则比值绘图
def create_multi_year_bar_chart(df, field_name, cos, years, divisor=1, unit_label="百万元", highlight_co="无", is_percentage=False):
    """
    绘制多年份分组柱状图（横轴为公司，每个公司多个柱子）
    用于保险服务收入、新旧准则比值等字段
    """
    df_clean = df.copy()
    df_clean['字段名'] = df_clean['字段名'].astype(str).str.strip()
    df_clean['公司'] = df_clean['公司'].astype(str).str.strip()
    df_clean['报告年份'] = df_clean['报告年份'].astype(str).str.replace('.0', '', regex=False).str.strip()
    df_clean = df_clean[df_clean['公司'].isin(cos)]

    # 若存在类别，限定为保费收入
    if '类别' in df_clean.columns:
        df_clean = df_clean[df_clean['类别'] == '保费收入']

    norm_field = normalize_field(field_name)
    df_plot = df_clean[df_clean['字段名'].apply(lambda x: normalize_field(x) == norm_field)].copy()
    if df_plot.empty:
        fig = go.Figure()
        fig.add_annotation(text=f"未找到{field_name}数据", x=0.5, y=0.5, showarrow=False)
        fig.update_layout(height=500)
        return fig

    # 按年份过滤
    available_years = sorted([int(y) for y in df_plot['报告年份'].unique() if y.isdigit()])
    years = [y for y in years if y in available_years]
    if not years:
        fig = go.Figure()
        fig.add_annotation(text="指定年份无数据", x=0.5, y=0.5, showarrow=False)
        return fig

    val_col = '(百万)人民币' if '(百万)人民币' in df_plot.columns else df_plot.columns[-1]
    pivot_df = df_plot.pivot_table(index='公司', columns='报告年份', values=val_col, aggfunc='first').reindex(cos)
    # 按最新年份排序
    latest_yr = str(max(years))
    if latest_yr in pivot_df.columns:
        pivot_df['_sort'] = pivot_df[latest_yr]
        pivot_df = pivot_df.sort_values('_sort', ascending=False).drop('_sort', axis=1)
    sorted_cos = pivot_df.index.tolist()

    # 颜色映射
    color_map = {str(y): KPMG_COLORS[i % len(KPMG_COLORS)] for i, y in enumerate(years)}
    if len(years) > len(KPMG_COLORS):
        import plotly.express as px
        extra = px.colors.qualitative.Plotly
        for i, y in enumerate(years):
            if y not in color_map:
                color_map[str(y)] = extra[i % len(extra)]

    fig = go.Figure()
    all_values = []
    for yr in years:
        yr_str = str(yr)
        if yr_str in pivot_df.columns:
            vals = pivot_df[yr_str].fillna(0).tolist()
            if not is_percentage:
                vals = [v / divisor for v in vals]
            else:
                vals = [v * 100 for v in vals]  # 百分比转为百分数显示
            all_values.extend(vals)
        else:
            vals = [0] * len(sorted_cos)
            all_values.extend(vals)
        texts = [f"{v:.1f}%" if is_percentage and v != 0 else f"{v:.1f}" if v != 0 else "" for v in vals]
        fig.add_trace(go.Bar(
            name=f"{yr}YE",
            x=sorted_cos,
            y=vals,
            marker_color=color_map[yr_str],
            text=texts,
            textposition='outside',
            textfont=dict(size=10),
            width=0.25,
            offsetgroup=yr_str
        ))

    # 高亮框
    if highlight_co != "无" and highlight_co in sorted_cos:
        idx = sorted_cos.index(highlight_co)
        fig.add_shape(
            type="rect",
            xref="x", yref="paper",
            x0=idx - 0.45, x1=idx + 0.45,
            y0=-0.08, y1=1,
            fillcolor="rgba(0,51,141,0.05)",
            line=dict(color="rgba(0,51,141,0.8)", width=1.5),
            layer="below"
        )

    fig = add_company_borders(fig, sorted_cos, sorted_cos, top_margin=70)
    fig.update_xaxes(showticklabels=False)

    y_max = max(all_values) if all_values else 1
    y_upper = max(y_max * 1.15, 1)
    y_title = "百分比（%）" if is_percentage else f"金额（{unit_label}）"
    fig.update_layout(
        barmode='group',
        height=500,
        margin=dict(t=70, b=40, l=50, r=20),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        yaxis=dict(
            range=[0, y_upper],
            tickformat=",.0f" if not is_percentage else ".1f",
            title_text=y_title,
            zeroline=True,
            zerolinecolor='gray'
        ),
        xaxis=dict(showgrid=False, zeroline=False, showline=False),
        bargap=0.15,
        bargroupgap=0.1
    )
    return fig


def create_multi_year_line_chart(df, field_name, cos, years, divisor=1, unit_label="百万元", highlight_co="无", is_percentage=False, decimal_places=1):
    """
    绘制多年份折线图（横轴为年份，每条线代表一个公司）
    用于投资成分占比、保费增长率等百分比指标，或比值指标。
    decimal_places: 数值标签保留的小数位数
    """
    df_clean = df.copy()
    df_clean['字段名'] = df_clean['字段名'].astype(str).str.strip()
    df_clean['公司'] = df_clean['公司'].astype(str).str.strip()
    df_clean['报告年份'] = df_clean['报告年份'].astype(str).str.replace('.0', '', regex=False).str.strip()
    df_clean = df_clean[df_clean['公司'].isin(cos)]
    if '类别' in df_clean.columns:
        df_clean = df_clean[df_clean['类别'] == '保费收入']

    norm_field = normalize_field(field_name)
    df_plot = df_clean[df_clean['字段名'].apply(lambda x: normalize_field(x) == norm_field)].copy()
    if df_plot.empty:
        fig = go.Figure()
        fig.add_annotation(text=f"未找到{field_name}数据", x=0.5, y=0.5, showarrow=False)
        fig.update_layout(height=400)
        return fig

    available_years = sorted([int(y) for y in df_plot['报告年份'].unique() if y.isdigit()])
    years = [y for y in years if y in available_years]
    if not years:
        fig = go.Figure()
        fig.add_annotation(text="指定年份无数据", x=0.5, y=0.5, showarrow=False)
        return fig

    val_col = '(百万)人民币' if '(百万)人民币' in df_plot.columns else df_plot.columns[-1]
    fig = go.Figure()
    color_map = {co: KPMG_COLORS[i % len(KPMG_COLORS)] for i, co in enumerate(cos)}
    latest_yr = max(years)
    sort_df = df_plot[df_plot['报告年份'] == str(latest_yr)].set_index('公司')[val_col]
    sorted_cos = sorted(cos, key=lambda x: sort_df.get(x, 0), reverse=True)

    for co in sorted_cos:
        co_data = df_plot[df_plot['公司'] == co]
        co_values = []
        for yr in years:
            yr_str = str(yr)
            val = co_data[co_data['报告年份'] == yr_str][val_col]
            v = val.iloc[0] if not val.empty else np.nan
            if is_percentage:
                v = v * 100 if pd.notna(v) else np.nan
            else:
                v = v / divisor if pd.notna(v) else np.nan
            co_values.append(v)
        valid_pairs = [(yr, v) for yr, v in zip(years, co_values) if pd.notna(v)]
        if valid_pairs:
            x_vals = [yr for yr, _ in valid_pairs]
            y_vals = [v for _, v in valid_pairs]
            fig.add_trace(go.Scatter(
                x=x_vals,
                y=y_vals,
                mode='lines+markers+text',
                name=co,
                line=dict(width=2, color=color_map[co]),
                marker=dict(size=8, color=color_map[co]),
                text=[f"{v:.{decimal_places}f}{'%' if is_percentage else ''}" for v in y_vals],
                textposition="top center",
                textfont=dict(size=10)
            ))

    if highlight_co != "无" and highlight_co in cos:
        for trace in fig.data:
            if trace.name == highlight_co:
                trace.line.width = 4
                trace.marker.size = 10

    y_min = min([min(y) for y in [trace.y for trace in fig.data] if y]) if fig.data else 0
    y_max = max([max(y) for y in [trace.y for trace in fig.data] if y]) if fig.data else 1
    padding = (y_max - y_min) * 0.15 if y_max != y_min else 5
    y_range = [max(0, y_min - padding), y_max + padding]

    fig.update_layout(
        height=450,
        margin=dict(t=40, b=50, l=50, r=20),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        xaxis=dict(
            title="年份",
            tickvals=years,
            ticktext=[f"{y}YE" for y in years],
            showgrid=False,
            zeroline=False
        ),
        yaxis=dict(
            range=y_range,
            tickformat="." + str(decimal_places) + "f",
            title_text="百分比（%）" if is_percentage else f"{unit_label}",
            zeroline=True,
            zerolinecolor='gray'
        ),
        hovermode='x unified'
    )
    return fig

#5.9 投资成分占比、保费增长率折线图
def create_line_subplots_chart(df, field_name, cos, years, divisor=1, unit_label="", highlight_co="无", is_percentage=False, decimal_places=1):
    """
    为每家公司生成独立的折线子图，横轴为年份，显示该公司的趋势。
    适用于投资成分占比、保费增长率等指标。
    """
    df_clean = df.copy()
    df_clean['字段名'] = df_clean['字段名'].astype(str).str.strip()
    df_clean['公司'] = df_clean['公司'].astype(str).str.strip()
    df_clean['报告年份'] = df_clean['报告年份'].astype(str).str.replace('.0', '', regex=False).str.strip()
    df_clean = df_clean[df_clean['公司'].isin(cos)]
    if '类别' in df_clean.columns:
        df_clean = df_clean[df_clean['类别'] == '保费收入']

    norm_field = normalize_field(field_name)
    df_plot = df_clean[df_clean['字段名'].apply(lambda x: normalize_field(x) == norm_field)].copy()
    if df_plot.empty:
        fig = go.Figure()
        fig.add_annotation(text=f"未找到{field_name}数据", x=0.5, y=0.5, showarrow=False)
        fig.update_layout(height=400)
        return fig

    available_years = sorted([int(y) for y in df_plot['报告年份'].unique() if y.isdigit()])
    years = [y for y in years if y in available_years]
    if not years:
        fig = go.Figure()
        fig.add_annotation(text="指定年份无数据", x=0.5, y=0.5, showarrow=False)
        return fig

    val_col = '(百万)人民币' if '(百万)人民币' in df_plot.columns else df_plot.columns[-1]
    # 构建数据字典
    data = {}
    for co in cos:
        co_data = df_plot[df_plot['公司'] == co]
        data[co] = {}
        for yr in years:
            yr_str = str(yr)
            val = co_data[co_data['报告年份'] == yr_str][val_col]
            v = val.iloc[0] if not val.empty else np.nan
            if is_percentage:
                v = v * 100 if pd.notna(v) else np.nan
            else:
                v = v / divisor if pd.notna(v) else np.nan
            data[co][yr] = v

    # 过滤掉所有年份数据都为nan的公司
    valid_cos = [co for co in cos if any(not np.isnan(data[co][yr]) for yr in years)]
    if not valid_cos:
        fig = go.Figure()
        fig.add_annotation(text="所有公司数据均为空", x=0.5, y=0.5, showarrow=False)
        return fig

    n = len(valid_cos)
    fig = make_subplots(rows=1, cols=n, shared_yaxes=True,
                        subplot_titles=[f"<b>{co}</b>" for co in valid_cos],
                        horizontal_spacing=0.02 if n > 1 else 0)

    # 颜色
    color_map = {co: KPMG_COLORS[i % len(KPMG_COLORS)] for i, co in enumerate(valid_cos)}

    for col_idx, co in enumerate(valid_cos):
        col = col_idx + 1
        x_vals = []
        y_vals = []
        for yr in years:
            v = data[co][yr]
            if not np.isnan(v):
                x_vals.append(yr)
                y_vals.append(v)
        if x_vals:
            fig.add_trace(go.Scatter(
                x=x_vals,
                y=y_vals,
                mode='lines+markers+text',
                name=co,  # 图例可能不显示，但保留
                line=dict(width=2, color=color_map[co]),
                marker=dict(size=8, color=color_map[co]),
                text=[f"{v:.{decimal_places}f}{'%' if is_percentage else ''}" for v in y_vals],
                textposition="top center",
                textfont=dict(size=10),
                showlegend=False  # 每个子图单独显示，不叠加图例
            ), row=1, col=col)

        # 添加灰色边框
        fig.add_shape(
            type="rect",
            xref="x domain", yref="y domain",
            x0=0, x1=1,
            y0=0, y1=1,
            fillcolor="rgba(200,200,200,0.05)",
            line=dict(color="#CCCCCC", width=1.2),
            layer="below",
            row=1, col=col
        )
        
        # 高亮边框
        if highlight_co != "无" and co == highlight_co:
            fig.add_shape(
                type="rect",
                xref="x domain", yref="y domain",
                x0=-0.05, x1=1.05, y0=-0.1, y1=1.1,
                fillcolor="rgba(0,51,141,0.05)",
                line=dict(color="rgba(0,51,141,0.8)", width=1.5),
                layer="above",
                row=1, col=col
            )

    # 调整布局
    fig.update_layout(
        height=450,
        margin=dict(t=60, b=40, l=40, r=20),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        bargap=0.15,
        bargroupgap=0.1
    )

    # 计算统一的 y 轴范围（所有子图共享）
    all_vals = [v for co in valid_cos for yr in years for v in [data[co][yr]] if not np.isnan(v)]
    if all_vals:
        y_min = min(all_vals)
        y_max = max(all_vals)
        padding = (y_max - y_min) * 0.15 if y_max != y_min else 5
        y_range = [max(0, y_min - padding), y_max + padding]
    else:
        y_range = [0, 1]

    fig.update_yaxes(range=y_range, tickformat=f".{decimal_places}f", title_text="百分比（%）" if is_percentage else unit_label, row=1, col=1)
    for i in range(1, n+1):
        fig.update_xaxes(tickvals=years, ticktext=[f"{y}YE" for y in years], row=1, col=i)

    # 调整子图标题位置
    for ann in fig.layout.annotations:
        if ann.text and "<b>" in ann.text:
            ann.update(y=1.02)

    return fig

# 5.10 折现率、非金融风险调整的表格
def create_disclosure_table(df, field_name, title, cos, years):
    """
    生成文本披露（如折现率、非金融风险调整）的表格
    df: 集成数据框，需包含 '公司', '报告年份', '字段名', '(百万)人民币' 等列
    field_name: 要展示的字段名（如 '折现率假设'）
    title: 表格标题
    cos: 公司列表（按显示顺序）
    years: 要显示的年份列表（如 [2024, 2025]）
    """
    # 筛选数据
    df_sub = df[df['字段名'] == field_name].copy()
    if df_sub.empty:
        fig = go.Figure()
        fig.add_annotation(text="暂无数据", x=0.5, y=0.5, showarrow=False)
        fig.update_layout(height=300)
        return fig

    # 透视：公司 × 年份 -> 值（文本）
    pivot = df_sub.pivot_table(index='公司', columns='报告年份', values='(百万)人民币', aggfunc='first')
    # 只保留指定的公司和年份
    pivot = pivot.reindex(index=cos, columns=[str(y) for y in years])
    # 将年份列名排序
    pivot = pivot[sorted(pivot.columns, key=lambda x: int(x))]

    # 准备表格数据
    header_vals = ['公司'] + [f'{y}年12月31日' for y in pivot.columns]
    cell_vals = [pivot.index.tolist()]
    for col in pivot.columns:
        cell_vals.append(pivot[col].fillna('未披露').tolist())

    fig = go.Figure(data=[go.Table(
        header=dict(
            values=header_vals,
            fill_color='#00338D',
            align='center',
            font=dict(color='white', size=13)
        ),
        cells=dict(
            values=cell_vals,
            fill_color=[['#F8F9FA', 'white'] * (len(pivot)//2 + 1)],
            align='center',
            font=dict(size=12)
        )
    )])
    fig.update_layout(
        title=title,
        margin=dict(l=10, r=10, t=50, b=10),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)'
    )
    return fig


# 5.11 非金融风险调整表格
def create_risk_margin_table(df, cos, title="非金融风险调整披露"):
    """
    生成非金融风险调整表格，展示每个公司的“方法”和“置信水平”（取最新年份）
    df: 集成数据，需包含 '公司', '字段名', '(百万)人民币' 等列
    cos: 公司列表（按显示顺序）
    title: 表格标题
    返回: Plotly 表格对象
    """
    field_name = '非金融风险调整'
    df_sub = df[df['字段名'] == field_name].copy()
    if df_sub.empty:
        fig = go.Figure()
        fig.add_annotation(text="暂无数据", x=0.5, y=0.5, showarrow=False)
        fig.update_layout(height=300)
        return fig

    # 取每个公司最新年份的数据（如果有多个年份，取最大的）
    df_sub['报告年份'] = df_sub['报告年份'].astype(int)
    df_sub = df_sub.sort_values(['公司', '报告年份'], ascending=[True, False])
    df_sub = df_sub.drop_duplicates(subset=['公司'], keep='first')
    df_sub = df_sub.set_index('公司').reindex(cos).reset_index()

    # 提取值并解析方法和置信水平
    values = df_sub['(百万)人民币'].fillna('').astype(str).tolist()
    methods = []
    confidences = []
    import re

    for val in values:
        if not val or val == 'nan' or val == '':
            methods.append('')
            confidences.append('')
            continue

        # ---- 1. 提取置信水平（百分比或范围） ----
        confidence_match = re.search(r'(\d+(?:\.\d+)?%\s*(?:[-~]\s*\d+(?:\.\d+)?%)?)', val)
        if confidence_match:
            confidence = confidence_match.group(1).strip()
            confidences.append(confidence)
        else:
            # 尝试提取纯数字（可能不带%）
            num_match = re.search(r'(\d+(?:\.\d+)?)', val)
            if num_match:
                confidences.append(num_match.group(1) + '%')
            else:
                confidences.append('未披露')

        # ---- 2. 提取方法：按逗号分割取第一部分 ----
        # 先尝试按逗号（中文或英文）分割
        parts = re.split(r'[，,]', val)
        if len(parts) >= 2:
            # 第一部分作为方法
            method = parts[0].strip()
            # 移除方法末尾可能多余的分隔符（如"、"）
            method = re.sub(r'[、，]\s*$', '', method)
            methods.append(method if method else '未披露')
        else:
            # 没有逗号，尝试提取含"法"的部分
            method_match = re.search(r'([^，,]+?(?:法|方法))', val)
            if method_match:
                methods.append(method_match.group(1).strip())
            else:
                # 如果都没有，则整体作为方法（不含置信水平）
                methods.append(val.strip())
                
    # 构建 Plotly 表格
    fig = go.Figure(data=[go.Table(
        header=dict(
            values=['公司名称', '方法', '置信水平'],
            fill_color='#00338D',
            align='center',
            font=dict(color='white', size=13)
        ),
        cells=dict(
            values=[cos, methods, confidences],
            fill_color=[['#F8F9FA', 'white'] * (len(cos)//2 + 1)],
            align='center',
            font=dict(size=12)
        )
    )])
    fig.update_layout(
        title=title,
        margin=dict(l=10, r=10, t=50, b=10),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)'
    )
    return fig
    
# 6.汇总表
def create_nonlife_summary_table(df, cos, highlight_co="无"):
    cy, py = latest_year, prev_year
    col_names, c1, c2, c3, c4, c5, c6, c7, c8, c9, c10 = [], [], [], [], [], [], [], [], [], [], []
    
    for co in cos:
        df_co = df[df['公司'] == co]

        def get_val(y, kw):
            s = df_co[(df_co['报告年份'].astype(str) == str(y)) & 
                      (df_co['字段名'].astype(str).str.contains(kw, na=False))].drop_duplicates(
                          subset=['公司','报告年份','字段名'])['(百万)人民币']
            if s.empty: return None
            v = s.sum(min_count=1)
            return None if pd.isna(v) else v

        def calc(n, d, is_growth=False, is_ratio=False):
            if n is None or d is None: return "未披露"
            if d == 0: return "-"
            if is_ratio:
                return f"{n/d:.1%}"
            ratio = (n / d) - 1 if is_growth else (n / d)
            if is_growth:
                is_positive = n > d
            else:
                is_positive = ratio >= 0
            return f"{abs(ratio):.0%}" if is_positive else f"-{abs(ratio):.0%}"

        prem_cy = get_val(cy, '保险业务收入') or get_val(cy, '已赚保费')
        prem_py = get_val(py, '保险业务收入') or get_val(py, '已赚保费')
        
        invest_comp_cy = get_val(cy, '投资成分')
        inv_ratio_cy = invest_comp_cy / prem_cy if (invest_comp_cy is not None and prem_cy) else None

        claim_cy = get_val(cy, '已发生赔款') or get_val(cy, '赔付支出')
        comm_cy = get_val(cy, '手续费及佣金支出')
        adm_cy  = get_val(cy, '业务及管理费')
        if prem_cy is None or prem_cy == 0:
            cor_cy = None
        else:
            total_cost = (claim_cy or 0) + (comm_cy or 0) + (adm_cy or 0)
            cor_cy = total_cost / prem_cy

        loss_ratio_cy = claim_cy / prem_cy if (claim_cy is not None and prem_cy) else None
        exp_ratio_cy = ((comm_cy or 0) + (adm_cy or 0)) / prem_cy if prem_cy else None

        current_claim_cy = get_val(cy, '当期发生赔款') or get_val(cy, '当期赔款及理赔费用')
        curr_claim_ratio = current_claim_cy / prem_cy if (current_claim_cy is not None and prem_cy) else None

        comm_ratio = comm_cy / prem_cy if (comm_cy is not None and prem_cy) else None

        loss_comp_cy = get_val(cy, '亏损部分') or get_val(cy, '亏损合同')
        lrc_nonloss_cy = get_val(cy, '未到期责任负债-非亏损部分') or get_val(cy, 'LRC非亏损部分')
        if loss_comp_cy is not None and lrc_nonloss_cy is not None:
            loss_comp_ratio_cy = loss_comp_cy / (loss_comp_cy + lrc_nonloss_cy)
        else:
            loss_comp_ratio_cy = None

        back_dev_cy = get_val(cy, '回溯偏差') or get_val(cy, '已发生赔款负债回溯偏差')
        back_dev_ratio = back_dev_cy / claim_cy if (back_dev_cy is not None and claim_cy) else None

        col_names.append(co)
        c1.append(prem_cy)
        c2.append(calc(prem_cy, prem_py, True))
        c3.append(calc(inv_ratio_cy, 1, False, True) if inv_ratio_cy is not None else "未披露")
        c4.append(calc(cor_cy, 1, False, True) if cor_cy is not None else "未披露")
        c5.append(calc(loss_ratio_cy, 1, False, True) if loss_ratio_cy is not None else "未披露")
        c6.append(calc(exp_ratio_cy, 1, False, True) if exp_ratio_cy is not None else "未披露")
        c7.append(calc(curr_claim_ratio, 1, False, True) if curr_claim_ratio is not None else "未披露")
        c8.append(calc(comm_ratio, 1, False, True) if comm_ratio is not None else "未披露")
        c9.append(calc(loss_comp_ratio_cy, 1, False, True) if loss_comp_ratio_cy is not None else "未披露")
        c10.append(calc(back_dev_ratio, 1, False, True) if back_dev_ratio is not None else "未披露")

    headers = [
        "公司名称", "保险业务收入", "保费增长率<br>(新旧准则)",
        "投资成分占比", "综合成本率<br>(COR)", "综合赔付率",
        "综合费用率", "当期赔款/理赔费用<br>占比", "手续费占比",
        "亏损成分占比", "回溯偏差"
    ]
    
    current_hl = str(highlight_co).strip()
    html = "<table style='width:100%; border-collapse:collapse; font-family:sans-serif; font-size:10px; margin-bottom:15px;'>"
    html += "<tr style='background-color:#00338D; color:white; text-align:center; font-weight:bold;'>"
    for idx, h in enumerate(headers):
        html += f"<th style='padding:4px 2px; {'text-align:left;' if idx==0 else 'text-align:center;'} border:1.5px solid white;'>{h}</th>"
    html += "</tr>"

    for i, co in enumerate(col_names):
        row_vals = [c1[i], c2[i], c3[i], c4[i], c5[i], c6[i], c7[i], c8[i], c9[i], c10[i]]
        is_hl = (str(co).strip() == current_hl)
        bg = HL_BOX_FILL if is_hl else ("white" if i % 2 == 0 else "#F8F9FA")
        base = (f"background-color:{bg}; padding:3px; font-size:10px; "
                + ("border-top:1.5px solid #00338D; border-bottom:1.5px solid #00338D; font-weight:bold;"
                   if is_hl else "border:1px solid #EAEAEA;"))
        s_first = base + f"text-align:left; color:#333333; {'border-left:1.5px solid #00338D;' if is_hl else ''}"
        s_mid   = base + f"text-align:center; {'color:#00338D; border-left:none; border-right:none;' if is_hl else 'color:#444444;'}"
        s_last  = base + f"text-align:center; {'color:#00338D; border-right:1.5px solid #00338D; border-left:none;' if is_hl else 'color:#444444;'}"

        html += f"<tr><td style='{s_first}'>{co}</td>"
        for idx, v in enumerate(row_vals):
            style = s_last if idx == len(row_vals)-1 else s_mid
            if v == "未披露":
                cell_style = style.replace(f"background-color:{bg};", "background-color:#CDCDCD;") \
                                  .replace("color:#444444;", "color:white;") \
                                  .replace("color:#00338D;", "color:white;") 
                html += f"<td style='{cell_style}'>未披露</td>"
            else:
                html += f"<td style='{style}'>{v}</td>"
        html += "</tr>"

    return html + "</table>"

# 7.主页面函数 show_step_7_content
def show_step_7_content():
    global notes_dict, ordered_modules
    # ----- 样式与前端 -----
    st.markdown("""
    <style>
    [data-testid="stSidebar"] { background: rgba(255,255,255,0.95) !important; border-right: 1px solid #EAEAEA !important; box-shadow: 2px 0px 15px rgba(0,0,0,0.08) !important; }
    .nav-floating-sign { position: fixed; left: 0; top: 50%; transform: translateY(-50%); background: rgba(0, 51, 141, 0.85); color: white; padding: 20px 8px; border-radius: 0 12px 12px 0; writing-mode: vertical-rl; text-orientation: mixed; font-size: 22px; font-weight: bold; letter-spacing: 3px; z-index: 9999; cursor: pointer; box-shadow: 3px 3px 12px rgba(0,0,0,0.25); transition: all 0.2s; }
    .nav-floating-sign:hover { background: rgba(0, 51, 141, 1); padding-left: 15px; }
    .stPlotlyChart { width: 100% !important; min-width: 0 !important; }
    .print-only { display: none !important; }
    .cover-page { position: relative !important; width: 338.67mm !important; height: 190.5mm !important; margin: 0 !important; padding: 0 !important; page-break-after: always !important; overflow: hidden !important; background: transparent !important; }
    .cover-page img { width: 100% !important; height: 100% !important; object-fit: cover !important; display: block !important; }
    .block-container { padding-top:0 !important; padding-right:10px !important; padding-left:10px !important; margin-top:0 !important; }
    .cover-text { forced-color-adjust: none !important; -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; color: white !important; -webkit-text-fill-color: white !important; }
    .element-container:first-child{ margin-top:0 !important; padding-top:0 !important; }
    /* 新增：保证每个模块整体不跨页 */
    .page-break-container { 
        page-break-inside: avoid; 
        break-inside: avoid; 
        margin: 0 !important; 
        padding: 0 !important; 
    }
    .page-break-title { 
        page-break-before: always; 
        break-before: page; 
        padding-top: 10px !important; 
        margin-top: 0 !important; 
        text-align: left !important; 
    }
    @media print {
        /* 页面设置：保留动态脚本覆盖，但默认边距稍大，避免内容溢出 */
        @page { 
            size: A4 portrait; 
            margin: 15mm 15mm 15mm 15mm; 
            @bottom-center { 
                content: counter(page); 
                font-size: 12px; 
                color: #666; 
                font-family: Microsoft YaHei, sans-serif; 
            } 
        }
        .print-only { display: block !important; }
        /* 移除固定宽高限制，允许内容自然展开，避免截断 */
        html, body { 
            width: 100% !important; 
            height: auto !important; 
            overflow: visible !important; 
            zoom: 100% !important; 
        }
        .main .block-container { 
            width: 100% !important; 
            max-width: 100% !important; 
            min-width: 0 !important; 
            padding: 0 10mm !important; 
        }
        .block-container { padding-top: 0rem !important; }
        .no-print, h1, .nav-floating-sign, [data-testid="collapsedControl"], header, footer, [data-testid="stHeader"], [data-testid="stSidebar"], section[data-testid="stSidebar"], [data-testid="stToolbar"], button[kind="secondary"], input, .stSlider, [data-testid="stSelectbox"], [data-testid="stRadio"], [data-testid="stExpander"], .stAlert, button[role="tab"], div[role="tablist"], [data-baseweb="tab-list"], hr { display: none !important; }
        /* 每个模块容器自动分页，内部不拆分 */
        .page-break-container { 
            page-break-inside: avoid !important; 
            break-inside: avoid !important; 
            margin: 0 !important; 
            padding: 0 !important; 
            padding-bottom: 5mm !important; 
        }
        .stApp { max-width: 100% !important; width: 100% !important; }
        .keep-columns [data-testid="stHorizontalBlock"]{ display:flex!important; flex-wrap:nowrap!important; align-items:flex-start!important; justify-content:space-between!important; gap:0!important; width:100%!important; }
        .keep-columns [data-testid="stHorizontalBlock"]>div{ width:49%!important; min-width:49%!important; max-width:49%!important; flex:0 0 49%!important; overflow:hidden!important; page-break-inside:avoid!important; break-inside:avoid!important; }
        /* 标题强制换页 */
        .page-break-title { 
            page-break-before: always !important; 
            break-before: page !important; 
            padding-top: 10px !important; 
            margin-top: 0 !important; 
            text-align: left !important; 
        }
        h2 { display: block !important; text-align: left !important; color: #00338D !important; font-size: 30px !important; font-weight: bold !important; border-bottom: 2px solid #00338D !important; padding-bottom: 6px !important; margin: 14px 0 10px 0 !important; }
        h3:not(.no-print) { display: block !important; text-align: left !important; color: #00338D !important; font-size: 30px !important; font-weight: bold !important; margin: 10px 0 8px 0 !important; page-break-after: avoid !important; }
        /* 图表自适应，不溢出，强制避免分页 */
        .plotly-graph-div,
        .stPlotlyChart {
            width: 100% !important;
            max-width: 100% !important;
            page-break-inside: avoid !important;
            break-inside: avoid !important;
            display: block !important;
        }
        div[data-testid="stDataFrame"], div[data-testid="stTable"] { zoom: 0.65 !important; margin: 0 auto 20px auto !important; max-width: 100% !important; page-break-inside: auto !important; }
        div[data-testid="stTable"] tr { page-break-inside: avoid !important; }
        .element-container { 
            page-break-inside: avoid !important; 
            break-inside: avoid !important; 
            width: 100% !important; 
        }
        .pdf-page-break { break-before: page !important; page-break-before: always !important; height: 0 !important; margin: 0 !important; padding: 0 !important; }
        table { page-break-inside: auto !important; }
        tr { page-break-inside: avoid !important; page-break-after: auto !important; }
        td, th { page-break-inside: avoid !important; }
        thead { display: table-header-group !important; }
    }
    @media print and (orientation: portrait) { .stPlotlyChart { margin-bottom: 10mm !important; } }
    @media print and (orientation: landscape) { .stPlotlyChart { margin-bottom: 6mm !important; } }
    .stPlotlyChart, div[data-testid="stDataFrame"] { display: flex !important; justify-content: center !important; }
    .highlight-blue-box { border: 1.5px solid rgba(0,51,141,0.85) !important; border-radius: 12px !important; padding: 10px !important; background: rgba(0,51,141,0.02) !important; box-shadow: 0px 4px 12px rgba(0,51,141,0.12) !important; margin-bottom: 25px !important; }
    </style>
    <div class="nav-floating-sign" id="custom-nav-trigger">展开导航栏</div>
    """, unsafe_allow_html=True)

    components.html("""<script>let t = setInterval(() => { const d = window.parent.document; const b = d.getElementById("custom-nav-trigger"); const c = d.querySelector('[data-testid="collapsedControl"]') || d.querySelector('button[kind="header"]'); if(b && c) { b.onclick = () => c.click(); clearInterval(t); } }, 500);</script>""", height=0, width=0)
    # ==========================================
    # 加载注释表（导航配置）
    # ==========================================

    st.markdown("<div class='no-print'>", unsafe_allow_html=True)
    with st.expander("📥 公司内容分析与注释输入", expanded=False):
        st.markdown("""
            <div style='background:linear-gradient(135deg,#00338D,#0865EE); 
            border-radius:10px; padding:14px 18px; margin-bottom:16px;
            display:flex; align-items:center; justify-content:space-between;'>
                <span style='color:white; font-size:15px; font-weight:bold;'>
                    获取官方注释表模板
                </span>
                <span style='color:rgba(255,255,255,0.8); font-size:12px;'>
                    需要安全码验证
                </span>
            </div>
        """, unsafe_allow_html=True)
    
        col_pwd, col_btn = st.columns([3, 1])
        with col_pwd:
            pwd_input = st.text_input("", placeholder="请输入安全码...", 
                                       type="password", key="template_pwd",
                                       label_visibility="collapsed")
        with col_btn:
            check_btn = st.button("🔓验证下载", use_container_width=True)
    
        if check_btn:
            if 'filled_notes_excel' in st.session_state:
                del st.session_state['filled_notes_excel']
                
            if pwd_input == "KPMG666":  # ✅ 保持与登录密码一致
                try:
                    import requests
                    from io import BytesIO
                    import openpyxl
        
                    # 🔧 修改点：替换为财险注释表模板的 URL
                    url = "https://github.com/z-xylym/my-actuary-tool/raw/refs/heads/main/step7-%E8%B4%A2%E9%99%A9%E6%A0%87%E6%B3%A8%E8%A1%A8_0806.xlsx"
                    r = requests.get(url, timeout=15)
        
                    if r.status_code == 200:
                        wb = openpyxl.load_workbook(BytesIO(r.content))
                        ws = wb.active
                        header = {cell.value: cell.column for cell in ws[1]}
                        mid_col = header.get('模块ID')
                        custom_col = header.get('分析内容-自定义')
        
                        if mid_col and custom_col and 'integrated_data' in st.session_state:
                            df_for_gen = st.session_state['integrated_data'].copy()
                            _cos = st.session_state.get('selected_cos_cache', [])
                            _valid_years = sorted([y for y in df_for_gen['报告年份'].dropna().astype(str).str.replace(".0","",regex=False).unique() if y.isdigit()])
                            _cy = int(_valid_years[-1]) if _valid_years else 2025
                            _py = int(_valid_years[-2]) if len(_valid_years) > 1 else 2024
        
                            # 遍历 Excel 行，填充分析内容（如果有自定义生成逻辑）
                            for row in ws.iter_rows(min_row=2):
                                m_id_val = row[mid_col - 1].value
                                if not m_id_val: continue
                                # 如果有自动生成话术的需求，可在这里调用
                                # 目前财险系统中暂无 generate_custom_analysis，所以跳过自动填充
                                # 后续如需自动填充，可在此添加逻辑
                                pass
        
                        output = BytesIO()
                        wb.save(output)
                        output.seek(0)
                        
                        st.session_state['filled_notes_excel'] = output.getvalue()
                        st.success("✅ 验证成功，请点击下方按钮下载模板")
                    else:
                        st.error("❌ 文件获取失败")
                except Exception as e:
                    st.error(f"❌ 错误：{e}")
            else:
                st.error("❌ 安全码错误")
        
        # 下载按钮
        if 'filled_notes_excel' in st.session_state:
            st.download_button(
                label="🔓点击下载注释表模板",
                data=st.session_state['filled_notes_excel'],
                file_name=f"注释表_{st.session_state.get('selected_cos_cache', [''])[0] if st.session_state.get('selected_cos_cache') else ''}_{2025}年.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
        st.divider()        
        use_default = st.toggle("使用默认注释表", value=True, key="use_default_notes")
        df_notes = None
        if use_default:
            try:
                # 🔧 修改点：替换为财险注释表模板的 URL
                df_notes = pd.read_excel("https://github.com/Polly031021/Annual-Report-System/raw/main/step7-%E8%B4%A2%E9%99%A9%E6%A0%87%E5%87%86%E6%B3%A8%E9%87%8A%E8%A1%A8_0824.xlsx")
                st.success("✅ 内置默认注释表加载成功")
            except Exception as e:
                st.error(f"❌ 加载失败：{e}")
        else:
            st.info("💡 请上传包含【图片文件名】【模块ID】【一级分类】【二级分类】【对应图表名称】【分析内容】【注释内容】的 Excel")
            notes_file = st.file_uploader("上传 Excel", type=['xlsx', 'xls'])
            if notes_file: 
                df_notes = pd.read_excel(notes_file)

        # ---- 解析注释表 ----
        if df_notes is not None:
            # 统一清洗文本前后的空格
            for col in ['一级分类', '二级分类', '对应图表名称', '模块ID']:
                if col in df_notes.columns:
                    df_notes[col] = df_notes[col].astype(str).str.strip().replace(['nan', 'NaN', 'NAN', 'None'], '')
            
            # 如果二级分类为空白，强制塞入"全部"
            if '二级分类' in df_notes.columns:
                df_notes['二级分类'] = df_notes['二级分类'].apply(lambda x: "全部" if str(x).strip() == "" else str(x).strip())
            
            # 填充 notes_dict 和 ordered_modules
            has_img = '图片文件名' in df_notes.columns
            for _, r in df_notes.iterrows():
                m_id = str(r.get('模块ID', '')).strip()
                if not m_id: 
                    continue
                img_val = str(r.get('图片文件名', '')).strip() if has_img and pd.notna(r.get('图片文件名')) else ''
                notes_dict[m_id] = {
                    'title': str(r.get('对应图表名称', '')).strip(),
                    'analysis_default': str(r.get('分析内容-默认', '')).strip() if pd.notna(r.get('分析内容-默认')) else '',
                    'analysis_custom': str(r.get('分析内容-自定义', '')).strip() if pd.notna(r.get('分析内容-自定义')) else '',
                    'note': str(r.get('注释内容', '')).strip() if pd.notna(r.get('注释内容')) else '',
                    'image_file': img_val if img_val.lower() != 'nan' else ''
                }
                if m_id not in ordered_modules:
                    ordered_modules.append(m_id)
            
            # 存入 session_state
            st.session_state['df_notes'] = df_notes
    
    st.markdown("</div>", unsafe_allow_html=True)
    
    # ----- 数据检查与年份提取 -----
    if 'integrated_data' not in st.session_state or st.session_state['integrated_data'] is None:
        # 判断用户角色
        if st.session_state.get('user_role') == "普通用户":
            st.info("📂 请上传已集成好的数据文件（Excel 或 CSV），格式需包含 '公司'、'报告年份'、'字段名'、'(百万)人民币' 等列。")
            
            # 提供示例模板
            with st.expander("📄 查看所需数据格式示例"):
                sample_df = pd.DataFrame({
                    "公司": ["示例公司A", "示例公司A"],
                    "报告年份": [2024, 2025],
                    "字段名": ["总资产", "总资产"],
                    "(百万)人民币": [100000, 110000]
                })
                st.dataframe(sample_df)
                st.caption("请确保你的数据包含以上四列，且每一行对应一个公司、一个年份、一个指标。")
            
            uploaded_file = st.file_uploader("上传集成后的数据表", type=["xlsx", "csv"], key="step7_upload")
            if uploaded_file is not None:
                try:
                    if uploaded_file.name.endswith('.csv'):
                        df = pd.read_csv(uploaded_file)
                    else:
                        df = pd.read_excel(uploaded_file)
                    # 检查必需列
                    required_cols = ['公司', '报告年份', '字段名', '(百万)人民币']
                    if all(col in df.columns for col in required_cols):
                        # 将年份列转换为字符串并去除 .0
                        df['报告年份'] = df['报告年份'].astype(str).str.replace('.0', '', regex=False)
                        st.session_state['integrated_data'] = df
                        st.success("✅ 数据加载成功！正在刷新...")
                        st.rerun()
                    else:
                        st.error(f"❌ 数据格式不正确，必须包含列：{required_cols}，你的数据列有：{df.columns.tolist()}")
                except Exception as e:
                    st.error(f"❌ 读取文件失败：{e}")
            # 阻止继续执行报告渲染（因为数据还不存在）
            st.stop()
        else:
            # 项目组成员：提示去 Step 6
            st.warning("⚠️ 请先在 Step 6 完成数据集成。")
            st.stop()
    df_raw = st.session_state['integrated_data'].copy()
    valid_years = sorted([y for y in df_raw['报告年份'].dropna().astype(str).str.replace(".0", "", regex=False).unique() if y.isdigit()])
    latest_year = int(valid_years[-1]) if valid_years else 2025
    prev_year = int(valid_years[-2]) if len(valid_years) > 1 else 2023

    # ----- 侧边栏导航 -----
    print_mode, active_m_id = False, None
    with st.sidebar:
        st.markdown("<h3 style='color: #00338D; font-size: 18px;'>公司报告导航</h3>", unsafe_allow_html=True)
        if 'df_notes' in st.session_state and st.session_state['df_notes'] is not None and not st.session_state['df_notes'].empty:
            df_n = st.session_state['df_notes']
            for _, r in df_n.iterrows():
                m_id = str(r.get('模块ID', '')).strip()
                if m_id:
                    notes_dict[m_id] = {
                        'title': str(r.get('对应图表名称', '')).strip(),
                        'analysis_default': str(r.get('分析内容-默认', '')).strip() if pd.notna(r.get('分析内容-默认')) else '',
                        'analysis_custom': str(r.get('分析内容-自定义', '')).strip() if pd.notna(r.get('分析内容-自定义')) else '',
                        'note': str(r.get('注释内容', '')).strip() if pd.notna(r.get('注释内容')) else '',
                        'image_file': str(r.get('图片文件名', '')).strip() if pd.notna(r.get('图片文件名')) else ''
                    }
                    if m_id not in ordered_modules:
                        ordered_modules.append(m_id)
            first_levels = [x for x in df_n['一级分类'].unique() if str(x).strip() != ""]
            main_nav = st.radio("📁 一级模块", first_levels + ["🖨️ 一键显示全部 (打印/导出)"], key="s7_m")

            if main_nav == "🖨️ 一键显示全部 (打印/导出)":
                print_mode = True
                if not st.session_state.get("pdf_export_mode", False):
                    st.info('竖版适合文字多的页面，横版适合宽图表。勾选"背景图形"以保留颜色。')
                    components.html("""
                    <div style="display:flex; flex-direction:column; gap:8px;">
                        <button onclick="printAs('portrait')" style="width:100%; padding:11px; background:#00338D; color:white; border:none; border-radius:6px; cursor:pointer; font-weight:bold; font-size:13px;">
                            🖨️ 导出竖版 A4 PDF
                        </button>
                        <button onclick="printAs('widescreen')" style="width:100%; padding:11px; background:#008578; color:white; border:none; border-radius:6px; cursor:pointer; font-weight:bold; font-size:13px;">
                            🖨️ 导出横版 16:9 PDF
                        </button>
                    </div>
                    <script>
                    function printAs(mode) {
                        const doc = window.parent.document;
                    
                        // 移除旧打印样式
                        const old = doc.getElementById('dynamic-print-style');
                        if (old) old.remove();
                    
                        const style = doc.createElement('style');
                        style.id = 'dynamic-print-style';
                    
                        if (mode === 'widescreen') {
                            style.innerHTML = `
                                @page {
                                    size: landscape;
                                    margin: 5mm 8mm;
                                }
                    
                                @page {
                                    -webkit-size: landscape;
                                }
                    
                                .main .block-container {
                                    width: 100% !important;
                                    max-width: 100% !important;
                                }
                            `;
                        } else {
                            style.innerHTML = `
                                @page {
                                    size: A4 portrait;
                                    margin: 10mm;
                                }
                    
                                @page {
                                    -webkit-size: A4 portrait;
                                }
                            `;
                        }
                    
                        doc.head.appendChild(style);
                    
                        // 强制重新计算布局
                        window.parent.getComputedStyle(doc.body).width;
                    
                        // ==========================================
                        // 等待 Plotly 图表完成渲染
                        // ==========================================
                        const startTime = Date.now();
                    
                        function checkChartsReady() {
                    
                            const charts = Array.from(
                                doc.querySelectorAll('.plotly-graph-div')
                            );
                    
                            // 检查所有 Plotly 图表
                            const allReady =
                                charts.length === 0 ||
                                charts.every(chart => {
                    
                                    const svg = chart.querySelector('.main-svg');
                                    const rect = chart.getBoundingClientRect();
                    
                                    return (
                                        svg &&
                                        rect.width > 100 &&
                                        rect.height > 100
                                    );
                                });
                    
                            // 图表全部完成
                            if (allReady) {
                    
                                // 再给浏览器两帧时间完成最终布局
                                requestAnimationFrame(() => {
                                    requestAnimationFrame(() => {
                    
                                        setTimeout(() => {
                                            window.parent.print();
                                        }, 300);
                    
                                    });
                                });
                    
                                return;
                            }
                    
                            // 最多等待 8 秒
                            if (Date.now() - startTime > 8000) {
                    
                                console.warn(
                                    'Plotly 图表等待超过 8 秒，继续打印'
                                );
                    
                                window.parent.print();
                                return;
                            }
                    
                            // 100ms 后再次检查
                            setTimeout(checkChartsReady, 100);
                        }
                    
                        checkChartsReady();
                    }
                    </script>
                    """, height=100)
            else:
                df_sub1 = df_n[df_n['一级分类'] == main_nav]
                raw_sec_levels = [x for x in df_sub1['二级分类'].unique() if x != "全部"]
                if len(raw_sec_levels) == 0:
                    charts = [x for x in df_sub1['对应图表名称'].unique() if x]
                    chart_nav = st.radio("具体图表", charts, key="s7_c")
                    matched = df_sub1[df_sub1['对应图表名称'] == chart_nav]
                    active_m_id = matched.iloc[0]['模块ID'] if not matched.empty else None
                else:
                    sub_nav = st.radio("📂 二级模块", ["全部"] + raw_sec_levels, key="s7_s")
                    if sub_nav != "全部":
                        df_sub2 = df_sub1[df_sub1['二级分类'] == sub_nav]
                        charts = [x for x in df_sub2['对应图表名称'].unique() if x]
                        chart_nav = st.radio("具体图表", charts, key="s7_c")
                        matched = df_sub2[df_sub2['对应图表名称'] == chart_nav]
                        active_m_id = matched.iloc[0]['模块ID'] if not matched.empty else None
                    else:
                        charts = [x for x in df_sub1['对应图表名称'].unique() if x]
                        chart_nav = st.radio("具体图表 (当前二级：全部)", charts, key="s7_c_all")
                        matched = df_sub1[df_sub1['对应图表名称'] == chart_nav]
                        active_m_id = matched.iloc[0]['模块ID'] if not matched.empty else None
        else:
            st.warning("⚠️ 请先加载包含层级信息的注释表")

    # ----- 配置与图片上传 -----
    st.markdown("<div class='no-print'>", unsafe_allow_html=True)
    with st.expander("⚙️ 公司级图表设置与图片覆盖", expanded=False):
        c0, c1, c2, c3, c4 = st.columns([1, 2, 1, 1, 1])   
        with c0:
            type_options = sorted([
                str(x).strip() for x in df_raw["公司类型"].dropna().unique()
                if str(x).strip() not in ["", "nan", "None"]
            ])
            if "s7_company_types" not in st.session_state:
                st.session_state["s7_company_types"] = ["全部"]
            selected_types_raw = st.multiselect(
                "公司类型",
                options=["全部"] + type_options,
                default=st.session_state["s7_company_types"],
                key="s7_company_types"
            )
            if (not selected_types_raw) or ("全部" in selected_types_raw):
                selected_types = ["全部"]
                df_filtered = df_raw.copy()
                st.write("df_filtered 行数:", len(df_filtered))
                st.dataframe(df_filtered.head(3))
            else:
                selected_types = selected_types_raw
                df_filtered = df_raw[df_raw["公司类型"].astype(str).str.strip().isin(selected_types)].copy()
        with c1: 
            raw_ordered_cos = list(dict.fromkeys(df_filtered['公司'].dropna().tolist()))
            selected_cos = st.multiselect("展示公司", options=raw_ordered_cos, default=raw_ordered_cos)
        with c2: 
            unit_label = st.selectbox("显示单位", ["十亿元", "亿元", "百万元", "十万元"])
            divisor = {"十亿元": 1000, "亿元": 100, "百万元": 1, "十万元": 0.1}[unit_label]
        with c3: highlight_co = st.selectbox("特定追踪", ["无"] + selected_cos)
        with c4: enable_ai = st.toggle("一键AI分析", value=False)
        
        st.markdown("<hr style='margin: 5px 0 15px 0;'>", unsafe_allow_html=True)
        sc1, sc2, sc3 = st.columns([2, 2, 3])
        with sc1:
            available_fields = sorted([str(x) for x in df_filtered['字段名'].unique() if str(x).strip() not in ['nan', 'None', '']])
            sort_field = st.selectbox("📊 图表展示顺序依据", ["默认（按列表原始顺序）"] + available_fields)
        with sc2:
            sort_order = st.radio("排序方向", ["降序 (从大到小) ⬇️", "升序 (从小到大) ⬆️"], horizontal=True)
        # ----- 统一公司排序逻辑 -----
        # 如果用户选择“默认”，则按保险服务收入降序排列（与“新准则保险服务收入排名”一致）
        if sort_field == "默认（按列表原始顺序）":
            # 使用保险服务收入作为排序依据
            sort_field_actual = "保险服务收入"
            is_desc = True  # 降序
            if sort_field_actual in available_fields:
                df_sort = df_filtered[(df_filtered['字段名'] == sort_field_actual) & (df_filtered['报告年份'].astype(str) == str(latest_year))]
                val_col = "(百万)人民币" if "(百万)人民币" in df_sort.columns else df_sort.columns[-1]
                sort_map = {}
                for _, r in df_sort.iterrows():
                    val = pd.to_numeric(r.get(val_col, np.nan), errors='coerce')
                    sort_map[str(r['公司']).strip()] = val
                default_val = float('-inf') if is_desc else float('inf')
                selected_cos = sorted(selected_cos, key=lambda x: sort_map.get(x) if pd.notna(sort_map.get(x)) else default_val, reverse=is_desc)
            else:
                # 若没有保险服务收入，则按公司名称排序
                selected_cos = sorted(selected_cos)
        elif sort_field != "默认（按列表原始顺序）" and selected_cos:
            df_sort = df_filtered[(df_filtered['字段名'] == sort_field) & (df_filtered['报告年份'].astype(str) == str(latest_year))]
            val_col = "(百万)人民币" if "(百万)人民币" in df_sort.columns else df_sort.columns[-1]
            is_desc = "降序" in sort_order
            sort_map = {}
            for _, r in df_sort.iterrows():
                val = pd.to_numeric(r.get(val_col, np.nan), errors='coerce')
                sort_map[str(r['公司']).strip()] = val
            default_val = float('-inf') if is_desc else float('inf')
            selected_cos = sorted(selected_cos, key=lambda x: sort_map.get(x) if pd.notna(sort_map.get(x)) else default_val, reverse=is_desc)
        st.session_state['selected_cos_cache'] = selected_cos
        global HL_BOX_FILL, HL_BOX_LINE
        HL_BOX_FILL = "rgba(0,51,141,0.08)"   # 适当加深背景
        HL_BOX_LINE = "rgba(0,51,141,0.8)"    # 加深边框
                
        st.markdown("---")
        st.caption("📸 手动上传图片（png或jpg）")
        if 'manual_upload_images' not in st.session_state:
            st.session_state.manual_upload_images = {}
        uploaded_files = st.file_uploader("拖入截图文件", type=['png', 'jpg'], accept_multiple_files=True)

        if uploaded_files and ordered_modules:
            cols = st.columns(2)
            for i, file in enumerate(uploaded_files):
                with cols[i % 2]:
                    def get_name(m):
                        if m == "不匹配/跳过":
                            return "不匹配/跳过"
                        return notes_dict.get(m, {}).get('title', m)
                    sel_mid = st.selectbox(
                        f"图片 {file.name} 对应：",
                        options=["不匹配/跳过"] + ordered_modules,
                        format_func=get_name,
                        key=f"up_{i}"
                    )
                    if sel_mid != "不匹配/跳过":
                        st.session_state.manual_upload_images[sel_mid] = file
                    st.image(file, use_column_width=True)
    st.markdown("</div>", unsafe_allow_html=True)


    # ----- 获取配置参数 -----
    st.session_state['df_filtered'] = df_filtered
    st.session_state['selected_cos_cache'] = selected_cos
    st.session_state['divisor'] = divisor
    st.session_state['latest_year'] = latest_year
    st.session_state['highlight_co'] = highlight_co
    st.session_state['enable_ai'] = enable_ai
    st.session_state['unit_label'] = unit_label
    st.session_state['print_mode'] = print_mode
    st.session_state['active_m_id'] = active_m_id
    st.session_state['prev_year'] = prev_year
    st.session_state['divisor'] = divisor
    st.session_state['HL_BOX_FILL'] = HL_BOX_FILL
    st.session_state['HL_BOX_LINE'] = HL_BOX_LINE
    st.session_state['highlight_co'] = highlight_co

    if not selected_cos:
        selected_cos = list(df_filtered['公司'].unique())
        st.session_state['selected_cos_cache'] = selected_cos

    st.markdown("<h3 class='no-print' style='font-weight:700;'>📊 公司级对标报告</h3>", unsafe_allow_html=True)

    # ----- 打印模式（一键显示全部）----- 
    if print_mode:
        # ===== 封面页（使用 <img> 标签，稳定且快速） =====
        cover_image_url = "https://raw.githubusercontent.com/Polly031021/Annual-Report-System/main/%E6%A0%87%E9%A2%98%E9%A1%B5.jpg"
    
        st.markdown(f"""
        <style>
        .cover-page {{
            position: relative;
            width: 100%;
            height: 100vh;
            overflow: hidden;
            page-break-after: always;
            margin: 0;
            padding: 0;
            -webkit-print-color-adjust: exact;
            print-color-adjust: exact;
        }}
        .cover-page img {{
            position: absolute;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            object-fit: cover;
            z-index: 1;
        }}
        .cover-content {{
            position: absolute;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            text-align: center;
            color: white !important;
            z-index: 2;
            font-family: "Microsoft YaHei", "PingFang SC", sans-serif;
            background: rgba(0, 0, 0, 0.3);
        }}
        .cover-title, .cover-subtitle, .cover-date {{
            color: white !important;
            text-shadow: 2px 2px 12px rgba(0,0,0,0.6);
        }}
        .cover-title {{
            font-size: 56px;
            font-weight: 700;
            letter-spacing: 4px;
            margin-bottom: 20px;
        }}
        .cover-subtitle {{
            font-size: 28px;
            font-weight: 300;
            letter-spacing: 8px;
            margin-bottom: 30px;
        }}
        .cover-date {{
            font-size: 20px;
            font-weight: 300;
            letter-spacing: 6px;
            opacity: 0.9;
        }}
        @media print {{
            .cover-page {{
                height: 100%;
                width: 100%;
                page-break-after: always;
            }}
            .cover-page img {{
                object-fit: cover;
            }}
        }}
        </style>
        <div class="cover-page">
            <img src="{cover_image_url}" alt="封面">
            <div class="cover-content">
                <div class="cover-title">2025年新会计准则业绩表现和洞察</div>
                <div class="cover-subtitle">保险公司</div>
                <div class="cover-date">2026年8月</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
        # 直接渲染所有模块，不显示进度条
        for idx, m_id in enumerate(ordered_modules):
            render_report_module(m_id, print_mode, is_first=(idx == 0))
    
        return

    # ----- 单模块模式 -----
    if not active_m_id:
        st.info("请在左侧导航栏中选择一个具体图表模块。")
        return

    render_report_module(active_m_id, print_mode, is_first=True)
    return
    
# 7.1 路由引擎
def render_pure_chart_entity(m_id, print_mode):
    global notes_dict
    
    # ====== 先读取所有必要的 session_state 变量 ======
    df_filtered = st.session_state.get('df_filtered', pd.DataFrame())
    selected_cos = st.session_state.get('selected_cos_cache', [])
    latest_year = st.session_state.get('latest_year', 2025)
    prev_year = st.session_state.get('prev_year', 2024)
    divisor = st.session_state.get('divisor', 1)
    unit_label = st.session_state.get('unit_label', "百万元")
    highlight_co = st.session_state.get('highlight_co', "无")
    global HL_BOX_FILL, HL_BOX_LINE
    HL_BOX_FILL = st.session_state.get('HL_BOX_FILL', "rgba(0,51,141,0.03)")
    HL_BOX_LINE = st.session_state.get('HL_BOX_LINE', "rgba(0,51,141,0.35)")
    current_hl = highlight_co if highlight_co else "无"
    
    # ====== 概览模块 ======
    if m_id in ["概览", "overview"]:
        fig = create_overview_table(df_filtered, selected_cos, latest_year, prev_year, unit_label)
        # 替换原来的 st.plotly_chart
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "概览")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return
    # ==========================================
    # 文本披露 - 折现率表格
    # ==========================================
    if m_id == "discount_rate":
        available_years = sorted([int(y) for y in df_filtered['报告年份'].unique() if y.isdigit()])
        years_to_show = available_years[-2:] if len(available_years) >= 2 else available_years
        fig = create_disclosure_table(
            df_filtered, 
            field_name='折现率假设',      # 请确认实际字段名
            title='折现率假设披露', 
            cos=selected_cos, 
            years=years_to_show
        )
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "折现率")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return
    
    # ==========================================
    # 文本披露 - 非金融风险调整表格
    # ==========================================
    if m_id == "risk_margin":
        fig = create_risk_margin_table(df_filtered, selected_cos, title='非金融风险调整披露')
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "非金融风险调整")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return
    
    # ==========================================
    # 1. 文本类披露（已在主调度中单独处理，此处留空以防误调用）
    # ==========================================
    if m_id == "policy_method":
        # 如果 policy_method 也需要表格，可以参照下面改为表格；否则保留原样
        display_textual_disclosures(df_filtered, selected_cos, latest_year)
        display_notes(m_id, df_filtered, "")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 2. 单指标柱状图（两年对比 + 增长率标注）
    # ==========================================
    if m_id == "premium_old":  # 保险业务收入（旧准则）
        # 自动获取数据中存在的年份（取最近三年，若不足则取全部）
        available_years = sorted([int(y) for y in df_filtered['报告年份'].unique() if y.isdigit()])
        # 取最近三年（若不足则取全部）
        years = available_years[-3:] if len(available_years) >= 3 else available_years
        if not years:
            st.warning("未找到有效年份数据")
            return
        fig = create_premium_old_chart(
            df_filtered, selected_cos, years, divisor, unit_label, current_hl
        )
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "保险业务收入")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return
    # ==========================================
    # 保险服务收入（新准则）—— 三年柱状图
    # ==========================================
    if m_id == "premium_ranking":
        available_years = sorted([int(y) for y in df_filtered['报告年份'].unique() if y.isdigit()])
        years = available_years[-3:] if len(available_years) >= 3 else available_years
        if not years:
            st.warning("未找到有效年份数据")
            return
        fig = create_multi_year_bar_chart(
            df_filtered, "保险服务收入", selected_cos, years,
            divisor=divisor, unit_label=unit_label, highlight_co=current_hl, is_percentage=False
        )
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "保险服务收入")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 新旧准则比值 —— 三年折线图
    # ==========================================
    if m_id == "new_old_ratio":
        df_temp = df_filtered[df_filtered['字段名'].str.strip() == '新旧准则比值']
        available_years = sorted([int(y) for y in df_temp['报告年份'].unique() if y.isdigit()])
        if not available_years:
            st.warning("未找到新旧准则比值数据")
            return
        years = available_years
        fig = create_line_subplots_chart(
            df_filtered, "新旧准则比值", selected_cos, years,
            divisor=1, unit_label="比值", highlight_co=current_hl,
            is_percentage=False, decimal_places=3
        )
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "新旧准则比值")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 投资成分占比 —— 三年折线图
    # ==========================================
    if m_id == "investment_component":
        df_temp = df_filtered[df_filtered['字段名'].str.strip() == '投资成分占比']
        available_years = sorted([int(y) for y in df_temp['报告年份'].unique() if y.isdigit()])
        if not available_years:
            st.warning("未找到投资成分占比数据")
            return
        years = available_years  # 显示所有存在的年份
        fig = create_line_subplots_chart(
            df_filtered, "投资成分占比", selected_cos, years,
            divisor=1, unit_label="百分比", highlight_co=current_hl,
            is_percentage=True, decimal_places=1
        )
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "投资成分占比", is_pct=True)
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 保费增长率（旧准则）—— 三年折线图
    # ==========================================
    if m_id == "premium_growth":
        # 从数据中提取保费增长率字段实际存在的年份
        df_temp = df_filtered[df_filtered['字段名'].str.strip() == '保费增长率（旧准则）']
        available_years = sorted([int(y) for y in df_temp['报告年份'].unique() if y.isdigit()])
        if not available_years:
            st.warning("未找到保费增长率数据")
            return
        # 直接使用存在的年份（通常为2024和2025）
        years = available_years
        fig = create_line_subplots_chart(
            df_filtered, "保费增长率（旧准则）", selected_cos, years,
            divisor=1, unit_label="百分比", highlight_co=current_hl,
            is_percentage=True, decimal_places=1
        )
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "保费增长率", is_pct=True)
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

        
    single_metric_map = {
        "loss_ratio": ("综合赔付率", False, True),
        "expense_ratio": ("综合费用率", False, True),
        "loss_component": ("亏损成分占比", False, True),
        "backtest_deviation": ("回溯偏差率", False, True),
    }

    # ==========================================
    # 综合成本率变化 + 综合费用率/赔付率 breakdown
    # ==========================================
    if m_id == "cor_trend":
        if not print_mode:
            c1, c2, c3 = st.columns(3)
            with c1:
                show_labels = st.toggle("显示标签", True, key=f"lab_{m_id}")
            with c2:
                pct_sz = st.slider("涨幅字号", 8, 24, 14, key=f"psz_{m_id}")   # 不再使用
            with c3:
                gap = st.slider("柱间距", 0.1, 0.8, 0.15, key=f"gap_{m_id}")   # 不再使用
        else:
            show_labels = st.session_state.get(f"lab_{m_id}", True)
            pct_sz = st.session_state.get(f"psz_{m_id}", 14)
            gap = st.session_state.get(f"gap_{m_id}", 0.15)

        # ---- 只保留第二张图：综合费用率与赔付率 breakdown ----
        raw_df2 = df_filtered.copy()
        cos_list = st.session_state.get('selected_cos_cache', [])
        prev_y = st.session_state.get('prev_year', 2024)
        latest_y = st.session_state.get('latest_year', 2025)
        years = [str(prev_y), str(latest_y)]

        # ===== 修改点1：同时提取"综合成本率"，作为灰色柱兜底数据 =====
        metrics = ["综合赔付率", "综合费用率", "综合成本率"]
        data_dict = {}
        for co in cos_list:
            data_dict[co] = {}
            for yr in years:
                data_dict[co][yr] = {}
                for m in metrics:
                    val = raw_df2[
                        (raw_df2['公司'].astype(str).str.strip() == str(co).strip()) &
                        (raw_df2['报告年份'].astype(str).str.replace('.0', '', regex=False).str.strip() == yr) &
                        (raw_df2['字段名'].astype(str).str.strip() == m)
                    ]['(百万)人民币']
                    if not val.empty:
                        num_val = pd.to_numeric(val.iloc[0], errors='coerce')
                        data_dict[co][yr][m] = 0 if pd.isna(num_val) else float(num_val)
                    else:
                        data_dict[co][yr][m] = 0

        # ===== 修改点2：构造数据，"无拆解但有COR"的位置走灰色柱 =====
        x_vals = []
        y_exp = []    # 费用率
        y_loss = []   # 赔付率
        y_cor = []    # 综合成本率（灰色兜底柱）
        total_vals = []  # 柱顶总标注
        missing_cos = []  # 记录未披露拆解的公司，用于脚注

        for co in cos_list:
            # 判断该公司是否所有年份都没有拆解数据（但有综合成本率）
            has_no_breakdown = all(
                data_dict[co][yr]["综合赔付率"] == 0 and
                data_dict[co][yr]["综合费用率"] == 0
                for yr in years
            ) and any(
                data_dict[co][yr]["综合成本率"] != 0
                for yr in years
            )
            if has_no_breakdown:
                missing_cos.append(co)

            base_idx = 2 * len(x_vals) // 2  # 保持每组两个柱子的位置逻辑
            group_start = 2 * cos_list.index(co)
            for j, yr in enumerate(years):
                x_vals.append(group_start + j)
                exp = data_dict[co][yr]["综合费用率"]
                loss = data_dict[co][yr]["综合赔付率"]
                cor = data_dict[co][yr]["综合成本率"]
                if exp == 0 and loss == 0 and cor != 0:
                    # 未披露拆解 → 整根灰色柱显示综合成本率
                    y_exp.append(0)
                    y_loss.append(0)
                    y_cor.append(cor)
                    total_vals.append(cor)
                else:
                    # 正常情况 → 蓝色堆叠柱，灰色柱为0不显示
                    y_exp.append(exp)
                    y_loss.append(loss)
                    y_cor.append(0)
                    total_vals.append(exp + loss)

        # 年份标签（每个柱子对应一个年份）
        year_labels = []
        for _ in cos_list:
            year_labels.append(f"{prev_y}YE")
            year_labels.append(f"{latest_y}YE")

        fig2 = go.Figure()

        # ===== 修改点3：新增灰色兜底柱（正常公司该值为0，不会显示） =====
        fig2.add_trace(go.Bar(
            name="综合成本率（未披露拆解）",
            x=x_vals,
            y=y_cor,
            marker_color="#B0BEC5",
            text=[f"{v:.1%}" if show_labels and v != 0 else "" for v in y_cor],
            textposition='inside',
            insidetextanchor='middle',
            textfont=dict(color="#333333", size=11),
            width=0.8,
        ))

        fig2.add_trace(go.Bar(
            name="综合费用率",
            x=x_vals,
            y=y_exp,
            marker_color="#1E49E2",
            text=[f"{v:.1%}" if show_labels and v != 0 else "" for v in y_exp],
            textposition='inside',
            insidetextanchor='middle',
            textfont=dict(color="white", size=11),
            width=0.8,
        ))

        fig2.add_trace(go.Bar(
            name="综合赔付率",
            x=x_vals,
            y=y_loss,
            marker_color="#00338D",
            text=[f"{v:.1%}" if show_labels and v != 0 else "" for v in y_loss],
            textposition='inside',
            insidetextanchor='middle',
            textfont=dict(color="white", size=11),
            width=0.8,
        ))

        # 添加总数值标注（柱顶；灰色柱显示综合成本率本身）
        for x, total in zip(x_vals, total_vals):
            if total != 0:
                fig2.add_annotation(
                    x=x,
                    y=total,
                    text=f"{total:.1%}",
                    showarrow=False,
                    font=dict(size=12, color="#333"),
                    yshift=5,
                )

        # ===== 修改点4：脚注，仿照 KPMG PPT 的 * 说明 =====
        if missing_cos:
            footnote = "*" + "、".join(missing_cos) + "年报中未详细披露综合费用率和综合赔付率"
            fig2.add_annotation(
                text=footnote,
                xref="paper", yref="paper",
                x=0, y=-0.30,
                showarrow=False,
                xanchor="left",
                font=dict(size=10, color="#888888", style="italic"),
            )

        fig2.update_layout(
            barmode='relative',
            title="综合费用率与赔付率 breakdown",
            xaxis=dict(
                tickvals=x_vals,
                ticktext=year_labels,
                tickangle=-30,
                tickfont=dict(size=11, family="Microsoft YaHei", color="#333", style="normal"),
            ),
            yaxis=dict(title="占保险服务收入比例", tickformat=".1%"),
            height=420,
            bargap=0.15,
            # ===== 修改点5：底部留出脚注空间 =====
            margin=dict(t=50, b=110),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
        )

        # 添加公司名 annotations（每组柱子中间下方）
        for i, co in enumerate(cos_list):
            fig2.add_annotation(
                x=2 * i + 0.5,
                y=-0.08,  # 相对于图表高度的比例，放在 x 轴标签下方
                text=co,
                showarrow=False,
                font=dict(size=12, family="Microsoft YaHei", color="#333", style="normal"),
                xref='x',
                yref='paper',
                yanchor='top',
                xanchor='center'
            )

        show_chart(fig2, print_mode, m_id)

        display_notes(m_id, df_filtered, "综合成本率")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 3. 综合赔付拆解（多因子分组柱状图）
    # ==========================================
    if m_id == "cor_components":
        # 过滤掉太平产险
        filtered_cos = [co for co in selected_cos if co != "太平产险"]
        fig = create_cor_breakdown_stacked_chart(
            df_filtered, filtered_cos, latest_year, divisor, unit_label, current_hl
        )
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "综合赔付率")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return
        
    # ==========================================
    # 4. 费用分类构成（使用通用多分类引擎）
    # ==========================================
    if m_id == "expense_classification":
        if not print_mode:
            c1, c2, c3 = st.columns(3)
            with c1:
                show_labels = st.toggle("显示标签", True, key=f"lab_{m_id}")
            with c2:
                label_size = st.slider("标签字号", 8, 20, 12, key=f"lsz_{m_id}")
            with c3:
                bar_width = st.slider("柱宽", 0.2, 1.0, 0.6, key=f"wid_{m_id}")
        else:
            show_labels = st.session_state.get(f"lab_{m_id}", True)
            label_size = st.session_state.get(f"lsz_{m_id}", 12)
            bar_width = st.session_state.get(f"wid_{m_id}", 0.6)
        
        # 只显示最近两年
        prev_year = st.session_state.get('prev_year', 2024)
        latest_year = st.session_state.get('latest_year', 2025)
        years_to_show = [str(prev_year), str(latest_year)]
        
        field_map = {
            "获取费用占比": "获取费用",
            "维持费用占比": "维持费用",
            "非履约费用占比": "非履约费用",
            "履约费用占比": "履约费用"
        }
        color_map = {
            "获取费用占比": KPMG_COLORS[0],  # 深蓝
            "维持费用占比": KPMG_COLORS[1],  # 亮蓝
            "非履约费用占比": KPMG_COLORS[10], # 橙红
            "履约费用占比": KPMG_COLORS[5],  # 深紫
        }

        fig, df_avg = create_kpmg_multi_composition_chart(
            df_filtered, field_map, color_map, "",
            show_labels=show_labels, label_size=label_size, bar_width=bar_width, 
            co_font_size=13, highlight_co=current_hl, add_zero_line=False,
            years_to_show=years_to_show
        )
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "费用分类")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 5. 税务分析：税前/净利润及有效税率
    # ==========================================
    if m_id == "tax_analysis":
        if not print_mode:
            c1, c2, c3, c4, c5 = st.columns(5)
            with c1:
                show_labels = st.toggle("显示标签", True, key=f"lab_{m_id}")
            with c2:
                bar_width = st.slider("柱宽", 0.1, 0.8, 0.35, key=f"wid_{m_id}")
            with c3:
                label_size = st.slider("字号", 8, 16, 12, key=f"lsz_{m_id}")
            with c4:
                co_font_size = st.slider("公司名字号", 10, 24, 14, key=f"cfs_{m_id}")
            with c5:
                co_y_offset = st.slider("标题偏移", 1.0, 1.2, 1.08, step=0.01, key=f"off_{m_id}")
        else:
            show_labels = st.session_state.get(f"lab_{m_id}", True)
            bar_width = st.session_state.get(f"wid_{m_id}", 0.35)
            label_size = st.session_state.get(f"lsz_{m_id}", 12)
            co_font_size = st.session_state.get(f"cfs_{m_id}", 14)
            co_y_offset = st.session_state.get(f"off_{m_id}", 1.08)
        
        # 数据预处理
        df_tax_sub = df_filtered[(df_filtered['字段名'].isin(['税前利润总额', '净利润'])) & (df_filtered['公司'].isin(selected_cos))].drop_duplicates(subset=['公司', '报告年份', '字段名']).copy()
        if not df_tax_sub.empty:
            df_tax_pivot = df_tax_sub.pivot_table(index=['公司', '报告年份'], columns='字段名', values='(百万)人民币').fillna(0).reset_index()
            denom = df_tax_pivot['税前利润总额'].replace(0, np.nan)
            df_tax_pivot['有效税率'] = np.where(
                df_tax_pivot['税前利润总额'] != 0,
                (df_tax_pivot['税前利润总额'] - df_tax_pivot['净利润']) / denom,
                0
            )
            df_tax_pivot['报告年份'] = df_tax_pivot['报告年份'].astype(str).str.replace(".0", "", regex=False) + "YE"
            fig = create_tax_subplot_chart(
                df_tax_pivot, selected_cos, 
                show_labels=show_labels, bar_width=bar_width, label_size=label_size, 
                co_font_size=co_font_size, co_y_offset=co_y_offset, highlight_co=current_hl
            )
        else:
            fig = go.Figure()
            fig.add_annotation(text="无税前/净利润数据", x=0.5, y=0.5, showarrow=False)
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "税前利润")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 6. 净资产变动明细表（纯 HTML 表格）
    # ==========================================
    if m_id == "equity_change_detail":
        html_table = create_equity_change_detail_table(
            df_raw=df_filtered,
            target_year=latest_year,
            cos=selected_cos,
            divisor=divisor,
            unit_label=unit_label,
            highlight_co=current_hl
        )
        st.markdown(html_table, unsafe_allow_html=True)
        display_notes(m_id, df_filtered, "净资产")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 7. 总资产 / 净资产趋势（带明细表）
    # ==========================================
    if m_id == "asset_trend":
        if not print_mode:
            c1, c2, c3 = st.columns(3)
            with c1:
                show_labels = st.toggle("显示标签", True, key=f"lab_{m_id}")
            with c2:
                pct_sz = st.slider("涨幅字号", 8, 24, 14, key=f"psz_{m_id}")
            with c3:
                gap = st.slider("柱间距", 0.1, 0.8, 0.15, key=f"gap_{m_id}")
        else:
            show_labels = st.session_state.get(f"lab_{m_id}", True)
            pct_sz = st.session_state.get(f"psz_{m_id}", 14)
            gap = st.session_state.get(f"gap_{m_id}", 0.15)
        
        fig = create_kpmg_chart(df_filtered, "总资产", "", show_labels, pct_sz, gap, 
                                sort_by_value=False, is_percentage=False)
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "总资产")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    if m_id == "equity_trend":
        if not print_mode:
            c1, c2, c3 = st.columns(3)
            with c1:
                show_labels = st.toggle("显示标签", True, key=f"lab_{m_id}")
            with c2:
                pct_sz = st.slider("涨幅字号", 8, 24, 14, key=f"psz_{m_id}")
            with c3:
                gap = st.slider("柱间距", 0.1, 0.8, 0.15, key=f"gap_{m_id}")
        else:
            show_labels = st.session_state.get(f"lab_{m_id}", True)
            pct_sz = st.session_state.get(f"psz_{m_id}", 14)
            gap = st.session_state.get(f"gap_{m_id}", 0.15)
        
        fig = create_kpmg_chart(df_filtered, "期末股东权益", "", show_labels, pct_sz, gap, 
                                sort_by_value=False, is_percentage=False)
        show_chart(fig, print_mode, m_id)
        # 下方补充净资产变动明细表
        html_table = create_equity_change_detail_table(
            df_raw=df_filtered,
            target_year=latest_year,
            cos=selected_cos,
            divisor=divisor,
            unit_label=unit_label,
            highlight_co=current_hl
        )
        st.markdown(html_table, unsafe_allow_html=True)
        display_notes(m_id, df_filtered, "净资产")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 8. 费用结构拆解图（堆叠柱状图）
    # ==========================================
    if m_id == "expense_structure":
        # 过滤掉太平产险
        filtered_cos = [co for co in selected_cos if co != "太平产险"]
        if not print_mode:
            c1, c2, c3 = st.columns(3)
            with c1:
                show_labels = st.toggle("显示标签", True, key=f"lab_{m_id}")
            with c2:
                label_size = st.slider("标签字号", 8, 20, 12, key=f"lsz_{m_id}")
            with c3:
                bar_width = st.slider("柱宽", 0.2, 1.0, 0.5, key=f"wid_{m_id}")
        else:
            show_labels = st.session_state.get(f"lab_{m_id}", True)
            label_size = st.session_state.get(f"lsz_{m_id}", 12)
            bar_width = st.session_state.get(f"wid_{m_id}", 0.5)
        
        fig, _ = create_expense_structure_chart(
            df_filtered, filtered_cos,
            show_labels=show_labels, label_size=label_size, bar_width=bar_width,
            co_font_size=13, highlight_co=current_hl
        )
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "费用")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 9. 六维度散点图矩阵
    # ==========================================
    if m_id == "six_dimensions":
        if not print_mode:
            c1, c2, c3 = st.columns(3)
            with c1:
                show_labels = st.toggle("显示标签", True, key=f"lab_{m_id}")
            with c2:
                label_size = st.slider("字号", 8, 16, 11, key=f"lsz_{m_id}")
            with c3:
                dot_size = st.slider("点大小", 4, 16, 11, key=f"dsz_{m_id}")
        else:
            show_labels = st.session_state.get(f"lab_{m_id}", True)
            label_size = st.session_state.get(f"lsz_{m_id}", 11)
            dot_size = st.session_state.get(f"dsz_{m_id}", 11)
        
        figs = create_nonlife_six_dimensional_charts(
            df_raw=df_filtered,
            target_year=latest_year,
            cos=selected_cos,
            divisor=divisor,
            unit_label=unit_label,
            highlight_co=current_hl,
            label_size=label_size,
            show_labels=show_labels,
            dot_size=dot_size
        )
        # 以 3 列网格展示 6 个子图
        for i, fig in enumerate(figs):
            if i % 3 == 0:
                cols = st.columns(3)
            with cols[i % 3]:
                st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})
        display_notes(m_id, df_filtered, "六维度")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 🆕 10. 综合赔付率趋势图（新增）
    # ==========================================
    if m_id == "claim_ratio_trend":
        if not print_mode:
            show_labels = st.toggle("显示标签", True, key=f"lab_{m_id}")
        else:
            show_labels = st.session_state.get(f"lab_{m_id}", True)
        figs = create_ratio_trend_chart(
            df_filtered, selected_cos,
            factor_list=[
                "当期发生赔款及理赔费用",
                "已发生赔款负债履约现金流变动",
                "亏损合同损益",
                "再保净成本"
            ],
            ratio_field="综合赔付率",
            title_prefix="综合赔付率拆解趋势",
            divisor=divisor,
            unit_label=unit_label,
            highlight_co=current_hl,
            show_labels=show_labels,
            ratio_first=True   # 综合图放在首位
        )
        for fig in figs:
            show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "综合赔付率")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 🆕 11. 综合费用率趋势图（新增）
    # ==========================================
    if m_id == "expense_ratio_trend":
        if not print_mode:
            show_labels = st.toggle("显示标签", True, key=f"lab_{m_id}")
        else:
            show_labels = st.session_state.get(f"lab_{m_id}", True)
        
        figs = create_ratio_trend_chart(
            df_filtered, selected_cos,
            factor_list=[
                "获取费用摊销",
                "维持费用"
            ],
            ratio_field="综合费用率",
            title_prefix="综合费用率拆解趋势",
            divisor=divisor,
            unit_label=unit_label,
            highlight_co=current_hl,
            show_labels=show_labels,
            ratio_first=True   # 综合图放在首位
        )
        
        # 逐个显示图表（每个因子一张图）
        for fig in figs:
            show_chart(fig, print_mode, m_id)
        
        # 显示注释和底部说明
        display_notes(m_id, df_filtered, "综合费用率")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return
    # ==========================================
    # 保险服务收入业务构成（堆叠图）
    # ==========================================
    if m_id == "premium_composition":
        fig = create_premium_stacked_chart(
            df_filtered, selected_cos, latest_year, divisor, unit_label, current_hl
        )
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "保险服务收入构成")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 承保利润业务构成（堆叠图）
    # ==========================================
    if m_id == "profit_composition":
        fig = create_profit_contribution_stacked_chart(
            df_filtered, selected_cos, latest_year, divisor, unit_label, current_hl
        )
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "承保利润构成")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 保险业务构成（两种方法占比堆叠图）
    # ==========================================
    if m_id == "premium_method_composition":
        fig = create_method_composition_chart(
            df_filtered, selected_cos, latest_year, divisor, unit_label
        )
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "保险业务构成")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 关键指标 - 总资产与净资产（柱状图）
    # ==========================================
    if m_id == "key_assets_equity":
        # 设置UI控件（仅非打印模式）
        if not print_mode:
            c1, c2, c3 = st.columns(3)
            with c1:
                show_labels = st.toggle("显示标签", True, key=f"lab_{m_id}")
            with c2:
                pct_sz = st.slider("涨幅字号", 8, 24, 14, key=f"psz_{m_id}")
            with c3:
                gap = st.slider("柱间距", 0.1, 0.8, 0.15, key=f"gap_{m_id}")
        else:
            show_labels = st.session_state.get(f"lab_{m_id}", True)
            pct_sz = st.session_state.get(f"psz_{m_id}", 14)
            gap = st.session_state.get(f"gap_{m_id}", 0.15)
        
        # 绘制总资产图
        fig_asset = create_kpmg_chart(
            df_filtered, "总资产", "", show_labels, pct_sz, gap,
            sort_by_value=False, is_percentage=False
        )
        # 绘制净资产图
        fig_equity = create_kpmg_chart(
            df_filtered, "净资产", "", show_labels, pct_sz, gap,
            sort_by_value=False, is_percentage=False
        )
        
        # 显示两个图表（上下排列）
        show_chart(fig_asset, print_mode, m_id)
        show_chart(fig_equity, print_mode, m_id)
        
        display_notes(m_id, df_filtered, "总资产与净资产")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 关键指标 - 承保利润（柱状图）
    # ==========================================
    if m_id == "key_underwriting_profit":
        if not print_mode:
            c1, c2, c3 = st.columns(3)
            with c1:
                show_labels = st.toggle("显示标签", True, key=f"lab_{m_id}")
            with c2:
                pct_sz = st.slider("涨幅字号", 8, 24, 14, key=f"psz_{m_id}")
            with c3:
                gap = st.slider("柱间距", 0.1, 0.8, 0.15, key=f"gap_{m_id}")
        else:
            show_labels = st.session_state.get(f"lab_{m_id}", True)
            pct_sz = st.session_state.get(f"psz_{m_id}", 14)
            gap = st.session_state.get(f"gap_{m_id}", 0.15)
        
        fig = create_kpmg_chart(df_filtered, "承保利润", "", show_labels, pct_sz, gap,
                                sort_by_value=False, is_percentage=False)
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "承保利润")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 关键指标 - 投资利润（柱状图）
    # ==========================================
    if m_id == "key_investment_profit":
        if not print_mode:
            c1, c2, c3 = st.columns(3)
            with c1:
                show_labels = st.toggle("显示标签", True, key=f"lab_{m_id}")
            with c2:
                pct_sz = st.slider("涨幅字号", 8, 24, 14, key=f"psz_{m_id}")
            with c3:
                gap = st.slider("柱间距", 0.1, 0.8, 0.15, key=f"gap_{m_id}")
        else:
            show_labels = st.session_state.get(f"lab_{m_id}", True)
            pct_sz = st.session_state.get(f"psz_{m_id}", 14)
            gap = st.session_state.get(f"gap_{m_id}", 0.15)
        
        fig = create_kpmg_chart(df_filtered, "投资利润", "", show_labels, pct_sz, gap,
                                sort_by_value=False, is_percentage=False)
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "投资利润")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 关键指标 - 净利润（柱状图）
    # ==========================================
    if m_id == "key_net_profit":
        if not print_mode:
            c1, c2, c3 = st.columns(3)
            with c1:
                show_labels = st.toggle("显示标签", True, key=f"lab_{m_id}")
            with c2:
                pct_sz = st.slider("涨幅字号", 8, 24, 14, key=f"psz_{m_id}")
            with c3:
                gap = st.slider("柱间距", 0.1, 0.8, 0.15, key=f"gap_{m_id}")
        else:
            show_labels = st.session_state.get(f"lab_{m_id}", True)
            pct_sz = st.session_state.get(f"psz_{m_id}", 14)
            gap = st.session_state.get(f"gap_{m_id}", 0.15)
        
        fig = create_kpmg_chart(df_filtered, "净利润", "", show_labels, pct_sz, gap,
                                sort_by_value=False, is_percentage=False)
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "净利润")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return

    # ==========================================
    # 关键指标 - 利润构成（承保 vs 投资 堆叠图）
    # ==========================================
    if m_id == "key_profit_composition":
        available_years = sorted([int(y) for y in df_filtered['报告年份'].unique() if str(y).isdigit()])
        if len(available_years) >= 2:
            years = [available_years[-2], available_years[-1]]
        else:
            years = [available_years[-1], available_years[-1]] if available_years else [2024, 2025]
        fig = create_profit_stacked_pct_chart(
            df_filtered, selected_cos, years, divisor, unit_label, current_hl
        )
        show_chart(fig, print_mode, m_id)
        display_notes(m_id, df_filtered, "利润构成")
        display_bottom_note(notes_dict.get(m_id, {}).get('note', ''))
        return
    # ==========================================
    # 未匹配
    # ==========================================
    st.warning(f"未识别的模块 ID：{m_id}，请检查 Excel 模板。")

# 7.2 统一渲染出口
def render_report_module(m_id, print_mode, is_first=False):
    global notes_dict  # ✅ 添加这一行
    """
    统一渲染出口：标题 + 手动截图覆盖 + 注释框 + 图表 + 底部注释
    内部调用 render_pure_chart_entity 绘制图表
    """
    mod_data = notes_dict.get(m_id, {})
    
    # ---- 1. 生成完整标题（从 notes_dict 读取） ----
    full_title = mod_data.get('title', m_id)
    if 'df_notes' in st.session_state and isinstance(st.session_state['df_notes'], pd.DataFrame):
        df_n = st.session_state['df_notes']
        if '模块ID' in df_n.columns:
            match = df_n[df_n['模块ID'] == m_id]
            if not match.empty:
                r = match.iloc[0]
                title_parts = []
                for field in ['一级分类', '二级分类', '对应图表名称']:
                    val = str(r.get(field, '')).strip()
                    if val and val.lower() != 'nan' and val != '全部':
                        title_parts.append(val)
                if title_parts:
                    full_title = " - ".join(title_parts)
    
    # ---- 2. 打印模式容器 ----
    if print_mode:
        st.markdown("<div class='page-break-container' style='margin:0;padding:0;page-break-inside:avoid;'>", unsafe_allow_html=True)
    if not print_mode:
        st.markdown(
            "<div class='no-print' style='height:2px; background:linear-gradient(to right, #00338D, #0865EE, #00338D); opacity:0.3; margin:0;'></div>",
            unsafe_allow_html=True
        )
    
    # ---- 3. 标题（打印和网页都执行） ----
    title_cls = "page-break-title" if (print_mode and not is_first) else ""
    mt = "0px" if (print_mode and is_first) else "20px"
    font_size = "35px" if print_mode else "30px"
    st.markdown(
        f"<h3 class='{title_cls}' style='"
        f"text-align:left; color:#00338D; font-size:{font_size}; font-weight:900; "
        f"font-family:Microsoft YaHei, 微软雅黑, sans-serif; "
        f"margin-top:{mt}; margin-bottom:20px; border:none; padding-bottom:0px;'>"
        f"{full_title}</h3>",
        unsafe_allow_html=True
    )
    
    # ---- 4. 手动截图覆盖 ----
    if 'manual_upload_images' in st.session_state and m_id in st.session_state.manual_upload_images:
        if print_mode:
            st.image(st.session_state.manual_upload_images[m_id], use_column_width=True)
        else:
            img_col_left, img_col_center, img_col_right = st.columns([1, 8, 1])
            with img_col_center:
                st.image(st.session_state.manual_upload_images[m_id], use_column_width=True)
        # 显示注释（display_notes 内部已包含注释显示，无需再单独调用 display_bottom_note）
        display_notes(m_id)
        if print_mode:
            st.markdown("</div>", unsafe_allow_html=True)
        return
    
    # ---- 5. AI / 注释框（预调用，获取注释文本，并显示注释） ----
    # display_notes 内部会生成 AI 洞察并显示分析内容，同时负责显示注释
    display_notes(m_id, ai_df=st.session_state.get('df_filtered', pd.DataFrame()), ai_field=mod_data.get('title', m_id))
    
    # ---- 6. 图表（区分打印/网页模式） ----
    if print_mode:
        render_pure_chart_entity(m_id, print_mode)
    else:
        chart_col_left, chart_col_center, chart_col_right = st.columns([1, 10, 1])
        with chart_col_center:
            render_pure_chart_entity(m_id, print_mode)
    
    # ---- 7. 不再单独调用 display_bottom_note，避免重复 ----
    # 注释已由第5步的 display_notes 统一显示
    
    if print_mode:
        st.markdown("</div>", unsafe_allow_html=True)
        
# 8.税务分析图表
def create_tax_subplot_chart(df_pivot, selected_cos, show_labels, bar_width, label_size, co_font_size, co_y_offset, highlight_co="无"):
    av_cos = [c for c in selected_cos if c in df_pivot['公司'].unique()]
    if not av_cos:
        return go.Figure()
    
    cols = ['税前利润总额', '净利润']
    g_max = df_pivot[cols].max().max() / divisor
    g_min = df_pivot[cols].min().min() / divisor
    y_top = g_max * 1.25 if pd.notna(g_max) else 100
    y_bot = min(0, g_min * 1.2) if pd.notna(g_min) else 0
    lbl_y = g_max * 1.15 if pd.notna(g_max) else 100
    ph_h = g_max * 0.4 if pd.notna(g_max) else 10
    all_yrs = sorted(df_pivot['报告年份'].astype(str).unique())
    
    n = len(av_cos)
    fig = make_subplots(rows=1, cols=n, shared_yaxes=True, 
                        column_titles=[f"<b><span style='color:#00338D;'>{c}</span></b>" for c in av_cos], 
                        horizontal_spacing=0.03 if n <= 1 else min(0.03, 0.8 / (n - 1)))
    
    for nm, c in zip(cols, ['#1E49E2', '#C7A0F7']):
        fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers", 
                                 marker=dict(symbol="square", size=12, color=c), 
                                 name=nm, showlegend=True), row=1, col=1)

    for i, co in enumerate(av_cos):
        ci = i + 1
        d = df_pivot[df_pivot['公司'] == co].sort_values('报告年份')
        x = d['报告年份'].astype(str).tolist()
        yp = (d[cols[0]] / divisor).tolist()
        yn = (d[cols[1]] / divisor).tolist()
        tr = d['有效税率'].tolist()
        
        pr, nr, pt, nt, my, mt = [], [], [], [], [], []
        for vp, vn in zip(yp, yn):
            if (pd.isna(vp) or vp == 0) and (pd.isna(vn) or vn == 0):
                pr.append(0); nr.append(0); pt.append(""); nt.append(""); my.append(ph_h); mt.append("未披露")
            else:
                pr.append(vp); nr.append(vn)
                pt.append(f"{vp:.1f}" if show_labels and pd.notna(vp) else "")
                nt.append(f"{vn:.1f}" if show_labels and pd.notna(vn) else "")
                my.append(0); mt.append("")
        
        fig.add_trace(go.Bar(x=x, y=pr, marker_color='#1E49E2', text=pt, 
                             textposition='outside', textfont=dict(size=label_size), 
                             width=bar_width, offsetgroup=1, showlegend=False, cliponaxis=False), row=1, col=ci)
        fig.add_trace(go.Bar(x=x, y=nr, marker_color='#C7A0F7', text=nt, 
                             textposition='outside', textfont=dict(size=label_size), 
                             width=bar_width, offsetgroup=2, showlegend=False, cliponaxis=False), row=1, col=ci)
        fig.add_trace(go.Bar(x=x, y=my, marker_color="#CDCDCD", text=mt, 
                             textposition='inside', insidetextanchor='middle', 
                             textfont=dict(size=12, color="white"), 
                             width=bar_width * 2, showlegend=False, cliponaxis=False, hoverinfo='skip'), row=1, col=ci)
        
        for j, yr in enumerate(x):
            if pd.notna(tr[j]) and mt[j] == "":
                fig.add_annotation(x=yr, y=lbl_y, text=f"<b>{tr[j]:.0%}</b>", 
                                   showarrow=False, font=dict(size=label_size + 2, color="#97014F" if tr[j] >= 0 else "#269924"),
                                   xref=f"x{ci}" if ci > 1 else "x", yref="y1")
        
        hl = (str(co).strip() == str(highlight_co).strip())
        fig.add_shape(type="rect", xref=f"x{ci} domain" if ci > 1 else "x domain", yref="paper",
                      x0=-0.05, x1=1.05, y0=-0.12, y1=1.1,
                      line=dict(color=HL_BOX_LINE if hl else "rgba(200,200,200,0.3)", width=1.5 if hl else 1),
                      fillcolor=HL_BOX_FILL if hl else "rgba(0,0,0,0)", layer="above")

    fig.update_layout(height=550, margin=dict(t=40, b=100, l=20, r=20),
                      plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                      legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5))
    
    for i in range(1, n + 1):
        fig.update_xaxes(type='category', categoryorder='array', categoryarray=all_yrs,
                         showline=True, linecolor='lightgray', linewidth=1,
                         showgrid=False, tickfont=dict(size=10), zeroline=False, ticks="", ticklen=0, row=1, col=i)
        fig.update_yaxes(showline=False, showgrid=False, showticklabels=False,
                         zeroline=True, zerolinecolor="#E0E0E0", range=[y_bot, y_top], row=1, col=i)
    
    for a in fig.layout.annotations:
        if "span" in str(a.text):
            a.update(y=co_y_offset, font_size=co_font_size)
    
    return fig

# 9.财务与投资分析图表
def create_asset_composition_chart(df, selected_cos, field_map, color_map, title_prefix, show_labels, label_size, bar_width, co_font_size, highlight_co="无"):
    fields = list(field_map.keys())
    d_struct = df[(df['公司'].isin(selected_cos)) & (df['字段名'].isin(fields))].copy()
    d_struct['报告年份'] = d_struct['报告年份'].astype(str).str.replace(".0", "", regex=False) + "YE"
    all_yrs = [f"{prev_year}YE", f"{latest_year}YE"]
    av_cos = [co for co in selected_cos if co in d_struct['公司'].unique()]
    if not av_cos:
        return go.Figure()

    hl_co = str(highlight_co).strip()
    fig = make_subplots(rows=1, cols=len(av_cos), shared_yaxes=True, 
                        column_titles=[f"<span style='color:#00338D;'><b>{co}</b></span>" for co in av_cos], 
                        horizontal_spacing=0.015)
    for i, co in enumerate(av_cos):
        d_co = d_struct[d_struct['公司'] == co].pivot(index='报告年份', columns='字段名', values='(百万)人民币').reindex(all_yrs).fillna(0)
        for f in fields:
            if f not in d_co.columns:
                d_co[f] = 0
        raw_total = d_co.sum(axis=1)
        d_co['T'] = raw_total.replace(0, 1)
        
        for fn in fields:
            val_pct = d_co[fn] / d_co['T'] * 100
            is_dark = any(x in color_map.get(fn, "") for x in ["00338D", "510DBC", "1E49E2"])
            txt_c = "white" if is_dark else "black"
            fig.add_trace(go.Bar(
                x=d_co.index, y=val_pct,
                name=field_map[fn] if i == 0 else None,
                marker_color=color_map.get(fn, "#CCCCCC"),
                text=[f"{v:.0f}%" if show_labels and raw_total.iloc[idx] > 0 else "" for idx, v in enumerate(val_pct)],
                textposition='inside', insidetextanchor='middle',
                textfont=dict(size=label_size, color=txt_c),
                constraintext='none', textangle=0, cliponaxis=False,
                width=bar_width, showlegend=(i == 0), legendgroup=fn, hoverinfo="skip"
            ), row=1, col=i + 1)
        
        fig.add_trace(go.Bar(
            x=d_co.index, y=[100 if t == 0 else 0 for t in raw_total],
            name="未披露", marker_color="#CDCDCD",
            text=["未披露" if t == 0 else "" for t in raw_total],
            textposition='inside', insidetextanchor='middle',
            textfont=dict(size=label_size, color="white"),
            constraintext='none', textangle=0, cliponaxis=False,
            width=bar_width, showlegend=False, hoverinfo="skip"
        ), row=1, col=i + 1)

        is_hl = (str(co).strip() == hl_co)
        fig.add_shape(
            type="rect", xref="x domain", yref="y domain",
            x0=-0.06, x1=1.06, y0=-0.12, y1=1.12,
            fillcolor=HL_BOX_FILL if is_hl else "rgba(0,0,0,0)",
            line=dict(color=HL_BOX_LINE, width=1.5) if is_hl else dict(color="rgba(0,0,0,0)", width=0),
            layer="above", row=1, col=i + 1
        )

    fig.update_layout(
        barmode='stack', height=550,
        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
        uniformtext=dict(minsize=label_size, mode='show'),
        margin=dict(t=50, b=120, l=40, r=40),
        legend=dict(orientation="h", yanchor="top", y=-0.2, xanchor="center", x=0.5,
                    font=dict(size=11), itemsizing="constant")
    )
    for i in range(1, len(av_cos) + 1):
        fig.update_xaxes(showgrid=False, showline=False, zeroline=False,
                         tickfont=dict(size=10), ticks="", ticklen=0, row=1, col=i)
        fig.update_yaxes(showgrid=False, range=[0, 101], showline=False, zeroline=False,
                         tickvals=[0, 25, 50, 75, 100], ticktext=["0%", "25%", "50%", "75%", "100%"],
                         showticklabels=(i == 1), row=1, col=i)
    for ann in fig.layout.annotations:
        ann.update(y=ann.y + 0.05, font_size=co_font_size)
    return fig

def create_oci_chart(df_raw, year, title, show_labels, co_font_size, bar_gap, selected_cos, highlight_co="无"):
    oci_fields = ['可转损益OCI合计', '不可转损益OCI合计', '净利润', '综合收益总额']
    d_sub = df_raw[(df_raw['字段名'].isin(oci_fields)) & 
                   (df_raw['报告年份'].astype(str) == str(year).replace(".0", ""))].drop_duplicates(
                       subset=['公司', '报告年份', '字段名'], keep='last').copy()
    if d_sub.empty:
        return go.Figure().update_layout(paper_bgcolor='rgba(0,0,0,0)')

    df_pivot = d_sub.pivot_table(index='公司', columns='字段名', values='(百万)人民币', aggfunc='sum').fillna(0)
    for f in oci_fields:
        if f not in df_pivot.columns:
            df_pivot[f] = 0.0
    df_pivot = df_pivot / divisor
    df_pivot['other'] = df_pivot['综合收益总额'] - df_pivot['净利润'] - df_pivot['可转损益OCI合计'] - df_pivot['不可转损益OCI合计']
    mc = {
        "净利润": {"c": "rgb(0,176,240)", "n": "净利润"},
        "可转损益OCI合计": {"c": "rgb(253,52,156)", "n": "可转损益OCI变动"},
        "不可转损益OCI合计": {"c": "rgb(114,19,234)", "n": "不可转损益OCI变动"},
        "other": {"c": "rgb(127,127,127)", "n": "其他"},
        "综合收益总额": {"c": "rgb(172,234,255)", "n": "综合收益总额"}
    }

    av_cos = [c for c in selected_cos if c in df_pivot.index]
    hl_co = str(highlight_co).strip()
    if not av_cos:
        return go.Figure()
    all_vals = df_pivot[['净利润', '可转损益OCI合计', '不可转损益OCI合计', 'other', '综合收益总额']].values.flatten()
    max_val = np.nanmax(all_vals)
    min_val = np.nanmin(all_vals)
    buffer = (max_val - min_val) * 0.8 if max_val != min_val else abs(max_val) * 0.3
    y_range = [min_val - buffer if min_val < 0 else min_val - abs(min_val) * 0.3, max_val + buffer]

    fig = make_subplots(rows=1, cols=len(av_cos), shared_yaxes=True, horizontal_spacing=0.015,
                        subplot_titles=[f"<b><span style='color:#00338D;'>{co}</span></b>" for co in av_cos])
    for col_idx, co in enumerate(av_cos):
        for m_key, m_info in mc.items():
            val = df_pivot.loc[co].get(m_key, 0)
            fig.add_trace(go.Bar(
                name=m_info["n"], x=[m_key], y=[val],
                text=[f"{val:.0f}" if (show_labels and val != 0) else ""],
                textposition='outside', textfont=dict(size=11, color='#00338D'),
                marker_color=m_info["c"], width=0.8,
                legendgroup=m_key, showlegend=(col_idx == 0),
                cliponaxis=False, constraintext='none'
            ), row=1, col=col_idx + 1)

        is_hl = (str(co).strip() == hl_co)
        bg_fill = HL_BOX_FILL if is_hl else "rgba(240, 240, 240, 0.35)"
        line_dict = dict(color=HL_BOX_LINE, width=1.5) if is_hl else dict(color="rgba(210, 210, 210, 0.6)", width=1)
        fig.add_shape(
            type="rect", xref="x domain", yref="y domain",
            x0=-0.04, x1=1.04, y0=-0.04, y1=1.13,
            fillcolor=bg_fill, line=line_dict, layer="above" if is_hl else "below",
            row=1, col=col_idx + 1
        )
        fig.update_xaxes(showticklabels=False, showline=False, zeroline=False,
                         showgrid=False, ticks="", ticklen=0, row=1, col=col_idx + 1)

    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
        barmode='group', bargap=bar_gap,
        height=420,
        margin=dict(t=50, b=40, l=40, r=30),
        legend=dict(orientation="h", yanchor="top", y=-0.2, xanchor="center", x=0.5, font=dict(size=11))
    )
    fig.update_yaxes(range=y_range, showline=False, zeroline=False, showgrid=False,
                     gridcolor='rgba(0,0,0,0)', gridwidth=0, row=1, col="all")
    for ann in fig.layout.annotations:
        ann.update(y=1.03, font_size=co_font_size)
    return fig

def create_asset_liab_oci_chart(df_raw, selected_cos, bar_gap, co_font_size, show_labels, highlight_co="无"):
    yrs = [str(prev_year), str(latest_year)]
    fns = ['可转损益的负债OCI', 'FVOCI债券公允价值']
    d = df_raw[df_raw['字段名'].isin(fns)].copy()
    d['v'] = d['(百万)人民币'] / divisor
    av_cos = [c for c in selected_cos if c in d['公司'].unique()]
    if not av_cos:
        return go.Figure()

    def gv(c, y, f):
        return d[(d['公司'] == c) & (d['报告年份'].astype(str) == y) & (d['字段名'] == f)]['v'].sum()

    all_v = [gv(c, y, f) for c in av_cos for y in yrs for f in fns]
    g_max = max(all_v + [0.1])
    g_min = min(all_v + [-0.1])
    ph_h = max(abs(g_max), abs(g_min)) * 0.4
    n = len(av_cos)
    colors = {fns[0]: "rgb(0, 184, 245)", fns[1]: "rgb(253, 52, 156)"}

    fig = make_subplots(rows=1, cols=n, shared_yaxes=True,
                        horizontal_spacing=0.01 if n <= 1 else min(0.01, 0.8 / (n - 1)),
                        subplot_titles=[f"<b><span style='color:#00338D;'>{c}</span></b>" for c in av_cos])
    for fn, c in colors.items():
        fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers",
                                 marker=dict(symbol="square", size=12, color=c),
                                 name=fn, showlegend=True), row=1, col=1)

    for i, co in enumerate(av_cos):
        ci = i + 1
        x_lbl = [f"{y}YE" for y in yrs]
        v1, v2, t1, t2, my = [], [], [], [], []
        for y in yrs:
            val1, val2 = gv(co, y, fns[0]), gv(co, y, fns[1])
            m = (val1 == 0 and val2 == 0)
            v1.append(0 if m else val1)
            v2.append(0 if m else val2)
            my.append(ph_h if m else 0)
            t1.append("" if m else (f"{val1:.0f}" if show_labels and val1 != 0 else ""))
            t2.append("" if m else (f"{val2:.0f}" if show_labels and val2 != 0 else ""))
            if m:
                fig.add_annotation(x=f"{y}YE", y=ph_h / 2, text="未披露", showarrow=False,
                                   font=dict(color="white", size=12),
                                   xref=f"x{ci}" if ci > 1 else "x", yref="y1",
                                   xanchor="center", yanchor="middle")

        c1_list = ["#ACEAFF", colors[fns[0]]]
        c2_list = ["#FFD6E8", colors[fns[1]]]

        fig.add_trace(go.Bar(x=x_lbl, y=v1, marker_color=c1_list, text=t1,
                             textposition='outside', textfont=dict(size=11, color='#00338D'),
                             width=0.4, offsetgroup=1, showlegend=False,
                             cliponaxis=False, constraintext='none'), row=1, col=ci)
        fig.add_trace(go.Bar(x=x_lbl, y=v2, marker_color=c2_list, text=t2,
                             textposition='outside', textfont=dict(size=11, color='#00338D'),
                             width=0.4, offsetgroup=2, showlegend=False,
                             cliponaxis=False, constraintext='none'), row=1, col=ci)
        fig.add_trace(go.Bar(x=x_lbl, y=my, marker_color="#CDCDCD",
                             width=0.8, offset=-0.4, showlegend=False,
                             cliponaxis=False, hoverinfo='skip'), row=1, col=ci)

        hl = (str(co).strip() == str(highlight_co).strip())
        xref_str = f"x{ci} domain" if ci > 1 else "x domain"
        fig.add_shape(type="rect", xref=xref_str, yref="y domain",
                      x0=-0.06, x1=1.06, y0=0, y1=1.08,
                      fillcolor=HL_BOX_FILL if hl else "rgba(245,245,245,0.5)",
                      line=dict(color=HL_BOX_LINE if hl else "#CCCCCC", width=1.5 if hl else 1),
                      layer="below")

    fig.update_layout(barmode='group', bargap=bar_gap,
                      paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                      height=420,
                      margin=dict(t=50, b=80, l=20, r=20),
                      legend=dict(orientation="h", yanchor="top", y=-0.28, x=0.5, xanchor="center"))
    y_rng = [g_min - abs(g_min) * 0.3, g_max + abs(g_max) * 0.3]

    for i in range(1, n + 1):
        fig.update_xaxes(type='category', categoryorder='array', categoryarray=x_lbl,
                         showline=False, zeroline=False, showgrid=False, ticks="", ticklen=0, row=1, col=i)
        fig.update_yaxes(range=y_rng, showline=False, zeroline=True, zerolinecolor="#E0E0E0",
                         zerolinewidth=1.02, showgrid=False, gridcolor='rgba(0,0,0,0)',
                         gridwidth=0, row=1, col=i)
    for ann in fig.layout.annotations:
        if "span" in str(ann.text):
            ann.update(y=1.08, font_size=co_font_size)

    return fig

def create_equity_change_detail_table(df_raw, target_year, cos, divisor=1, unit_label="", highlight_co="无"):
    df = df_raw.copy()
    df["报告年份"] = df["报告年份"].astype(str).str.replace(".0", "", regex=False).str.strip()
    df["公司"] = df["公司"].astype(str).str.strip()
    df["字段名"] = df["字段名"].astype(str).str.strip()

    year_str = str(target_year)

    fields = [
        ("1", "净利润", "净利润"),
        ("2", "其他综合收益", "其他综合收益"),
        ("3", "股东权益分配", "股东权益分配"),
        ("4", "股东增资", "股东增资"),
        ("5", "发债", "发债"),
        ("6", "子公司合并", "子公司合并"),
        ("7", "其他权益变动", "其他权益变动"),
        ("8=sum(1-7)", "净资产变动值", "净资产变动值"),
    ]

    val_col = "(百万)人民币" if "(百万)人民币" in df.columns else df.columns[-1]

    d = df[
        (df["报告年份"] == year_str) &
        (df["公司"].isin(cos)) &
        (df["字段名"].isin([x[2] for x in fields]))
    ].drop_duplicates(subset=["公司", "报告年份", "字段名"], keep="first").copy()

    def to_num(v):
        if pd.isna(v):
            return np.nan
        s = str(v).strip()
        if s.lower() in ["", "nan", "none", "null", "未披露", "—", "-"]:
            return np.nan
        try:
            return float(
                s.replace(",", "")
                 .replace("(", "-")
                 .replace(")", "")
            )
        except:
            return np.nan

    data_map = {}
    for _, r in d.iterrows():
        data_map[(str(r["公司"]).strip(), str(r["字段名"]).strip())] = to_num(r.get(val_col, np.nan))

    for co in cos:
        co = str(co).strip()
        raw_net_change = data_map.get((co, "净资产变动值"), np.nan)
        if pd.isna(raw_net_change):
            vals = [data_map.get((co, f[2]), np.nan) for f in fields[:7]]
            vals_valid = [x for x in vals if pd.notna(x)]
            if vals_valid:
                data_map[(co, "净资产变动值")] = sum(vals_valid)

    def fmt(v):
        if pd.isna(v):
            return "未披露", True
        v = v / divisor
        if abs(v) < 1e-9:
            return "0.0", False
        return f"{v:,.1f}", False

    current_hl = str(highlight_co).strip()

    html = (
        "<style>"
        "@media print {"
        ".equity-detail-wrap { margin-top:4px!important; margin-bottom:6px!important; }"
        ".equity-detail-title { font-size:10px!important; margin-bottom:3px!important; }"
        ".equity-detail-table { font-size:8.5px!important; }"
        ".equity-detail-table th { padding:3px 4px!important; line-height:1.15!important; }"
        ".equity-detail-table td { padding:2.5px 4px!important; line-height:1.15!important; }"
        "}"
        "</style>"
        "<div class='equity-detail-wrap' style='margin-top:8px; margin-bottom:10px;'>"
        f"<div class='equity-detail-title' style='font-size:13px; font-weight:bold; color:#00338D; margin-bottom:5px;'>"
        f"{year_str}年净资产变动明细表"
        + (f"（单位：{unit_label}）" if unit_label else "")
        + "</div>"
    )

    html += (
        "<table class='equity-detail-table' style='width:100%; border-collapse:collapse; "
        "font-family:Microsoft YaHei, sans-serif; font-size:11px; table-layout:fixed;'>"
    )

    html += (
        "<tr style='background-color:#00338D; color:white; text-align:center; font-weight:bold;'>"
        "<th style='padding:5px 6px; border:1px solid white; width:10%;'>序号</th>"
        "<th style='padding:5px 6px; border:1px solid white; width:16%;'>项目</th>"
    )
    for co in cos:
        co_str = str(co).strip()
        html += f"<th style='padding:5px 6px; border:1px solid white;'>{co_str}</th>"
    html += "</tr>"

    for row_idx, (no, display_name, field_name) in enumerate(fields):
        row_bg = "#F8F9FA" if row_idx % 2 == 0 else "white"

        html += f"<tr style='background-color:{row_bg};'>"
        html += f"<td style='padding:4px 6px; text-align:center; border:1px solid #EAEAEA; color:#333;'>{no}</td>"
        html += f"<td style='padding:4px 6px; text-align:left; border:1px solid #EAEAEA; font-weight:bold; color:#333;'>{display_name}</td>"

        for co in cos:
            co_str = str(co).strip()
            val, missing = fmt(data_map.get((co_str, field_name), np.nan))

            if co_str == current_hl and not missing:
                cell_style = (
                    "background-color:rgba(0,51,141,0.08); color:#00338D; "
                    "font-weight:bold; border:1px solid #00338D;"
                )
            elif missing:
                cell_style = "background-color:#CDCDCD; color:white; border:1px solid #EAEAEA;"
            else:
                cell_style = "color:#333; border:1px solid #EAEAEA;"

            html += f"<td style='padding:4px 6px; text-align:center; {cell_style}'>{val}</td>"

        html += "</tr>"

    html += "</table></div>"
    return html

def create_expense_structure_chart(df, selected_cos, show_labels, label_size, bar_width, co_font_size, highlight_co="无"):
    prev_year = st.session_state.get('prev_year', 2024)
    latest_year = st.session_state.get('latest_year', 2025)
    years = [str(prev_year), str(latest_year)]
    field_map = {
        "获取费用摊销": "获取费用摊销",
        "维持费用": "维持费用",
    }
    fields = list(field_map.keys())
    
    raw_df = df.copy()
    raw_df.columns = raw_df.columns.str.strip()
    raw_df['公司'] = raw_df['公司'].astype(str).str.strip()
    raw_df['字段名'] = raw_df['字段名'].astype(str).str.strip()
    raw_df['报告年份'] = raw_df['报告年份'].astype(str).str.replace('.0', '', regex=False).str.strip()
    raw_df = raw_df[raw_df['报告年份'] != '']
    raw_df = raw_df[~raw_df['报告年份'].str.lower().isin(['nan', 'none'])]
    
    # ===== 修改点 1：分母取数限定类别为 '保费收入' =====
    service_revenue = {}
    for co in selected_cos:
        co_clean = co.strip()
        for yr in years:
            mask = (raw_df['公司'] == co_clean) & \
                   (raw_df['报告年份'] == yr) & \
                   (raw_df['字段名'] == '保险服务收入') & \
                   (raw_df['类别'] == '保费收入')   # ← 新增类别过滤
            rev_series = raw_df.loc[mask, '(百万)人民币']
            rev = rev_series.sum() if not rev_series.empty else 0
            if pd.isna(rev) or rev == 0:
                mask2 = (raw_df['公司'] == co_clean) & \
                        (raw_df['报告年份'] == yr) & \
                        (raw_df['字段名'] == '保险业务收入') & \
                        (raw_df['类别'] == '保费收入')   # ← 新增类别过滤
                rev_series2 = raw_df.loc[mask2, '(百万)人民币']
                rev = rev_series2.sum() if not rev_series2.empty else 0
            if rev == 0:
                st.warning(f"⚠️ 公司 {co} 在 {yr} 年未找到保险服务收入或保险业务收入！")
                rev = 1
            service_revenue[(co, yr)] = rev
    
    fig = make_subplots(rows=1, cols=len(selected_cos), horizontal_spacing=0.02,
                        subplot_titles=[f"<span style='color:#00338D;'><b>{co}</b></span>" for co in selected_cos])
    
    for i, co in enumerate(selected_cos):
        co_clean = co.strip()
        # ===== 修改点 2：取费用因子时限定类别为 '综合成本率拆解' =====
        cd = raw_df[(raw_df['公司'] == co_clean) & (raw_df['类别'] == '综合成本率拆解')]
        vals = {}
        for f in fields:
            vals[f] = {}
            for yr in years:
                mask_f = (cd['报告年份'] == yr) & (cd['字段名'] == f)
                v_series = cd.loc[mask_f, '(百万)人民币']
                v = v_series.sum() if not v_series.empty else 0
                denominator = service_revenue.get((co, yr), 1)
                ratio = v / denominator * 100 if denominator != 0 else 0
                vals[f][yr] = ratio if pd.notna(ratio) else 0
        
        x_axis = [f"{y}YE" for y in years]
        for f in fields:
            y_vals = [vals[f].get(y, 0) for y in years]
            fig.add_trace(go.Bar(
                x=x_axis, y=y_vals,
                name=field_map[f],
                marker_color="#00338D" if f == "获取费用摊销" else "#0865EE",
                width=bar_width,
                showlegend=(i == 0),
                text=[f"{v:.1f}%" if show_labels and v != 0 else "" for v in y_vals],
                textposition='inside',
                textfont=dict(size=label_size, color="white"),
                constraintext='none'
            ), row=1, col=i+1)
        
        for yr in years:
            total = sum(vals[f].get(yr, 0) for f in fields)
            if total > 0:
                fig.add_annotation(
                    x=f"{yr}YE", y=total + 2,
                    text=f"<b>{total:.1f}%</b>",
                    showarrow=False,
                    font=dict(size=label_size, color="#222"),
                    row=1, col=i+1
                )
    
    fig.update_layout(
        barmode='stack',
        height=550,
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        margin=dict(t=60, b=80, l=10, r=10),
        legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5)
    )
    fig.update_yaxes(tickformat=".0f", title_text="占保险服务收入比例", row=1, col=1)
    return fig, (0,0,0)

def create_ratio_trend_chart(df, cos, factor_list, ratio_field, title_prefix,
                             divisor=1, unit_label="", highlight_co="无",
                             show_labels=True, line_width=2, ratio_first=False):
    """
    每个指标生成一个独立折线图，每条线代表一家公司。
    只绘制有分母（保险服务收入或保险业务收入）的年份，确保比率有意义。
    """
    raw = df.copy()
    raw['报告年份'] = raw['报告年份'].astype(str).str.replace('.0', '', regex=False)
    raw['公司'] = raw['公司'].astype(str).str.strip()
    raw['字段名'] = raw['字段名'].astype(str).str.strip()

    needed = factor_list + [ratio_field, '保险服务收入', '保险业务收入']
    raw = raw[raw['公司'].isin(cos) & raw['字段名'].isin(needed)]

    years_raw = raw['报告年份'].unique()
    years = sorted(
        [y for y in years_raw if str(y).strip() not in ['', 'nan', 'NaN', 'None']],
        key=lambda x: int(x)
    )
    if not years:
        return []

    # 计算分母
    denom = {}
    for co in cos:
        for yr in years:
            v = raw[(raw['公司']==co) & (raw['报告年份']==yr) & (raw['字段名']=='保险服务收入')]['(百万)人民币']
            if not v.empty and v.iloc[0] != 0:
                denom[(co, yr)] = v.iloc[0]
            else:
                v2 = raw[(raw['公司']==co) & (raw['报告年份']==yr) & (raw['字段名']=='保险业务收入')]['(百万)人民币']
                if not v2.empty and v2.iloc[0] != 0:
                    denom[(co, yr)] = v2.iloc[0]
                else:
                    denom[(co, yr)] = None

    # 构建数据
    all_items = [ratio_field] + factor_list if ratio_first else factor_list + [ratio_field]
    metric_data = {item: {co: {} for co in cos} for item in all_items}

    for co in cos:
        for yr in years:
            den = denom.get((co, yr))
            for f in factor_list:
                if den is None:
                    metric_data[f][co][yr] = np.nan
                else:
                    s = raw[(raw['公司']==co) & (raw['报告年份']==yr) & (raw['字段名']==f)]['(百万)人民币']
                    val = s.iloc[0] if not s.empty else 0
                    metric_data[f][co][yr] = val / den if den != 0 else np.nan
            s_ratio = raw[(raw['公司']==co) & (raw['报告年份']==yr) & (raw['字段名']==ratio_field)]['(百万)人民币']
            metric_data[ratio_field][co][yr] = s_ratio.iloc[0] if not s_ratio.empty else np.nan

    # 颜色：使用自定义强对比色（从 KPMG 色板中精选）
    custom_colors = ["#00338D", "#510DBC", "#FD349C", "#00C0AE", "#FB8E7E", "#7213EA"]
    company_colors = {co: custom_colors[i % len(custom_colors)] for i, co in enumerate(cos)}
    marker_symbols = ['circle', 'square', 'diamond', 'triangle-up', 'star', 'pentagon', 'x', 'cross']

    figs = []
    for item in all_items:   # 按调整后的顺序循环
        fig = go.Figure()
        is_ratio = (item == ratio_field)
        title = f"{title_prefix} - {item}" if title_prefix else item
        if is_ratio:
            title = f"{title_prefix} - {item}（综合）" if title_prefix else f"{item}（综合）"

        all_y = []

        for idx, co in enumerate(cos):
            data = metric_data[item][co]
            valid_years = [yr for yr in years if not np.isnan(data.get(yr, np.nan))]
            if not valid_years:
                continue
            x_vals = valid_years
            y_vals = [data[yr] for yr in valid_years]
            all_y.extend(y_vals)

            color = company_colors[co]
            is_hl = (str(co).strip() == str(highlight_co).strip())
            width = line_width + 2 if is_hl else line_width
            marker_size = 10 if is_hl else 8
            symbol = marker_symbols[idx % len(marker_symbols)]

            fig.add_trace(go.Scatter(
                x=x_vals,
                y=y_vals,
                mode='lines+markers' + ('+text' if show_labels else ''),
                name=co,
                line=dict(width=width, color=color),
                marker=dict(size=marker_size, color=color, symbol=symbol,
                            line=dict(width=1.5, color='white')),
            ))

        if not all_y:
            fig.update_layout(title=title, height=400)
            figs.append(fig)
            continue

        y_min = min(all_y)
        y_max = max(all_y)
        range_width = y_max - y_min
        if range_width == 0:
            range_width = 1
        padding = range_width * 0.1
        y_lower = y_min - padding
        y_upper = y_max + padding

        fig.update_layout(
            title=title,
            xaxis_title="年份",
            yaxis_title="比率",
            xaxis=dict(
                type='category',
                tickmode='array',
                tickvals=years,
                ticktext=years,
                gridcolor='rgba(0,0,0,0)',   # 去掉网格线
                gridwidth=0,
            ),
            yaxis=dict(
                tickformat=".2%",
                range=[y_lower, y_upper],
                gridcolor='rgba(0,0,0,0)',   # 去掉网格线
                gridwidth=0,
                zeroline=True,
                zerolinecolor='gray',
                zerolinewidth=1,
            ),
            height=400,
            margin=dict(t=50, b=50, l=40, r=20),
            legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5,
                        font=dict(size=11)),
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)'
        )
        figs.append(fig)

    return figs


def create_nonlife_six_dimensional_charts(df_raw, target_year, cos, divisor=1, unit_label="百万元", 
                                          highlight_co="无", label_size=11, show_labels=False, dot_size=11):
    df = df_raw.copy()
    df[["字段名", "公司", "报告年份"]] = df[["字段名", "公司", "报告年份"]].astype(str).apply(
        lambda x: x.str.strip().str.replace(".0", "", regex=False))
    
    needed = ["净利润", "期初股东权益", "期末股东权益", "总资产", "投资收益率", "综合成本率", "保险业务收入"]
    needed_prev = ["保险业务收入"]
    
    df = df[(df["报告年份"] == str(target_year)) & df["公司"].isin(cos) & df["字段名"].isin(needed)].drop_duplicates(
        subset=["公司", "报告年份", "字段名"])
    
    df_prev = df_raw[(df_raw["报告年份"] == str(int(target_year)-1)) & df_raw["公司"].isin(cos) & 
                     df_raw["字段名"].isin(needed_prev)].drop_duplicates(subset=["公司", "报告年份", "字段名"])
    
    if df.empty:
        return []
    
    a_cols = ["(百万)人民币", "(亿元)人民币", "(百万)原币", "(亿元)原币", "人民币", "原币", "数值", "值"]
    r_cols = ["(%)原币", "(%)", "百分比", "比例", "占比", "比率"]
    
    def get_v(r, is_ratio):
        for c in (r_cols + a_cols if is_ratio else a_cols + r_cols):
            if c in r.index and pd.notna(r[c]) and str(r[c]).strip().lower() not in ["", "nan", "none", "null", "-"]:
                s = str(r[c]).strip().replace(",", "")
                try:
                    v = float(s[:-1])/100.0 if s.endswith("%") else float(s)
                    return v/100.0 if is_ratio and ("%" in str(r[c]) or abs(v) > 1) else v
                except:
                    continue
        return np.nan
    
    records = [{"公司": r["公司"], "字段名": r["字段名"], "数值": get_v(r, r["字段名"] in ["投资收益率", "综合成本率"])} 
               for _, r in df.iterrows()]
    df_p = pd.DataFrame(records).dropna(subset=["数值"]).pivot_table(
        index="公司", columns="字段名", values="数值", aggfunc="first").reindex(cos)
    df_p.index.name = None
    
    def get_col(n):
        return pd.to_numeric(df_p[n], errors="coerce") if n in df_p.columns else pd.Series(np.nan, index=df_p.index)
    
    net_profit = get_col("净利润")
    eq_b = get_col("期初股东权益")
    eq_e = get_col("期末股东权益")
    ta = get_col("总资产")
    inv_return = get_col("投资收益率")
    cor = get_col("综合成本率")
    
    prem_curr = get_col("保险业务收入")
    prem_prev = pd.Series(np.nan, index=cos)
    for _, r in df_prev.iterrows():
        prem_prev[str(r["公司"]).strip()] = get_v(r, False)
    prem_growth = (prem_curr - prem_prev) / prem_prev.abs() if not prem_prev.empty else pd.Series(np.nan, index=cos)
    
    pd_data = pd.DataFrame({
        "公司": df_p.index,
        "净利润": net_profit,
        "净资产": eq_e,
        "保费增长率": prem_growth,
        "投资收益率": inv_return,
        "综合成本率": cor,
        "总资产": ta,
        "期初净资产": eq_b
    })
    
    pd_data["承保利润率"] = np.where((eq_b + eq_e) != 0, net_profit / ((eq_b + eq_e) / 2), np.nan)
    pd_data["财务杠杆率"] = np.where(ta != 0, eq_e / ta, np.nan)
    pd_data["净资产增长率"] = np.where(eq_b != 0, (eq_e - eq_b) / eq_b, np.nan)
    
    cfgs = [
        ("承保利润率", "承保利润率", "净资产", ".1%"),
        ("保费增长", "保费增长率", "净资产", ".1%"),
        ("财务杠杆", "财务杠杆率", "净资产", ".1%"),
        ("投资能力", "投资收益率", "净资产", ".1%"),
        ("财务稳定", "净资产增长率", "净资产", ".1%"),
        ("承保效益", "综合成本率", "净资产", ".1%")
    ]
    
    cmap = {co: px.colors.qualitative.Plotly[i % 10] for i, co in enumerate(cos)}
    hl = str(highlight_co).strip()
    rd = divisor or 1
    figs = []
    
    for title_text, y_col, x_col, y_fmt in cfgs:
        d_plt = pd_data[["公司", y_col, x_col]].copy().replace([np.inf, -np.inf], np.nan).dropna()
        
        display_y = y_col
        if y_col == "承保利润率":
            display_y = "承保利润率（净利润/净资产平均值）"
        elif y_col == "财务杠杆率":
            display_y = "财务杠杆率（净资产/总资产）"
        elif y_col == "承保效益":
            display_y = "综合成本率（COR，越低越好）"
        
        title_html = (f"<span style='font-size:13px'><b>{title_text}</b></span><br>"
                      f"<span style='font-size:11px;color:#666'>Y轴={display_y}，X轴={x_col}</span>")
        
        fig = go.Figure()
        
        if d_plt.empty:
            fig.update_layout(title=dict(text=title_html, x=0.02), height=250,
                              margin=dict(l=20, r=15, t=40, b=10),
                              paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
            figs.append(fig)
            continue
        
        d_plt["x_p"] = pd.to_numeric(d_plt[x_col], errors="coerce") / rd
        
        for _, r in d_plt.sort_values("公司").iterrows():
            co_name = str(r["公司"])
            is_hl = (co_name == hl)
            fig.add_trace(go.Scatter(
                x=[r["x_p"]], y=[r[y_col]],
                mode="markers+text" if show_labels else "markers",
                name=co_name,
                text=[co_name] if show_labels else None,
                textposition="top center",
                textfont=dict(size=label_size, color="#333"),
                marker=dict(
                    size=dot_size * 1.45 if is_hl else dot_size,
                    color=cmap.get(co_name, "#1f77b4"),
                    line=dict(color="white", width=1.8 if is_hl else 1.2),
                    opacity=0.95
                ),
                customdata=[[f"{r['x_p']:,.2f}", f"{r[y_col]:{y_fmt}}", f"{r[x_col]:,.2f}"]],
                hovertemplate=(f"<b>{co_name}</b><br>{x_col}: %{{customdata[0]}}<br>"
                              f"{y_col}: %{{customdata[1]}}<br>{x_col}: %{{customdata[2]}}<extra></extra>")
            ))
        
        fig.update_layout(
            title=dict(text=title_html, x=0.02),
            width=260, height=250,
            margin=dict(l=20, r=20, t=45, b=15),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            showlegend=False,
            hovermode="closest",
            xaxis=dict(showgrid=True, gridcolor="rgba(180,180,180,0.2)",
                      zeroline=True, zerolinecolor="rgba(150,150,150,0.3)"),
            yaxis=dict(tickformat=y_fmt, showgrid=True, gridcolor="rgba(180,180,180,0.2)",
                      zeroline=True, zerolinecolor="rgba(150,150,150,0.3)")
        )
        
        fig.add_hline(y=0, line_dash="dash", line_color="rgba(120,120,120,0.7)", line_width=1)
        fig.add_vline(x=0, line_dash="dash", line_color="rgba(120,120,120,0.7)", line_width=1)
        figs.append(fig)
    
    return figs
# ==========================================
# 登录后主系统界面（主逻辑）
# ==========================================
api_key = st.session_state.get('api_key', "")
base_url = st.session_state.get('base_url', "https://api.deepseek.com")
model_name = st.session_state.get('model_name', "deepseek-chat")
user_role = st.session_state.get('user_role', "普通用户")

# 顶部状态栏
top_col1, top_col2 = st.columns([8, 1])
with top_col1:
    st.markdown(
        f"<div class='no-print' style='color: #6c757d; font-size: 14px; margin-top: 8px;'>"
        f"当前身份：<b>{user_role}</b> | 欢迎使用财险数智分析平台</div>",
        unsafe_allow_html=True
    )
with top_col2:
    if st.button("退出登录"):
        st.session_state['logged_in'] = False
        st.rerun()

st.markdown("<hr style='margin-top: 5px; margin-bottom: 15px;'>", unsafe_allow_html=True)
st.markdown("<h1 class='no-print' style='font-weight:700; padding-top:10px;'>财险数智・年报处理平台</h1>", unsafe_allow_html=True)

# ---------- 9. 财险公司默认列表 ----------
DEFAULT_COMPANIES = [
    {"类别": "上市", "公司": "平安产险", "类型": "财险", "链接地址": "https://property.pingan.com/gongkaixinxipilu/nianduxinxipilubaogao.shtml"},
    {"类别": "上市", "公司": "人保产险", "类型": "财险", "链接地址": "https://property.picc.com.cn/gkxx/ndxx/"},
    {"类别": "上市", "公司": "太平产险", "类型": "财险", "链接地址": "https://caixian.cntaiping.com/info-ndxxpl/"},
    {"类别": "上市", "公司": "众安产险", "类型": "财险", "链接地址": "https://www.zhongan.com/channel/public/publicInfo_ndxx2018.html"},
    {"类别": "上市", "公司": "阳光产险", "类型": "财险", "链接地址": "https://www.4000-000-000.com/#/pltable?menuId=628&type=menu"},
    {"类别": "上市", "公司": "太保财险", "类型": "财险", "链接地址": "https://property.cpic.com.cn/xccbx/gkxxbl/ndxx/?subMenu=2&inSub=1"},
    {"类别": "上市", "公司": "平安产险（偿付能力报告）", "类型": "财险", "链接地址": "https://property.pingan.com/gongkaixinxipilu/changfunenglixinxipilubaogao.shtml"},
    {"类别": "上市", "公司": "人保产险（偿付能力报告）", "类型": "财险", "链接地址": "https://property.picc.com.cn/gkxx/zxxx/cfnl/"},
    {"类别": "上市", "公司": "太平产险（偿付能力报告）", "类型": "财险", "链接地址": "https://caixian.cntaiping.com/info-cfnljdbgzy/"},
    {"类别": "上市", "公司": "众安产险（偿付能力报告）", "类型": "财险", "链接地址": "http://zhongan.com/channel/public/publicInfo_cfnl2018.html"},
    {"类别": "上市", "公司": "阳光产险（偿付能力报告）", "类型": "财险", "链接地址": "https://www.4000-000-000.com/#/pltable?menuId=633&type=menu"},
    {"类别": "上市", "公司": "太保财险（偿付能力报告）", "类型": "财险", "链接地址": "https://property.cpic.com.cn/xccbx/gkxxbl/cfnlxxzq/?subMenu=4&inSub=3"},
    {"类别": "非上市", "公司": "华泰财险", "类型": "年报", "链接地址": "https://pc.ehuatai.com/annual_info.html"}
]

# ---------- 10. Step 0~8 Tabs ----------
if st.session_state['user_role'] == "项目组成员":
    tab0, tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
        " 🌐 Step 0 ／ 官网年报监控 ",
        " 📑 Step 1 ／ 智能页码检索 ",
        " ⚡ Step 2 ／ 表格智能转换 ",
        " 📝 Step 3 ／ 目标表标准填报 ",
        " 🔍 Step 4 ／ 数据勾稽检查 ",
        " ⛓️‍💥 Step 5 ／ 多公司合并 ",
        " 📊 Step 6 ／ 自定义对标分析 ",
        " 🖼️ Step 7 ／ 公司级对标报告 ",
        " 📈 Step 8 ／ 行业分类统计分析 "
    ])
    
    # ----- Step 0：官网年报监控（财险版）-----
    with tab0:
        st.markdown("### 🌐 财险公司官网年报监控")
        
        col_t0_1, _ = st.columns([1, 2])
        with col_t0_1:
            target_year = st.number_input("📅 请选择监控年份", min_value=2010, max_value=2050, value=2025, step=1)
        
        st.markdown(f"""
        <div class="info-card green">
            <h4>功能说明</h4>
            <p>系统将自动扫描各财险公司官网，检测网页中是否出现 <b>{target_year}</b> 字样。您可以<b>双击下方表格修改网址</b>。</p>
        </div>
        """, unsafe_allow_html=True)
    
        if 'company_df' not in st.session_state:
            st.session_state.company_df = pd.DataFrame(DEFAULT_COMPANIES)
        
        st.markdown("#### 🏢 监控目标名单 (双击下方链接可修改)")
        edited_df = st.data_editor(
            st.session_state.company_df,
            column_config={
                "链接地址": st.column_config.TextColumn("🌐 网页链接地址 (支持双击编辑)", max_chars=1000, width="large"),
                "类别": st.column_config.TextColumn(disabled=True),
                "公司": st.column_config.TextColumn(disabled=True),
                "类型": st.column_config.TextColumn(disabled=True),
            },
            use_container_width=True,
            hide_index=True,
            key="company_data_editor_step0"
        )
        st.session_state.company_df = edited_df
    
        st.markdown("---")
        
        col_btn1, col_btn2 = st.columns([1, 1])
        with col_btn1:
            if not edited_df.empty:
                target_company = st.selectbox("🎯 单家快速检索：", options=edited_df['公司'].tolist())
                single_scan = st.button(f"🔍 检索【{target_company}】", use_container_width=True)
            else:
                st.warning("请先在表格中添加财险公司数据")
                single_scan = False
                target_company = None
        with col_btn2:
            st.write("")
            batch_scan = st.button("🚀 启动批量扫描", type="primary", use_container_width=True)
    
        if single_scan or batch_scan:
            if edited_df.empty:
                st.warning("⚠️ 请先在表格中添加财险公司数据。")
            else:
                tasks = edited_df.to_dict('records') if batch_scan else edited_df[edited_df['公司'] == target_company].to_dict('records')
                results = []
                my_bar = st.progress(0, text="准备启动扫描...")
                total_tasks = len(tasks)
                headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
    
                for index, row in enumerate(tasks):
                    company_name = row['公司']
                    url = str(row['链接地址'])
                    my_bar.progress((index + 1) / total_tasks, text=f"正在扫描 ({index+1}/{total_tasks}): {company_name}")
                    
                    status = "🔴 未更新 / 无法访问"
                    if url.lower().endswith('.pdf'):
                        status = "⚠️ 直接PDF链接 (需手动核实)"
                    elif "http" in url:
                        try:
                            response = requests.get(url, headers=headers, timeout=8)
                            response.encoding = response.apparent_encoding
                            soup = BeautifulSoup(response.text, 'html.parser')
                            page_text = soup.get_text()
                            if str(target_year) in page_text:
                                status = "🟢 极可能已更新!"
                        except Exception as e:
                            status = "🟡 网站拦截/超时，需手动查看"
                    else:
                        status = "无效链接"
                    
                    results.append({
                        "公司名称": company_name,
                        "监控状态": status,
                        "检测结果描述": f"网页中已检索到 {target_year} 字样" if "🟢" in status else "未发现关键字",
                        "直达链接": url
                    })
                    time.sleep(0.3)
    
                my_bar.empty()
                st.success(f"🎉 {target_year}年度检测扫描完成！")
                df_result = pd.DataFrame(results)
                st.data_editor(
                    df_result,
                    column_config={
                        "直达链接": st.column_config.LinkColumn("点击前往网页"),
                        "监控状态": st.column_config.TextColumn("状态", width="medium")
                    },
                    hide_index=True,
                    use_container_width=True,
                    key="scan_result_display"
                )
                st.info("💡 提示：请点击标记为 🟢 的公司链接下载 PDF，下载后请前往 [Step 1] 进行页码定位。")
    
    # ----- Step 1：智能页码检索（财险版）-----
    with tab1:
        st.markdown("### 📑 智能页码检索")
        uploaded_file = st.file_uploader(
            "拖拽或选择一份已经下载好的年报 PDF 文件",
            type="pdf",
            help="推荐上传一份财险公司的完整年报"
        )
    
        if uploaded_file:
            if 'pdf_bytes' not in st.session_state or st.session_state.get('pdf_name') != uploaded_file.name:
                st.session_state['pdf_bytes'] = uploaded_file.read()
                st.session_state['pdf_name'] = uploaded_file.name
                if 'found_pages' in st.session_state:
                    del st.session_state['found_pages']
                if 'edited_pages' in st.session_state:
                    del st.session_state['edited_pages']
            
            pdf_bytes = st.session_state['pdf_bytes']
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            total_pages = len(doc)
            
            st.caption(f"当前加载文件：{uploaded_file.name}　|　文档共 {total_pages} 页")
            st.markdown("---")
            
            col_left, col_spacer, col_right = st.columns([1, 0.05, 1.2])
    
            with col_left:
                st.markdown("#### 检索目标设定")
                # 公司选择：决定"保险合同负债及资产"的差异化识别规则
                company_options = ["平安产险", "人保产险", "太平产险", "众安产险", "阳光产险", "太保财险", "其他/自动识别"]
                default_company_idx = 0
                if target_company:
                    for _ci, _co in enumerate(company_options):
                        if _co in target_company or target_company in _co:
                            default_company_idx = _ci
                            break
                selected_company = st.selectbox("🏢 当前年报对应公司：", options=company_options, index=default_company_idx, help="用于精准匹配《保险合同负债及资产》表，请选择与上传PDF一致的公司")
                target_tables = st.multiselect(
                    "请选择需要定位的报表：",
                    [
                        "保险合同负债及资产",
                        "公司利润表",
                        "业务及管理费",
                        "公司资产负债表",
                        "主要经营指标",
                        "保险产品经营信息",
                        "折现率披露",
                        "非金融风险调整披露"  # 新增
                    ],
                    default=["保险合同负债及资产", "公司利润表", "公司资产负债表"]
                )
                
                st.markdown("")
                btn_search = st.button("启动智能检索", use_container_width=True)
                
                if btn_search:
                    current_api_key = st.session_state.get('api_key', "").strip()
                    if not current_api_key:
                        st.error("⚠️ 未检测到 API Key！请刷新页面返回登录界面填写。")
                    elif not target_tables:
                        st.error("请至少选择一张报表。")
                    else:
                        with st.spinner("需要一些时间，请稍等~"):
                            try:
                                result = ai_find_pages(pdf_bytes, current_api_key, target_tables, base_url, model_name, target_company if single_scan else "")
                                st.write("当前公司名称:", target_company)
                                main_table_pairs = [
                                    ("资产负债表（合并）", "资产负债表（公司）"),
                                    ("利润表（合并）", "利润表（公司）"),
                                    ("现金流量表（合并）", "现金流量表（公司）"),
                                    ("股东/所有者权益变动表（合并）", "股东/所有者权益变动表（公司）")
                                ]
                                for merge_key, company_key in main_table_pairs:
                                    merge_pages = result.get(merge_key, [0])
                                    company_pages = result.get(company_key, [0])
                                    if merge_pages != [0] and company_pages == [0]:
                                        result[company_key] = merge_pages
                                    elif company_pages != [0] and merge_pages == [0]:
                                        result[merge_key] = company_pages
                                st.session_state['found_pages'] = result
                                st.success("检索完成！请在下方核对物理页码。")
                            except Exception as e:
                                st.error(f"引擎出现异常：{e}")
    
                if 'found_pages' in st.session_state:
                    st.markdown("---")
                    st.markdown("#### 结果核对")
                    st.caption("提示：若表格跨页，请以英文逗号分隔页码（如 64, 65）。可结合右侧预览进行校准。")
                    
                    edited_pages = {}
                    for table_name in target_tables:
                        page_data = st.session_state['found_pages'].get(table_name, [0])
                        if isinstance(page_data, int):
                            page_data = [page_data]
                        str_val = ", ".join(map(str, page_data))
                        user_input = st.text_input(f"{table_name}", value=str_val, key=f"page_{table_name}")
                        try:
                            edited_pages[table_name] = [int(x.strip()) for x in user_input.split(",") if x.strip().isdigit()]
                        except:
                            edited_pages[table_name] = page_data
                    
                    st.session_state['edited_pages'] = edited_pages
                    
                    st.markdown("")
                    if st.button("确认页码，进入下一步", use_container_width=True):
                        st.success("页码已确认！请前往 Step 2 进行表格转换。")
    
            with col_right:
                st.markdown("#### 页面预览")
                if 'found_pages' in st.session_state and 'edited_pages' in st.session_state:
                    table_to_view = st.selectbox("选择要预览的报表：", options=list(st.session_state['edited_pages'].keys()))
                    pages_to_preview = st.session_state['edited_pages'].get(table_to_view, [0])
                    if not pages_to_preview:
                        pages_to_preview = [0]
                    
                    if len(pages_to_preview) > 1:
                        current_page = st.radio("该报表包含多页，请切换预览：", options=pages_to_preview, horizontal=True)
                    else:
                        current_page = pages_to_preview[0]
                    
                    preview_idx = current_page - 1
                    if 0 <= preview_idx < total_pages:
                        page = doc.load_page(preview_idx)
                        try:
                            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                            if pix is None:
                                st.warning("无法生成该页预览，可能是空白页。")
                            else:
                                img_data = pix.tobytes("png")   # 明确指定 PNG 格式
                                if img_data:
                                    st.image(img_data, caption=f"当前预览：第 {current_page} 页 / 共 {total_pages} 页", use_container_width=True)
                                else:
                                    st.warning("生成的图像数据为空，请检查 PDF 内容。")
                        except Exception as e:
                            st.warning(f"预览失败：{e}")
                    elif current_page == 0:
                        st.info("尚未识别到页码，请在左侧输入框中手动填入。")
                    else:
                        st.warning("该页码超出了文档总页数，请检查左侧修改。")
                else:
                    st.info("上传文件并启动检索后，此处将显示对应页面。")
    # ----- Step 2：表格智能转换 -----
    with tab2:
        st.markdown("### ⚡ 表格智能转换")
        st.markdown("""
        <div class="info-card blue">
            <h4>功能说明</h4>
            <p>系统将调用大模型对多页 PDF 进行网格化对齐重构，并自动拼接同表跨页数据，生成标准化 Excel。<b>如果出现提取失败或错位</b>，请开启下方的【图片扫描模式】重试。</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("")
        
        st.markdown("#### PDF类型选择")
        use_vision = st.toggle("📸 开启图片扫描模式 (适用于扫描版、纯图片PDF)", value=False)
        if use_vision:
            st.caption("提示：图片扫描模式请确保您登录时选择的模型支持视觉功能（如 GPT-4o, 阿里云通义千问等）。DeepSeek 暂不支持直接读图。")
        st.markdown("---")
        
        if 'edited_pages' not in st.session_state or 'pdf_bytes' not in st.session_state:
            st.warning("⚠️ 请先在 Step 1 中上传文件并完成【页码确认】。")
            st.button("开始提取结构化数据", disabled=True, use_container_width=True)
        else:
            current_api_key = st.session_state.get('api_key', "").strip()
            current_base_url = st.session_state.get('base_url', "https://api.deepseek.com")
            current_model_name = st.session_state.get('model_name', "deepseek-chat")
            
            can_run = True
            if not current_api_key:
                st.error("⚠️ 未检测到 API Key！请刷新页面返回登录界面填写授权密钥。")
                can_run = False
    
            if not can_run:
                st.button("开始提取结构化数据", disabled=True, use_container_width=True)
            else:
                valid_tasks = {k: v for k, v in st.session_state['edited_pages'].items() if v != [0]}
                
                if not valid_tasks:
                    st.warning("⚠️ 没有找到任何有效的页码，无需提取。")
                else:
                    if st.button("开始提取", use_container_width=True):
                        with st.spinner("☕ 趁现在喝口水吧，正在后台拼命打工中..."):
                            extracted_sheets = {}
                            global_unit = "未能自动提取，需人工核对"
                            all_tasks = []
                            for table_name, pages in valid_tasks.items():
                                for page_num in pages:
                                    all_tasks.append({"table": table_name, "page": page_num})
                            
                            total_pages_to_process = len(all_tasks)
                            pages_done = 0
                            
                            progress_bar = st.progress(0)
                            status_box = st.status(f"正在转化 {total_pages_to_process} 个任务...", expanded=True)
                            temp_results = {table_name: {} for table_name in valid_tasks.keys()}
                            
                            with status_box:
                                workers = 2 if use_vision else 5
                                
                                with ThreadPoolExecutor(max_workers=workers) as executor:
                                    future_to_task = {}
                                    
                                    for task in all_tasks:
                                        if use_vision:
                                            future = executor.submit(
                                                extract_single_page_vision, 
                                                st.session_state['pdf_bytes'], 
                                                task["page"], 
                                                task["table"], 
                                                current_api_key,
                                                current_base_url,
                                                current_model_name
                                            )
                                        else:
                                            future = executor.submit(
                                                extract_single_page, 
                                                st.session_state['pdf_bytes'], 
                                                task["page"], 
                                                task["table"], 
                                                current_api_key,
                                                current_base_url,
                                                current_model_name
                                            )
                                        future_to_task[future] = task
                                    
                                    for future in as_completed(future_to_task):
                                        task = future_to_task[future]
                                        t_name = task["table"]
                                        p_num = task["page"]
                                        try:
                                            df, raw_text = future.result()
                                            if df is not None and not df.empty:
                                                temp_results[t_name][p_num] = df
                                                st.write(f"✅ [{t_name}] - 第 {p_num} 页提取完成！")
                                                if global_unit == "未能自动提取，需人工核对":
                                                    unit = get_report_unit(raw_text)
                                                    if unit:
                                                        global_unit = unit
                                                        st.write(f"&nbsp;&nbsp;&nbsp;&nbsp;🔎 识别到该公司报表单位：【{global_unit}】")
                                            else:
                                                st.write(f"⚠️ [{t_name}] - 第 {p_num} 页未提取到有效数据。")
                                        except Exception as e:
                                            st.error(f"❌ [{t_name}] - 第 {p_num} 页提取失败: {str(e)}")
                                        
                                        pages_done += 1
                                        progress_bar.progress(pages_done / total_pages_to_process)
                            
                            for table_name, pages in valid_tasks.items():
                                table_dfs = []
                                for p_num in pages: 
                                    if p_num in temp_results[table_name]:
                                        table_dfs.append(temp_results[table_name][p_num])
                                
                                if table_dfs:
                                    merged_df = pd.concat(table_dfs, ignore_index=True)
                                    safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '', table_name)[:30]
                                    extracted_sheets[safe_sheet_name] = merged_df
                                        
                            status_box.update(label="🎉 所有任务提取完成！", state="complete", expanded=False)
                            st.session_state['extracted_data'] = extracted_sheets
                            st.session_state['global_unit'] = global_unit
                        
        if 'extracted_data' in st.session_state:
            st.markdown("---")
            st.markdown("#### 📊 提取结果预览")
            extracted_data = st.session_state['extracted_data']
            company_name = st.session_state.get('pdf_name', '未命名').replace(".pdf", "")
            
            unit_df = pd.DataFrame([
                {"项目": "源文件名称", "信息": company_name},
                {"项目": "探测到的报表单位", "信息": st.session_state.get('global_unit', '')}
            ])
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                unit_df.to_excel(writer, sheet_name="基本信息_单位", index=False)
                for sheet_name, df in extracted_data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            excel_data = output.getvalue()
            
            st.download_button(
                label="⬇️ 一键下载结构化提取表 (Excel)",
                data=excel_data,
                file_name=f"{company_name}_数据提取.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
            st.markdown("")
            sheet_names = list(extracted_data.keys())
            if sheet_names:
                tabs = st.tabs(sheet_names)
                for i, tab in enumerate(tabs):
                    with tab:
                        st.dataframe(extracted_data[sheet_names[i]], use_container_width=True)
   
    # ----- Step 3：目标表标准填报 -----
    with tab3:
        st.markdown("### 📝 目标表标准填报")
        st.markdown('<div class="info-card pink"><h4>功能说明</h4><p>AI 将自动寻找科目数据填充数值，生成 Excel 计算公式。支持内置模板或上传自定义模板。</p></div>', unsafe_allow_html=True)
        
        # 从财险公司列表获取类型映射
        COMPANY_TYPE_MAP = {item["公司"]: item["类别"] for item in DEFAULT_COMPANIES}
        
        col_t1, col_t2 = st.columns([1, 1])
        with col_t1: use_default = st.toggle("使用系统默认模板", value=True, help="开启后直接使用内置的财险样例表")
        
        template_file = (
"https://github.com/Polly031021/Annual-Report-System/raw/main/%E6%8C%87%E6%A0%87%E6%A0%B7%E4%BE%8B%E8%A1%A8.xlsx" 
            if use_default 
            else st.file_uploader("上传自定义目标表模板 (.xlsx)", type="xlsx", key="unique_template_uploader")
        )
    
        if template_file:
            COL_COMPANY, COL_CATEGORY, COL_FIELD_NAME, COL_FIELD_TYPE, COL_NOTE, COL_RULE, COL_CO_TYPE = "公司", "类别", "字段名", "字段类型", "注释", "计算规则", "公司类型"
    
            if 'extracted_data' not in st.session_state:
                st.warning("⚠️ 尚未找到提取的数据，请先完成 Step 1 和 Step 2。")
            elif st.button("启动智能填报与公式生成", use_container_width=True):
                with st.status("此过程略慢，请稍等~", expanded=True) as status_box:
                    st.write("📄 正在获取并解析模板表结构...")
                    
                    try:
                        if use_default:
                            import requests, io
                            res = requests.get(template_file, timeout=15)
                            res.raise_for_status()
                            template_df = pd.read_excel(io.BytesIO(res.content))
                        else:
                            template_df = pd.read_excel(template_file)
                    except Exception as e:
                        st.error(f"❌ 读取模板失败，请检查网络连接或文件格式：{e}")
                        st.stop()
                    
                    year_cols = sorted([c for c in template_df.columns if re.search(r'20\d{2}', str(c))])
                    if len(year_cols) < 2:
                        st.error("❌ 错误：模板中未能自动识别出两个或以上的年份数据列！请检查表头规范。")
                        st.stop()
                    
                    col_prev, col_curr = year_cols[-2], year_cols[-1]
                    st.session_state['col_prev'] = col_prev
                    st.session_state['col_curr'] = col_curr
                    
                    raw_name = st.session_state.get('pdf_name', '未命名').replace(".pdf", "")
                    company_short = re.sub(r'202\d年.*', '', raw_name).strip()
                    matched_type = next((c_type for c_name, c_type in COMPANY_TYPE_MAP.items() if c_name[:2] in company_short or company_short[:2] in c_name), "其他")
                    working_df = template_df.copy()   # ⬅️ 添加这一行
                    if COL_COMPANY in working_df.columns: working_df[COL_COMPANY] = company_short
                    if COL_CO_TYPE in working_df.columns: working_df[COL_CO_TYPE] = matched_type
                    
                    st.write("正在准备目标指标清单...")
                    input_items = working_df[working_df[COL_FIELD_TYPE].astype(str).str.strip() == "输入"]
                    ai_target_list = []
                    for _, r in input_items.drop_duplicates(subset=[COL_FIELD_NAME]).iterrows():
                        item = {
                            "类别": str(r.get(COL_CATEGORY, "")),
                            "标准字段名": str(r.get(COL_FIELD_NAME, "")),
                            "别名参考": str(r.get(COL_NOTE, "")) if pd.notna(r.get(COL_NOTE)) else "",
                            "提取说明": str(r.get(COL_NOTE, "")) if pd.notna(r.get(COL_NOTE)) else ""
                        }
                        ai_target_list.append(item)                    
                        
                    st.write("正在分析上下文...")
                    extracted_data = st.session_state['extracted_data']
                    context_text = "".join([f"\n[表名: {name}]\n{df.to_csv(index=False, sep='|')}\n" for name, df in extracted_data.items()])

                    # =========================================================
                    # 🌟 财险版系统提示词（已改造，强调提取说明）
                    # =========================================================
                    SYSTEM_PROMPT = """你是一个资深的财险精算审计专家。任务：将 PDF 提取的财务明细精准填入目标底稿，并严格区分上一年度和最新年份。

【通用执行准则：绝对原样提取】
1. ⚠️ 绝对指令：除了特殊的加总需求外，所有指标必须【原封不动地照抄】原文里的文本！
- 必须保留括号（表示负数）、逗号、正负号。例如原文是 "(295,992.00)"，必须输出为 "(295,992.00)"。
- 严禁擅自删除符号、严禁自行进行四舍五入。
2. 别名匹配：若标准名找不到，请查看"别名参考"列，或利用你的精算知识寻找同义词。
3. 空值处理：若原文为"—"或"无"请输出 "0"。若完全找不到该指标，请输出 null。
4. 自定义提取说明（最高优先级）：对于每个指标，如果“提取说明”不为空，请严格按照该说明提取数据。说明中可能包含“两个值相加”、“四个值相加”等操作，你需要根据说明从原文中定位相关数值并进行算术运算后填入。在匹配表格时，如果表格标题不明确，请根据表格内容中的关键描述词（如“分出的再保险合同”、“保险合同负债”等）来识别正确的表格。对于行匹配，请使用包含（contains）方式，即只要行文本中包含您指定的关键词即可，不必完全一致。如果说明为空，则使用“别名参考”或默认知识进行匹配。在定位数值时，务必以该行与指定列标题交叉处的单元格为准，不要取该列下方或其他行的汇总数。在提取数值时，务必严格区分“已发生赔款负债”列与“合计”列。只取指定列在该行的单元格值，不要取任何合计列的值，除非注释中明确要求取合计。
5. 在匹配表格时，如果表格标题不明确，请根据表格内容中的关键描述词（如“分出的再保险合同”、“保险合同负债”等）来识别正确的表格。对于行匹配，请使用包含（contains）方式，即只要行文本中包含您指定的关键词即可，不必完全一致。

【特殊指令：财险利润表科目映射】
当提取财险利润表科目时，请使用以下映射规则：
1. 【保险业务收入】：提取"保险业务收入"或"已赚保费"或"签单保费"。
2. 【分出保费】：提取"分出保费"或"分保费用"。
3. 【提取未到期责任准备金】：提取"未到期责任准备金"或"UPR"。
4. 【已发生赔款】：提取"已发生赔款"或"赔付支出"或"赔款支出"。
5. 【手续费及佣金支出】：提取"手续费及佣金支出"。
6. 【业务及管理费】：提取"业务及管理费"或"管理费用"。
7. 【承保利润】：若无法直接提取，可用"保险业务收入 - 分出保费 - 提取未到期责任准备金 - 已发生赔款 - 手续费及佣金支出 - 业务及管理费"计算。

【特殊指令：财险资产负债表科目映射】
当提取财险资产负债表科目时，请使用以下映射规则：
1. 【应收保费】：提取"应收保费"。
2. 【应收分保账款】：提取"应收分保账款"或"应收分保款项"。
3. 【应付保费】：提取"应付保费"或"预收保费"。
4. 【未到期责任准备金】：提取"未到期责任准备金"。
5. 【未决赔款准备金】：提取"未决赔款准备金"或"IBNR"。
6. 【总资产】：提取"总资产"或"资产总计"。
7. 【总负债】：提取"总负债"或"负债合计"。
8. 【股东权益】：提取"股东权益"或"所有者权益"或"净资产"。

【特殊指令：财险综合成本率拆解】
当提取综合成本率相关科目时：
1. 【综合成本率】：若有直接披露，优先提取；否则用"赔付率 + 费用率"计算。
2. 【综合赔付率】：提取"赔付率"或"综合赔付率"。
3. 【综合费用率】：提取"费用率"或"综合费用率"。

【时间维度与排版防反转绝对指令】
中国企业财报的标准排版规则是：【左边的数据列为当期（最新年份），右边的数据列为上期（历史年份）】。
- 当填报要求中包含"较晚/最新年份"（如当期/本年/年末）的键时：必须从原表表头的"本期"、"本年"、"年末"或【最左侧的数据列】取数。
- 当填报要求中包含"较早/历史年份"（如上期/上年/年初）的键时：必须从原表表头的"上期"、"上年"、"年初"或【紧挨着它的右侧数据列】取数。
⚠️ 警告：大语言模型极易犯"线性思维"错误（误以为小年份排在左边，大年份排在右边）。要求你必须打破定势，仔细核对表头！永远记住：左边的列才是最新当期！

【提取说明优先级】
- 如果“提取说明”字段有内容，必须优先按照该说明执行，忽略其他默认映射。
- 如果“提取说明”为空，则按“别名参考”或标准映射提取。

仅输出合法的 JSON 格式，严禁带有任何 Markdown 标记或文字说明。
格式要求：
- 顶层键为“公司名称”（与模板表中的“公司”列一致）。
- 每个公司下为字段名到数值的映射。
- 每个字段包含 "prev"（上一年）和 "curr"（最新年）两个键。
格式示例：{"平安产险": {"保险服务收入-车险": {"prev": "0", "curr": "228494928861"}, "营业利润-车险": {"prev": "0", "curr": "123456"}}, "人保产险": {...}}"""

                    st.write("正在呼叫大模型进行语义映射与精准取数...")
                    
                    current_api_key = st.session_state.get('api_key', "").strip()
                    current_base_url = st.session_state.get('base_url', "https://api.deepseek.com")
                    current_model_name = st.session_state.get('model_name', "deepseek-chat")
                    
                    client = OpenAI(api_key=current_api_key, base_url=current_base_url)
                    def process_final_val(val):
                        """将 AI 返回的数值字符串转为浮点数，支持括号负数、逗号分隔、空值处理"""
                        if val is None:
                            return 0.0
                        if isinstance(val, (int, float)):
                            return float(val)
                        s = str(val).strip()
                        if s in ['', 'null', 'None', 'nan', 'NaN']:
                            return 0.0
                        # 处理 (123,456.78) -> -123456.78
                        if s.startswith('(') and s.endswith(')'):
                            s = '-' + s[1:-1]
                        s = s.replace(',', '').replace(' ', '')
                        try:
                            return float(s)
                        except ValueError:
                            return 0.0
                    try:
                        # 收集所有需要提取的公司
                        companies_to_extract = list(set([
                            str(r.get(COL_COMPANY, '')).strip()
                            for _, r in working_df.iterrows()
                            if str(r.get(COL_COMPANY, '')).strip() != ''
                        ]))
                    
                        # 构建用户消息，明确列出每个指标的提取说明
                        user_content = f"""目标公司列表：{companies_to_extract}
                    
                    目标指标列表（含提取说明）：
                    {json.dumps(ai_target_list, ensure_ascii=False, indent=2)}
                    
                    请按照以下要求提取数据：
                    1. 对于每个指标，如果“提取说明”不为空，必须严格按照说明中的描述进行提取。
                    2. 如果“提取说明”为空，则根据“别名参考”或你的精算知识寻找匹配项。
                    3. 注意时间维度：最新年份对应左侧数据列，历史年份对应右侧数据列。
                    4. 输出格式必须为合法 JSON，顶层键为公司名称，每个公司下是字段名到 {{"prev": ..., "curr": ...}} 的映射。
                    
                    数据内容：
                    {context_text}"""
                    
                        response = client.chat.completions.create(
                            model=current_model_name,
                            messages=[
                                {"role": "system", "content": SYSTEM_PROMPT},
                                {"role": "user", "content": user_content}
                            ],
                            temperature=0.0
                        )
                        ai_res = response.choices[0].message.content.strip()
                        ai_res = re.sub(r'^```(json)?\n?', '', ai_res, flags=re.MULTILINE).replace("```", "").strip()
                        ai_data = json.loads(ai_res)
                        st.write("✅ 数据映射成功！正在回填表格...")
                        
                        # ----- 1. 回填输入数据（按公司和字段名双重匹配） -----
                        for idx, row in working_df.iterrows():
                            ftype = str(row.get(COL_FIELD_TYPE, "")).strip()
                            if ftype != "输入":
                                continue
                            fname = str(row.get(COL_FIELD_NAME, "")).strip()
                            if not fname:
                                continue
                            company_name = str(row.get(COL_COMPANY, "")).strip()
                            
                            if company_name:
                                # 有公司名，精确匹配
                                if company_name in ai_data and fname in ai_data[company_name]:
                                    item_data = ai_data[company_name][fname]
                                    if isinstance(item_data, dict):
                                        prev_val = process_final_val(item_data.get("prev"))
                                        curr_val = process_final_val(item_data.get("curr"))
                                        # 对“获取费用”和“维持费用”取绝对值
                                        if fname in ["获取费用", "维持费用"]:
                                            if isinstance(prev_val, (int, float)):
                                                prev_val = abs(prev_val)
                                            if isinstance(curr_val, (int, float)):
                                                curr_val = abs(curr_val)
                                        working_df.at[idx, col_prev] = prev_val
                                        working_df.at[idx, col_curr] = curr_val
                            else:
                                # 没有公司名，从所有公司中查找该字段
                                found_companies = [co for co, data in ai_data.items() if fname in data]
                                if len(found_companies) == 1:
                                    co = found_companies[0]
                                    item_data = ai_data[co][fname]
                                    if isinstance(item_data, dict):
                                        prev_val = process_final_val(item_data.get("prev"))
                                        curr_val = process_final_val(item_data.get("curr"))
                                        if fname in ["获取费用", "维持费用"]:
                                            if isinstance(prev_val, (int, float)):
                                                prev_val = abs(prev_val)
                                            if isinstance(curr_val, (int, float)):
                                                curr_val = abs(curr_val)
                                        working_df.at[idx, col_prev] = prev_val
                                        working_df.at[idx, col_curr] = curr_val
                                elif len(found_companies) > 1:
                                    # 多个公司都有该字段，默认取第一个公司的值（可根据需要调整）
                                    co = found_companies[0]
                                    item_data = ai_data[co][fname]
                                    if isinstance(item_data, dict):
                                        prev_val = process_final_val(item_data.get("prev"))
                                        curr_val = process_final_val(item_data.get("curr"))
                                        if fname in ["获取费用", "维持费用"]:
                                            if isinstance(prev_val, (int, float)):
                                                prev_val = abs(prev_val)
                                            if isinstance(curr_val, (int, float)):
                                                curr_val = abs(curr_val)
                                        working_df.at[idx, col_prev] = prev_val
                                        working_df.at[idx, col_curr] = curr_val
                                # 如果找不到，则留空
                        # ----- 2. 计算“计算”类型字段的值（用于预览）-----
                        # 构建字段名->数值的映射（用于计算）
                        field_prev_vals = {}
                        field_curr_vals = {}
                        for _, r in working_df.iterrows():
                            fname = str(r.get(COL_FIELD_NAME, "")).strip()
                            if fname:
                                field_prev_vals[fname] = r[col_prev] if pd.notna(r[col_prev]) else 0.0
                                field_curr_vals[fname] = r[col_curr] if pd.notna(r[col_curr]) else 0.0
                        
                        # 遍历每一行，对“计算”类型执行计算
                        for idx, row in working_df.iterrows():
                            ftype = str(row.get(COL_FIELD_TYPE, "")).strip()
                            if ftype != "计算":
                                continue
                            rule = str(row.get(COL_RULE, "")).strip()
                            if not rule or rule in ["nan", "None", ""]:
                                continue
                            # 替换 [字段名] 为数值
                            def replace_field(match):
                                field = match.group(1)
                                # 注意：计算规则中可能包含当前年份或其他字段，我们分别计算prev和curr
                                return f"field_prev_vals['{field}']"  # 但需要在eval时使用局部变量
                            # 更安全的方法是构造表达式字符串，然后用eval计算
                            # 我们分别计算prev和curr
                            try:
                                # 提取所有 [字段名]
                                fields = re.findall(r'\[(.*?)\]', rule)
                                # 构造prev表达式
                                prev_expr = rule
                                curr_expr = rule
                                for f in fields:
                                    prev_expr = prev_expr.replace(f"[{f}]", str(field_prev_vals.get(f, 0.0)))
                                    curr_expr = curr_expr.replace(f"[{f}]", str(field_curr_vals.get(f, 0.0)))
                                # 安全计算（只允许基本运算）
                                # 使用eval，但限制命名空间
                                safe_dict = {"abs": abs, "min": min, "max": max, "round": round}
                                # 对prev和curr分别计算
                                try:
                                    prev_result = eval(prev_expr, {"__builtins__": None}, safe_dict)
                                    working_df.at[idx, col_prev] = prev_result if isinstance(prev_result, (int, float)) else 0.0
                                except:
                                    working_df.at[idx, col_prev] = 0.0
                                try:
                                    curr_result = eval(curr_expr, {"__builtins__": None}, safe_dict)
                                    working_df.at[idx, col_curr] = curr_result if isinstance(curr_result, (int, float)) else 0.0
                                except:
                                    working_df.at[idx, col_curr] = 0.0
                            except Exception as e:
                                # 如果计算失败，留空
                                pass
                    
                        st.write("⚙️ 正在生成可追溯的单元格公式...")
                        
                        # 生成Excel（含公式）
                        temp_buffer = io.BytesIO()
                        working_df.to_excel(temp_buffer,index=False)
                        temp_buffer.seek(0)
                        
                        wb = openpyxl.load_workbook(temp_buffer)
                        ws = wb.active
                        
                        cols = {cell.value:cell.column for cell in ws[1]}
                        
                        prev_idx = cols.get(col_prev)
                        curr_idx = cols.get(col_curr)
                        
                        # 确保原有的公式单元格不被破坏
                        for r in range(2,ws.max_row+1):
                            for c_idx in [prev_idx,curr_idx]:
                                if not c_idx:
                                    continue
                                cell = ws.cell(r,c_idx)
                                if isinstance(cell.value,str) and cell.value.startswith("="):
                                    cell.value = cell.value
                        
                        field_to_row = {
                            str(ws.cell(r,cols[COL_FIELD_NAME]).value).strip():r
                            for r in range(2,ws.max_row+1)
                            if ws.cell(r,cols[COL_FIELD_NAME]).value
                        }
                        
                        sorted_fields = sorted(field_to_row.keys(),key=len,reverse=True)
                        
                        prev_letter = get_column_letter(prev_idx) if prev_idx else ""
                        curr_letter = get_column_letter(curr_idx) if curr_idx else ""
                        
                        for r in range(2,ws.max_row+1):
                            ftype = str(ws.cell(r,cols[COL_FIELD_TYPE]).value).strip()
                            rule = str(ws.cell(r,cols[COL_RULE]).value).strip()
                            if ftype!="计算" or rule in ["nan","None",""]:
                                continue
                            prev_formula = rule
                            curr_formula = rule
                            for f in sorted_fields:
                                if f in rule:
                                    row_num = field_to_row[f]
                                    if prev_letter:
                                        prev_formula = prev_formula.replace(f,f"{prev_letter}{row_num}")
                                    if curr_letter:
                                        curr_formula = curr_formula.replace(f,f"{curr_letter}{row_num}")
                            try:
                                if prev_idx:
                                    ws.cell(r,prev_idx).value = f"={prev_formula}"
                                if curr_idx:
                                    ws.cell(r,curr_idx).value = f"={curr_formula}"
                            except:
                                pass
                        
                        # 格式化数字
                        for r in range(2,ws.max_row+1):
                            for c_idx in [prev_idx,curr_idx]:
                                if not c_idx:
                                    continue
                                cell = ws.cell(r,c_idx)
                                cell.number_format = '#,##0_ ;(#,##0);"-"'
                                if cell.value and not str(cell.value).startswith("="):
                                    try:
                                        clean_val = str(cell.value).replace(',','').replace('(','-').replace(')','')
                                        cell.value = float(clean_val)
                                    except:
                                        pass
                        
                        final_buffer = io.BytesIO()
                        wb.save(final_buffer)
                        final_buffer.seek(0)
                        
                        st.session_state['filled_excel'] = final_buffer.getvalue()
                        st.session_state['final_df'] = working_df
                        st.session_state['col_prev'] = col_prev
                        st.session_state['col_curr'] = col_curr
                    
                    except Exception as e:
                        st.error(f"❌ 流程中断: {str(e)}")
                        status_box.update(label="处理失败", state="error", expanded=True)

        if 'filled_excel' in st.session_state:
            st.markdown("---")
            st.markdown("#### 📋 智能填报结果预览")
            company_name = st.session_state.get('pdf_name', '未命名').replace(".pdf", "")
            
            st.download_button(
                label="⬇️ 下载已填报的底稿 (含公式)",
                data=st.session_state['filled_excel'],
                file_name=f"{company_name}_自动填报表.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.markdown("💡 *预览表仅展示数值，下载后的 Excel 已包含自动关联公式。*")
            st.dataframe(st.session_state['final_df'], use_container_width=True)
    # ----- Step 4 分析与校验（财险版）-----
    with tab4:
        # 🎨 注入自定义 CSS（保持不变）
        st.markdown("""
            <style>
            div[data-testid="stPopover"] > button {
                padding: 2px 8px !important;
                min-height: 26px !important;
                height: 26px !important;
                font-size: 12px !important;
                background-color: #f8f9fa !important;
                border: 1px solid #ddd !important;
            }
            div[data-testid="stColumn"] button[kind="secondary"] {
                padding: 2px 10px !important;
                height: 26px !important;
                font-size: 12px !important;
                white-space: nowrap !important;
                min-width: 70px !important;
                display: flex !important;
                align-items: center !important;
                justify-content: center !important;
            }
            .stSubheader { margin-bottom: -10px !important; }
            </style>
        """, unsafe_allow_html=True)
    
        # 1. 顶部工具栏布局
        col_title, col_btn_sample, col_btn_refresh, col_manual_area = st.columns([3.5, 1.8, 1.1, 1.8])
        
        with col_title:
            st.subheader("🏁 精算数据勾稽关系检查")
            
        with col_btn_sample:
            with st.popover("校验表获取", use_container_width=True):
                st.caption("将人工校验两列复制到需要校对的表格对应位置")
                import requests
                # 🔧 修改点：替换为财险勾稽规则样例表 URL
                url = "https://github.com/Polly031021/Annual-Report-System/raw/main/%E5%8B%BE%E7%A8%BD%E8%A7%84%E5%88%99%E6%A0%B7%E4%BE%8B%E8%A1%A8.xlsx"
                try:
                    with st.spinner("获取中..."):
                        response = requests.get(url, timeout=10)
                        response.raise_for_status()
                    st.download_button(
                        label="⬇️ 下载样例文件", 
                        data=response.content, 
                        file_name="财险人工校验样例.xlsx", 
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error("网络请求失败，请检查网络或 GitHub 链接是否有效！")
        
        with col_btn_refresh:
            if st.button("🔄 刷新", use_container_width=True):
                st.rerun() 
    
        # --- 📖 核对重点与语音（财险版）---
        with col_manual_area:
            c1, c2 = st.columns([1, 0.4])
            
            # ========== 🔧 修改点：财险版核对指南 ==========
            manual_content = """
            财险数据人工核对指南
            
            1. 模板准备： 请将样例表中的"人工校验"两列手动复制到当前需要核对的表格中。
            2. 年度对齐： 重点检查 最新年 和 上一年 的数据是否有填反现象。
            3. 综合成本率： 检查 COR = 综合赔付率 + 综合费用率 是否成立。
            4. 承保利润： 检查 承保利润 = 已赚保费 - 赔付支出 - 手续费及佣金 - 业务及管理费。
            5. 赔付率拆解： 检查 当期赔款占比、回溯偏差率等因子是否合理。
            6. 费用分类： 检查 手续费及佣金、业务及管理费 的分类是否准确。
            7. 偿付能力： 如数据中有偿付能力充足率，需与最低资本、实际资本勾稽。
            8. 准备金评估： 未到期责任准备金、未决赔款准备金（含 IBNR）需与相关科目勾稽。
            """
    
            with c1:
                with st.popover("🕶️核对指南", use_container_width=True):
                    html_code = f"""
    <div style="font-family: 'Microsoft YaHei', sans-serif; padding: 5px;">
        <h4 style="color: #1f4e79; margin: 0 0 10px 0; font-size: 1rem; border-bottom: 2px solid #e1e4e8; padding-bottom: 5px;">
            财险数据人工核对指南
        </h4>
        <div style="line-height: 1.6; font-size: 0.85rem; color: #333333;">
            <p style="margin-bottom: 8px;">
                <strong style="color: #d32f2f;">1. 模板准备：</strong> 
                将样例表的"人工校验"两列复制到核对表中。
            </p>
            <p style="margin-bottom: 8px;">
                <strong style="color: #d32f2f;">2. 年度对齐：</strong> 
                检查 <span style="background-color: #fff3cd; padding: 1px 3px;">去年</span> 和 <span style="background-color: #fff3cd; padding: 1px 3px;">最新年</span> 数据是否有填反。
            </p>
            <p style="margin-bottom: 8px;">
                <strong style="color: #d32f2f;">3. 综合成本率：</strong> 
                检查 COR = 综合赔付率 + 综合费用率 是否成立。
            </p>
            <p style="margin-bottom: 8px;">
                <strong style="color: #d32f2f;">4. 承保利润：</strong> 
                承保利润 = 已赚保费 - 赔付支出 - 手续费 - 管理费。
            </p>
            <p style="margin-bottom: 8px;">
                <strong style="color: #d32f2f;">5. 赔付率拆解：</strong> 
                检查当期赔款占比、回溯偏差率等因子是否合理。
            </p>
            <p style="margin-bottom: 8px;">
                <strong style="color: #d32f2f;">6. 费用分类：</strong> 
                检查手续费及佣金、业务及管理费的分类是否准确。
            </p>
            <p style="margin-bottom: 8px;">
                <strong style="color: #d32f2f;">7. 偿付能力：</strong> 
                偿付能力充足率需与最低资本、实际资本勾稽。
            </p>
            <p style="margin-bottom: 0;">
                <strong style="color: #d32f2f;">8. 准备金评估：</strong> 
                未到期责任准备金、未决赔款准备金（含 IBNR）需与相关科目勾稽。
            </p>
        </div>
    </div>
    """
                    st.markdown(html_code, unsafe_allow_html=True)
                    
            with c2:
                clean_voice_text = manual_content.replace('\n', '。').replace('**', '')
                tts_html = f"""
                <div style="display: flex; align-items: center; height: 26px;">
                    <button id="tts-btn" onclick="toggleSpeak()" style="
                        border: 1px solid #ddd;
                        background-color: #f8f9fa;
                        border-radius: 4px;
                        cursor: pointer;
                        width: 26px;
                        height: 26px;
                        display: flex;
                        align-items: center;
                        justify-content: center;
                        font-size: 12px;
                    ">📢</button>
                </div>
                <script>
                var msg = new SpeechSynthesisUtterance();
                msg.text = "{clean_voice_text}";
                msg.lang = 'zh-CN';
                msg.rate = 1.2;
                function toggleSpeak() {{
                    const btn = document.getElementById('tts-btn');
                    if (window.speechSynthesis.speaking) {{
                        window.speechSynthesis.cancel();
                        btn.style.backgroundColor = "#f8f9fa";
                        btn.innerText = "📢";
                    }} else {{
                        window.speechSynthesis.speak(msg);
                        btn.style.backgroundColor = "#ffebeb"; 
                        btn.innerText = "⏹";
                        msg.onend = function() {{
                            btn.style.backgroundColor = "#f8f9fa";
                            btn.innerText = "📢";
                        }};
                    }}
                }}
                </script>
                """
                st.components.v1.html(tts_html, height=30)
    
        # ----- 校验执行逻辑（保持不变，框架完全通用）-----
        if 'final_df' in st.session_state:
            df = st.session_state['final_df'].copy()
            
            with st.expander("🛠️ 勾稽规则配置 (点击展开/修改公式)"):
                st.info("curr代表当前年度数据；prev代表上一年度数据。")
                
                # 🔧 修改点：使用空列表或财险规则，不再使用寿险 DEFAULT_RULES
                # ===== 财险默认校验规则（预置 5 条核心规则） =====
                DEFAULT_RULES_PC = [
                    {
                        "规则名称": "1. 综合成本率 = 赔付率 + 费用率",
                        "公式": "abs(curr['综合成本率'] - (curr['综合赔付率'] + curr['综合费用率'])) < 0.001",
                        "类型": "single",
                        "描述": "COR 应等于赔付率与费用率之和（误差<0.1%）"
                    },
                    {
                        "规则名称": "2. 承保利润勾稽",
                        "公式": "abs(curr['承保利润'] - (curr['保险业务收入'] - curr['已发生赔款'] - curr['手续费及佣金支出'] - curr['业务及管理费'])) < 5",
                        "类型": "single",
                        "描述": "承保利润 = 收入 - 赔款 - 手续费 - 管理费"
                    },
                    {
                        "规则名称": "3. 总资产 = 总负债 + 股东权益",
                        "公式": "abs(curr['总资产'] - (curr['总负债'] + curr['股东权益'])) < 5",
                        "类型": "single",
                        "描述": "资产负债表恒等式"
                    },
                    {
                        "规则名称": "4. 净资产变动（跨年）",
                        "公式": "abs(curr['期末股东权益'] - prev['期末股东权益'] - curr['净利润'] - curr['其他综合收益']) < 10",
                        "类型": "cross",
                        "描述": "期末净资产 ≈ 上期净资产 + 净利润 + OCI"
                    },
                    {
                        "规则名称": "5. 新旧准则比值合理",
                        "公式": "abs(curr['新旧准则比值'] - (curr['保险服务收入'] / curr['保险业务收入'])) < 0.001",
                        "类型": "single",
                        "描述": "新旧准则比值 = 保险服务收入 / 保险业务收入"
                    }
                ]
                
                # 使用预置规则作为默认值
                rules_df = st.data_editor(
                    pd.DataFrame(DEFAULT_RULES_PC),
                    num_rows="dynamic",
                    key="rules_editor_v3"
                )
                # 方案2：预置财险规则（取消下面注释）
                # DEFAULT_RULES_PC = [
                #     {"规则名称": "1. 综合成本率 = 赔付率 + 费用率", 
                #      "公式": "abs(curr['综合成本率'] - (curr['赔付率'] + curr['费用率'])) < 0.001", 
                #      "类型": "single"},
                #     ...
                # ]
                # rules_df = st.data_editor(pd.DataFrame(DEFAULT_RULES_PC), num_rows="dynamic", key="rules_editor_v3")
    
                target_prev = str(st.session_state.get('col_prev'))
                target_curr = str(st.session_state.get('col_curr'))
                
                col_prev_name = next(
                    (col for col in df.columns if str(target_prev) in str(col)),
                    None
                )
                
                col_curr_name = next(
                    (col for col in df.columns if str(target_curr) in str(col)),
                    None
                )
                
                if not col_prev_name or not col_curr_name:
                    st.error(
                        f"❌ 表格中找不到包含 '{target_prev}' 和 '{target_curr}' 的列。"
                    )
                    st.stop()
    
            def parse_finance_num(val):
                if pd.isnull(val): return 0.0
                v_str = str(val).strip()
                if v_str in ['-', '']: return 0.0
                if v_str.startswith('(') and v_str.endswith(')'): v_str = '-' + v_str[1:-1]
                v_str = v_str.replace(',', '')
                try: return float(v_str)
                except ValueError: return 0.0
    
            # SmartDict（保持不变）
            class SmartDict(dict):
                def _clean(self, k):
                    if not isinstance(k, str): return k
                    return k.replace('（','(').replace('）',')').replace('：',':').replace(' ','').replace('\n','').replace('\xa0','')
                
                def __init__(self, mapping):
                    super().__init__(mapping)
                    self._map = {self._clean(k): k for k in mapping.keys()}
                    
                def __getitem__(self, k):
                    clean_k = self._clean(k)
                    if clean_k in self._map:
                        return super().__getitem__(self._map[clean_k])
                    raise KeyError(k)
    
            base_prev = {
                str(k).strip(): parse_finance_num(v)
                for k, v in zip(df['字段名'], df[col_prev_name])
            }
            
            base_curr = {
                str(k).strip(): parse_finance_num(v)
                for k, v in zip(df['字段名'], df[col_curr_name])
            }
            
            prev_year_map = SmartDict(base_prev)
            curr_year_map = SmartDict(base_curr)
            
            check_results = []
            
            for _, rule in rules_df.iterrows():
                rule_name = str(rule.get('规则名称', ''))
                formula = str(rule.get('公式', ''))
                rule_type = str(rule.get('类型', 'single')).strip().lower()
            
                targets = (
                    [(f"{target_prev}-{target_curr}", None)]
                    if rule_type == "cross"
                    else [
                        (target_prev, prev_year_map),
                        (target_curr, curr_year_map)
                    ]
                )
            
                for year_label, curr_map in targets:
                    ctx = {
                        'curr': curr_map,
                        'prev': prev_year_map,
                        'prev_year': prev_year_map,
                        'curr_year': curr_year_map,
                        'abs': abs,
                        'min': min,
                        'max': max,
                        'round': round
                    }
            
                    try:
                        is_pass = bool(eval(formula, {"__builtins__": None}, ctx))
                        status = "✅ PASS" if is_pass else "❌ FALSE"
                        detail = "勾稽无误" if is_pass else "差额超过允许阈值"
            
                    except KeyError as e:
                        status = "⚠️ ERROR"
                        detail = f"缺失字段: {str(e)}"
            
                    except Exception as e:
                        status = "⚠️ ERROR"
                        detail = f"公式错误: {str(e)}"
            
                    check_results.append({
                        "年度": year_label,
                        "规则名称": rule_name,
                        "检查结果": status,
                        "诊断详情": detail
                    })
    
            res_display_df = pd.DataFrame(check_results)
            def color_status(val):
                if '✅' in str(val):
                    return 'background-color: #C6EFCE; color: #006100;'
                elif '❌' in str(val):
                    return 'background-color: #FFC7CE; color: #9C0006;'
                elif '⚠️' in str(val):
                    return 'background-color: #FFEB9C; color: #9C5700;'
                return ''
            # 在 st.dataframe 之前插入
            if res_display_df.empty:
                res_display_df = pd.DataFrame(columns=["年度", "规则名称", "检查结果", "诊断详情"])
            else:
                # 可选：确保列都存在
                for col in ["年度", "规则名称", "检查结果", "诊断详情"]:
                    if col not in res_display_df.columns:
                        res_display_df[col] = ""
            
            # 然后正常显示
            st.dataframe(
                res_display_df.style.map(color_status, subset=['检查结果']),
                use_container_width=True, hide_index=True
            )
            
            def color_status(val):
                if '✅' in str(val): return 'background-color: #C6EFCE; color: #006100;'
                elif '❌' in str(val): return 'background-color: #FFC7CE; color: #9C0006;'
                elif '⚠️' in str(val): return 'background-color: #FFEB9C; color: #9C5700;'
                return ''
    
            st.dataframe(
                res_display_df.style.map(color_status, subset=['检查结果']),
                use_container_width=True, hide_index=True
            )
    
            # 导出逻辑（保持不变）
            if st.button("📥 导出带勾稽结果的报告"):
                if 'filled_excel' not in st.session_state:
                    st.error("❌ 找不到填报后的数据，请先重新运行 Step 3")
                    st.stop()
                    
                input_buffer = io.BytesIO(st.session_state['filled_excel'])
                workbook = openpyxl.load_workbook(input_buffer)
                
                ws_data = workbook.active
                ws_data.title = "数据明细"
                
                if "勾稽报告" in workbook.sheetnames:
                    del workbook["勾稽报告"]
                ws_res = workbook.create_sheet("勾稽报告")
                
                headers = list(res_display_df.columns)
                for c_idx, header in enumerate(headers, 1):
                    ws_res.cell(row=1, column=c_idx).value = header
                
                from openpyxl.styles import PatternFill
                fill_green = PatternFill(start_color='C6EFCE', fill_type='solid')
                fill_red = PatternFill(start_color='FFC7CE', fill_type='solid')
                fill_yellow = PatternFill(start_color='FFEB9C', fill_type='solid')
                
                for r_idx, row_data in enumerate(res_display_df.values, 2):
                    for c_idx, value in enumerate(row_data, 1):
                        cell = ws_res.cell(row=r_idx, column=c_idx)
                        cell.value = value
                        if c_idx == 3:
                            val_str = str(value)
                            if "PASS" in val_str: cell.fill = fill_green
                            elif "FALSE" in val_str: cell.fill = fill_red
                            elif "ERROR" in val_str: cell.fill = fill_yellow
                
                ws_res.column_dimensions['B'].width = 30
                ws_res.column_dimensions['C'].width = 15
                ws_res.column_dimensions['D'].width = 40
    
                output = io.BytesIO()
                workbook.save(output)
                
                st.download_button(
                    label="点击下载精算复核底稿", 
                    data=output.getvalue(), 
                    file_name=f"勾稽复核底稿_{st.session_state.get('pdf_name','未命名').replace('.pdf','')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.info("💡 提取结果将在此处显示。请先在上方点击“开始提取”按钮。")    
    # ----- Step 5：多公司数据集成 -----
    with tab5:
        col_title, col_btn = st.columns([4, 1])
        with col_title:
            st.subheader("⛓️‍💥 多公司数据集成与汇率/单位转换")
        st.info("功能说明：支持上传单文件多Sheet或多文件。系统将自动提取所有公司，请在下方为不同公司配置对应的汇率和原始表格的单位。")
    
        uploaded_files = st.file_uploader("请上传已完成勾稽检查的底稿 (支持多文件或单文件多Sheet)", type="xlsx", accept_multiple_files=True)
    
        @st.cache_data(show_spinner=False)
        def run_data_integration(temp_data_list, rate_cfg, unit_cfg):
            # =========================================================
            # 🌟 修改点：财险版豁免字段列表（比率/文本类不换算）
            # =========================================================
            exact_exempt_fields = [
                # 比率类（已为百分比，无需换算）
                "综合成本率",
                "综合赔付率",
                "综合费用率",
                "赔付率",
                "费用率",
                "投资收益率",
                "综合偿付能力充足率",
                "核心偿付能力充足率",
                # 文本类（无需换算）
                "折现率假设",
                "风险边际",
                "非金融风险调整",
                "计量方法",
                "会计政策",
            ]
            
            combined_list = []
            
            def clean_to_float(val):
                try:
                    if isinstance(val, (int, float)): return float(val)
                    if isinstance(val, str):
                        val = val.replace(',','').replace('(','-').replace(')','').strip()
                        return 0.0 if val in ['-',''] else float(val)
                    return 0.0
                except: return 0.0
    
            for item in temp_data_list:
                df_single, comp_name = item["df"], item["comp"]
                rate, unit_mult = rate_cfg[comp_name], unit_cfg[comp_name]
    
                year_cols = {}
                for c in df_single.columns:
                    import re
                    m = re.search(r'(20\d{2})', str(c))
                    if m: year_cols[m.group(1)] = c
    
                if len(year_cols) < 1: continue
    
                base_cols = ["公司类型","公司","类别","字段名","字段类型"]
                existing_base = [c for c in base_cols if c in df_single.columns]
    
                for year_label, col_name in year_cols.items():
                    df_year = df_single[existing_base + [col_name]].copy()
                    df_year["公司"] = comp_name
                    df_year["报告年份"] = year_label
                    df_year["汇率"] = rate
                    df_year["(百万)原币"] = None
                    df_year["(百万)人民币"] = None
                    df_year["汇率"] = df_year["汇率"].astype(object)
    
                    for idx in df_year.index:
                        raw_val = df_year.loc[idx, col_name]
                        f_name = str(df_year.loc[idx, "字段名"]).strip()
                        # 获取字段类型（若存在）
                        field_type = str(df_year.loc[idx, "字段类型"]).strip() if "字段类型" in df_year.columns else ""
                        # 如果是“计算”字段，直接保留原值，不做任何换算
                        if field_type == "计算" and f_name != "再保净成本":
                            df_year.at[idx, "(百万)原币"] = raw_val
                            df_year.at[idx, "(百万)人民币"] = raw_val
                            df_year.at[idx, "汇率"] = "豁免换算"
                            continue
                                                
                        # 判断是否为豁免字段（比率/文本）
                        if f_name in exact_exempt_fields:
                            # 文本字段直接保留原值
                            if any(x in f_name for x in ["假设", "方法", "政策", "边际"]):
                                df_year.at[idx,"(百万)原币"] = str(raw_val) if pd.notna(raw_val) else ""
                                df_year.at[idx,"(百万)人民币"] = str(raw_val) if pd.notna(raw_val) else ""
                                df_year.at[idx,"汇率"] = "豁免换算"
                            else:
                                # 比率字段：去除%符号后转为小数
                                dec_val = float(raw_val.replace('%','').strip())/100.0 if isinstance(raw_val,str) and '%' in raw_val else clean_to_float(raw_val)
                                df_year.at[idx,"(百万)原币"] = dec_val
                                df_year.at[idx,"(百万)人民币"] = dec_val
                                df_year.at[idx,"汇率"] = "豁免换算"
                        else:
                            c_val = clean_to_float(raw_val)
                            df_year.at[idx,"(百万)原币"] = c_val * unit_mult
                            df_year.at[idx,"(百万)人民币"] = c_val * unit_mult * rate
    
                    final_cols = ["公司类型","公司","类别","字段名","字段类型","报告年份","(百万)原币","汇率","(百万)人民币"]
                    actual_cols = [c for c in final_cols if c in df_year.columns]
                    combined_list.append(df_year[actual_cols])
    
            return pd.concat(combined_list, ignore_index=True) if combined_list else pd.DataFrame()
    
        # ✅ 从列名自动识别单位的辅助函数
        def detect_unit_from_col(col_name):
            """从列名里读取单位关键词，返回对应的下拉选项文字"""
            s = str(col_name)
            if "十亿" in s: return "原表为: 十亿元 (× 1,000)"
            if "亿元" in s: return "原表为: 亿元 (× 100)"
            if "百万" in s: return "原表为: 百万元 (无需转换)"
            if "万元" in s or "万" in s: return "原表为: 万元 (÷ 100)"
            if "千元" in s: return "原表为: 千元 (÷ 1,000)"
            if "元" in s:   return "原表为: 元 (÷ 1,000,000)"
            return "原表为: 元 (÷ 1,000,000)"  # ← 默认改为“元”
    
        unit_multipliers = {
            "原表为: 百万元 (无需转换)": 1.0,
            "原表为: 万元 (÷ 100)": 0.01,
            "原表为: 元 (÷ 1,000,000)": 0.000001,
            "原表为: 千元 (÷ 1,000)": 0.001,
            "原表为: 亿元 (× 100)": 100.0,
            "原表为: 十亿元 (× 1,000)": 1000.0,
        }
    
        if uploaded_files:
            all_temp_data, found_companies = [], {}
    
            for file in uploaded_files:
                xl = pd.ExcelFile(file)
                for sheet_name in xl.sheet_names:
                    df_raw = pd.read_excel(file, sheet_name=sheet_name)
                    current_company = str(df_raw["公司"].iloc[0]) if "公司" in df_raw.columns and not df_raw["公司"].empty else sheet_name
                    if "基本信息" in current_company or "Sheet" in current_company: continue
    
                    # ✅ 从列名自动检测默认单位
                    import re
                    detected_unit = "原表为: 元 (÷ 1,000,000)"  # 先设默认值
                    for c in df_raw.columns:
                        if re.search(r'20\d{2}', str(c)):
                            detected_unit = detect_unit_from_col(c)
                            break
    
                    found_companies[current_company] = detected_unit
                    all_temp_data.append({"comp": current_company, "df": df_raw, "source": f"{file.name} - {sheet_name}"})
    
            st.markdown("#### 💵 汇率与数值单位配置盘")
            st.caption("目标表统一要求以【百万元人民币】展示。系统已根据列名自动识别默认单位，请人工核对后调整。")
    
            rate_config, unit_config = {}, {}
            rate_cols = st.columns(3)
            for i, (comp, default_unit) in enumerate(sorted(found_companies.items())):
                with rate_cols[i % 3]:
                    with st.container(border=True):
                        st.markdown(f"**🏢 {comp}**")
                        rate_config[comp] = st.number_input("汇率 (相对于RMB)", value=1.0, step=0.0001, format="%.4f", key=f"rate_{comp}")
                        unit_options = list(unit_multipliers.keys())
                        default_idx = unit_options.index(default_unit) if default_unit in unit_options else 0
                        unit_choice = st.selectbox("原表数值单位", unit_options, index=default_idx, key=f"unit_{comp}")
                        unit_config[comp] = unit_multipliers[unit_choice]
    
            if st.button("开始集成并换算数据", type="primary", use_container_width=True):
                with st.spinner("正在后台执行极速数据合并与换算..."):
                    final_all_df = run_data_integration(all_temp_data, rate_config, unit_config)
    
                if not final_all_df.empty:
                    st.session_state['integrated_data'] = final_all_df
                    years_found = sorted(final_all_df['报告年份'].unique().tolist())
                    st.success(f"✅ 集成与换算完毕！共处理 {len(found_companies)} 家公司，识别到年份：{' / '.join(years_found)}，生成 {len(final_all_df)} 条对标数据。")
                    st.dataframe(final_all_df, use_container_width=True)
    
                    import io
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        final_all_df.to_excel(writer, index=False, sheet_name='行业集成分析表')
                    st.download_button(label="下载行业集成目标表 (长表格式)", data=output.getvalue(), file_name="行业集成目标数据表.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                else:
                    st.error("未能从上传的文件中提取到有效数据，请检查列名中是否包含年份信息（如 2024YE、2025YE 等）。")

    # ----- Step 6 可视化分析面板 -----
        with tab6:
            st.markdown("### 📊 自定义对标分析")
        
            # ── 数据源选择 ───────────────────────────────────────────────────────
            if 'integrated_data' not in st.session_state:
                st.session_state['integrated_data'] = None
        
            source_choice = st.radio(
                "数据源选择", ["直接引用集成后的数据", "上传集成表 Excel"], horizontal=True
            )
            df_raw = None
        
            if source_choice == "上传集成表 Excel":
                viz_file = st.file_uploader("上传行业集成目标表", type=["xlsx"])
                if viz_file:
                    df_raw = pd.read_excel(viz_file)
                    st.session_state['integrated_data'] = df_raw
            else:
                df_raw = st.session_state.get('integrated_data')
        
            if df_raw is None:
                st.info("💡 请先完成数据集成或上传目标底稿。")
                st.stop()
        
            # ==========================================
            # 🌟 提速秘籍 1：锁定数据预处理 (只算一次)
            # ==========================================
            @st.cache_data(show_spinner=False)
            def prepare_viz_data(df_in):
                df_clean = df_in.copy()
                df_clean.columns = (
                    df_clean.columns.astype(str)
                    .str.strip()
                    .str.replace('\n', '', regex=False)
                    .str.replace('\r', '', regex=False)
                    .str.replace('\ufeff', '', regex=False)
                )
                df_clean['报告年份'] = df_clean['报告年份'].astype(str).str.replace('.0', '', regex=False)
                val_col = "(百万)人民币" if "(百万)人民币" in df_clean.columns else df_clean.columns[-1]
        
                # 去重
                dedup_cols = ['公司', '报告年份', '字段名']
                if '类别' in df_clean.columns:
                    dedup_cols.append('类别')
                df_clean = df_clean.drop_duplicates(subset=dedup_cols, keep='first')
        
                # 透视表
                df_pivot = (
                    df_clean.groupby(['公司', '报告年份', '字段名'])[val_col]
                    .sum().unstack('字段名').reset_index()
                )
        
                # 提取选项列表
                all_fields = sorted([x for x in df_clean['字段名'].unique().tolist() if isinstance(x, str) and x.strip()])
                all_types    = sorted(df_clean['类别'].dropna().astype(str).unique().tolist()) if '类别'    in df_clean.columns else []
                all_co_types = sorted(df_clean['公司类型'].dropna().astype(str).unique().tolist()) if '公司类型' in df_clean.columns else []
        
                all_years_sorted = sorted(
                    [y for y in df_clean['报告年份'].unique() if str(y).isdigit()],
                    key=lambda x: int(x)
                )
                
                return df_clean, df_pivot, val_col, all_fields, all_types, all_co_types, all_years_sorted
        
            # 调用缓存的数据处理函数
            df_clean, df_pivot, val_col, all_fields, all_types, all_co_types, all_years_sorted = prepare_viz_data(df_raw)
        
            # ── 动态年份颜色计算 (极速，无需缓存) ──────────────────────────────────
            KPMG_COLOR_MAP = {
                "Lightest": "#BFE8FF",
                "Light": "#76D2FF",
                "Primary": "#1E49E2",
                "Dark": "#00338D",
            }
            
            n_years = len(all_years_sorted)
            dynamic_year_colors = {}
            target_colors = []
            
            if n_years == 3:
                target_colors = [KPMG_COLOR_MAP["Light"], KPMG_COLOR_MAP["Primary"], KPMG_COLOR_MAP["Dark"]]
            elif n_years == 2:
                target_colors = [KPMG_COLOR_MAP["Primary"], KPMG_COLOR_MAP["Dark"]]
            else:
                target_colors = [KPMG_COLOR_MAP["Lightest"], KPMG_COLOR_MAP["Light"], KPMG_COLOR_MAP["Primary"], KPMG_COLOR_MAP["Dark"]]
            
            num_colors_to_use = len(target_colors)
            if n_years > 0:
                for idx, year in enumerate(all_years_sorted):
                    color_index = 0
                    if n_years > 1:
                        proportion = idx / (n_years - 1)
                        color_index = int(proportion * (num_colors_to_use - 1))
                    color_index = min(color_index, num_colors_to_use - 1)
                    dynamic_year_colors[str(year)] = target_colors[color_index]

            # ── KPMG 色卡 ────────────────────────────────────────────────────────
            # 定义 KPMG 官方色卡（可根据需要增删颜色分类）
            KPMG_CATEGORIES = {
                "KPMG 品牌色": {
                    "深蓝": "#00338D",
                    "浅蓝": "#0865EE",
                    "红色": "#C00000"
                },
                "辅助色": {
                    "绿色": "#92D050",
                    "紫色": "#7030A0",
                    "橙色": "#EF9867",
                    "青色": "#61CBF4",
                    "粉色": "#FEAED7"
                }
            }
            with st.expander("🎨 查看 KPMG 官方色卡"):
                for cat_name, cat_colors in KPMG_CATEGORIES.items():
                    st.markdown(f"**{cat_name}**")
                    html_str = "".join([
                        f'<div style="display:inline-block;margin-right:15px;margin-bottom:8px;">'
                        f'<div style="width:14px;height:14px;background-color:{c};display:inline-block;'
                        f'border-radius:3px;vertical-align:middle;border:1px solid #ddd;"></div>'
                        f'<span style="font-size:13px;vertical-align:middle;"> {n} <b>({c})</b></span></div>'
                        for n, c in cat_colors.items()
                    ])
                    st.markdown(html_str, unsafe_allow_html=True)
        
            st.divider()
            # ==========================================
            # 🌟 提速秘籍 2：锁定 UI 与绘图更新范围 (片段刷新)
            # ==========================================
            @st.fragment
            def render_viz_dashboard(df_clean, df_pivot, val_col, all_fields, all_types, all_co_types, dynamic_year_colors):
                import re, plotly.graph_objects as go, plotly.express as px
            
                def safe_num(s):
                    return pd.to_numeric(s.astype(str).str.strip().str.replace(",", "", regex=False).str.replace("(", "-", regex=False).str.replace(")", "", regex=False).replace(["", "nan", "None", "未披露", "—", "-"], np.nan), errors="coerce")
            
                def safe_formula_eval(wide_df, formula):
                    expr, local_dict = str(formula).strip(), {}
                    for i, field in enumerate(re.findall(r"\[(.*?)\]", expr)):
                        key = f"V{i}"
                        local_dict[key] = safe_num(wide_df[field]) if field in wide_df.columns else pd.Series(np.nan, index=wide_df.index)
                        expr = expr.replace(f"[{field}]", key)
                    try:
                        return pd.Series(pd.eval(expr, local_dict=local_dict, engine="python"), index=wide_df.index).replace([np.inf, -np.inf], np.nan)
                    except Exception:
                        return pd.Series(np.nan, index=wide_df.index)
            
                with st.expander("🛠️ 核心配置面板", expanded=True):
                    r1c1, r1c2, r1c3 = st.columns([1.5, 1, 1])
                    is_pct_stack_mode = False
            
                    with r1c1:
                        chart_type = st.selectbox("📈 1. 图表类型", ["簇状柱状图", "堆积柱状图", "折线对比图", "饼图", "内外环结构对比图", "散点图"])
            
                        if chart_type == "散点图":
                            calc_mode, plot_df_base = "散点图", pd.DataFrame()
                            st.caption("先选择字段；如需运算，开启高级公式，用 [字段名] 引用指标。")
                            x_field = st.selectbox("X轴字段", all_fields, index=0, key="scatter_x_field")
                            y_field = st.selectbox("Y轴字段", all_fields, index=min(1, len(all_fields)-1), key="scatter_y_field")
                            use_formula = st.checkbox("高级公式模式", value=False, key="scatter_formula_on")
                            x_formula = st.text_input("X轴公式", value=f"[{x_field}]", key="scatter_x_formula") if use_formula else f"[{x_field}]"
                            y_formula = st.text_input("Y轴公式", value=f"[{y_field}]", key="scatter_y_formula") if use_formula else f"[{y_field}]"
                            x_axis_title_custom = st.text_input("X轴显示名称", value=x_field, key="scatter_x_title")
                            y_axis_title_custom = st.text_input("Y轴显示名称", value=y_field, key="scatter_y_title")
            
                            all_cos_list = sorted(df_pivot['公司'].unique().tolist())
                            selected_cos = st.multiselect("🏗️ 选择公司", all_cos_list, default=all_cos_list[:min(8, len(all_cos_list))], key="scatter_selected_cos")
                            years_sorted = sorted(df_pivot['报告年份'].unique().tolist(), key=lambda x: int(x) if str(x).isdigit() else 9999)
                            scatter_years = st.multiselect("📅 选择年份", years_sorted, default=years_sorted[-1:], key="scatter_years")
            
                        elif "环" in chart_type or chart_type == "饼图":
                            calc_mode = "结构分析"
                            selected_multi_fields = st.multiselect("🎯 选择构成指标", all_fields, default=all_fields[:min(3, len(all_fields))])
                            selected_cos = st.selectbox("🏗️ 选择展示公司", sorted(df_pivot['公司'].unique().tolist()))
                            plot_df_base = df_clean[(df_clean['字段名'].isin(selected_multi_fields)) & (df_clean['公司'] == selected_cos)].copy().rename(columns={val_col: 'final_val'})
            
                        else:
                            if chart_type == "堆积柱状图":
                                stack_sub = st.radio("堆积模式", ["单指标 / 公式", "多指标占比"], horizontal=True, key="stack_sub_mode")
                                is_pct_stack_mode = (stack_sub == "多指标占比")
            
                            if is_pct_stack_mode:
                                calc_mode = "多指标占比"
                                pct_field_pool = sorted(df_clean[df_clean['类别'] == st.selectbox("🗂️ 按类别筛选指标", ["全部类别"] + all_types, key="pct_type_filter")]['字段名'].unique().tolist()) if all_types and st.session_state.get("pct_type_filter") not in [None, "全部类别"] else all_fields
                                selected_pct_fields = st.multiselect("🎯 选择堆积指标", pct_field_pool, default=pct_field_pool[:min(3, len(pct_field_pool))], key="pct_stack_fields")
                                plot_df_base = df_clean[df_clean['字段名'].isin(selected_pct_fields)][['公司', '报告年份', '字段名', val_col]].copy().rename(columns={val_col: 'final_val'}) if selected_pct_fields else pd.DataFrame(columns=['公司', '报告年份', '字段名', 'final_val'])
                            else:
                                calc_mode = st.radio("数据模式", ["单指标直显", "自定义公式运算"], horizontal=True)
                                if calc_mode == "单指标直显":
                                    filtered_fields = sorted(df_clean[df_clean['类别'] == st.selectbox("🗂️ 按类别筛选指标", ["全部类别"] + all_types, key="type_filter_single")]['字段名'].unique().tolist()) if all_types and st.session_state.get("type_filter_single") not in [None, "全部类别"] else all_fields
                                    target_field = st.selectbox("🎯 选择显示指标", filtered_fields)
                                    plot_df_base = df_pivot[['公司', '报告年份', target_field]].rename(columns={target_field: 'final_val'}).copy()
                                else:
                                    v1, v2 = st.columns(2)
                                    var_a = v1.selectbox("变量 A", all_fields, index=0)
                                    var_b = v2.selectbox("变量 B", ["无"] + all_fields, index=1 if len(all_fields) > 1 else 0)
                                    formula_input = st.text_input("✏️ 自定义公式 (使用 A 和 B)", value="A - B")
                                    try:
                                        A_data, B_data = safe_num(df_pivot[var_a]).fillna(0), (safe_num(df_pivot[var_b]).fillna(0) if var_b != "无" else 0)
                                        df_pivot['final_val'] = pd.Series(pd.eval(formula_input, local_dict={'A': A_data, 'B': B_data}), index=df_pivot.index).replace([np.inf, -np.inf], np.nan).fillna(0)
                                        plot_df_base = df_pivot[['公司', '报告年份', 'final_val']].copy()
                                    except Exception as e:
                                        st.error(f"公式无效: {e}")
                                        plot_df_base = pd.DataFrame(columns=['公司', '报告年份', 'final_val'])
            
                            all_cos_list = sorted(df_pivot['公司'].unique().tolist())
                            if all_co_types:
                                co_type_sel = st.selectbox("🏢 按公司类型快速选择", ["不按公司类型筛选"] + all_co_types, key="co_type_sel")
                                if co_type_sel != "不按公司类型筛选" and st.session_state.get('_prev_co_type_sel') != co_type_sel:
                                    st.session_state['selected_cos_ms'] = sorted(df_clean[df_clean['公司类型'] == co_type_sel]['公司'].unique().tolist())
                                    st.session_state['_prev_co_type_sel'] = co_type_sel
                            selected_cos = st.multiselect("🏗️ 选择对比公司", all_cos_list, default=all_cos_list[:min(2, len(all_cos_list))], key="selected_cos_ms")
            
                    with r1c2:
                        x_axis_mode = "散点视角" if chart_type == "散点图" else (st.radio("🔍 布局视角", ["以公司为横轴", "以年份为横轴"]) if "环" not in chart_type else "结构视角")
                        decimals = st.number_input("🔢 小数位数", 0, 4, 0)
                        show_value = st.toggle("✅ 显示数据标签", value=True)
            
                    with r1c3:
                        unit_options = {"原始数值": 1.0, "亿元": 0.01, "十亿元": 0.001, "百分比(%)": 100.0}
                        selected_unit = st.selectbox("📏 数值单位换算", list(unit_options.keys()))
                        multiplier = unit_options[selected_unit]
                        y_axis_title = st.text_input("📝 Y轴单位显示修改", value=f"单位: {selected_unit.split(' ')[0]}")
                        is_transparent = st.toggle("🌈 开启透明背景模式")
                        show_avg = st.checkbox("平均值线") if "柱" in chart_type else False
                        avg_color = st.color_picker("基准线颜色", value="#ED2124") if show_avg else "#ED2124"
            
                if chart_type == "散点图":
                    scatter_df = df_pivot.copy()
                    scatter_df["X"], scatter_df["Y"] = safe_formula_eval(scatter_df, x_formula) * multiplier, safe_formula_eval(scatter_df, y_formula) * multiplier
                    plot_df = scatter_df[scatter_df["公司"].isin(selected_cos) & scatter_df["报告年份"].isin(scatter_years)].dropna(subset=["X", "Y"]).copy()
                    if plot_df.empty:
                        st.warning("散点图暂无可用数据，请检查字段、公式、公司或年份。")
                        st.stop()
                    color_val = "公司"
                else:
                    if plot_df_base.empty:
                        st.warning("暂无数据，请检查筛选条件。")
                        st.stop()
                    if calc_mode in ("结构分析", "多指标占比"):
                        cos_filter = [selected_cos] if isinstance(selected_cos, str) else selected_cos
                        plot_df, color_val = plot_df_base[plot_df_base['公司'].isin(cos_filter)].copy(), "字段名"
                    else:
                        plot_df = plot_df_base[plot_df_base['公司'].isin(selected_cos)].copy()
                        color_val = "报告年份" if x_axis_mode == "以公司为横轴" else "公司"
                    plot_df['绘制金额'] = safe_num(plot_df['final_val']).fillna(0) * multiplier
            
                st.markdown("#### 🎨 自定义图例标签与颜色")
                legend_col = "公司" if chart_type == "散点图" else color_val
                unique_items = sorted(plot_df[legend_col].dropna().unique().tolist())
                rename_map, color_map, c_cols = {}, {}, st.columns(4)
                for i, item in enumerate(unique_items):
                    with c_cols[i % 4]:
                        st.caption(f"原始值: {item}")
                        new_label = st.text_input("显示名称", value=str(item), key=f"rename_{chart_type}_{item}")
                        default_c = dynamic_year_colors.get(str(item), DEFAULT_COLORS[i % len(DEFAULT_COLORS)] if 'DEFAULT_COLORS' in globals() else "#1E49E2")
                        new_color = st.color_picker("选择颜色", value=default_c, key=f"color_{chart_type}_{item}")
                        rename_map[item], color_map[new_label] = new_label, new_color
                if chart_type != "散点图":
                    plot_df[color_val] = plot_df[color_val].map(rename_map)
                else:
                    plot_df["_legend_show"] = plot_df[color_val].map(rename_map).fillna(plot_df[color_val])
            
                fig, fmt = go.Figure(), f'.{decimals}f'
                suffix, label_fmt = ("%" if selected_unit == "百分比(%)" else ""), f'%{{y:.{decimals}f}}{"%" if selected_unit == "百分比(%)" else ""}'
            
                if chart_type == "散点图":
                    for item in rename_map.values():
                        d = plot_df[plot_df["_legend_show"] == item]
                        if d.empty: continue
                        fig.add_trace(go.Scatter(
                            x=d["X"], y=d["Y"], name=str(item),
                            mode="markers+text" if show_value else "markers",
                            text=d["公司"] if show_value else None,
                            textposition="top center",
                            marker=dict(size=14, color=color_map.get(item, "#1E49E2"), line=dict(width=1.4, color="white")),
                            customdata=np.stack([d["公司"], d["报告年份"]], axis=-1),
                            hovertemplate="<b>%{customdata[0]}</b><br>年份：%{customdata[1]}<br>X=%{x}<br>Y=%{y}<extra></extra>"
                        ))
                    fig.update_xaxes(title=x_axis_title_custom, zeroline=True, zerolinecolor="#999", showgrid=True, gridcolor="#f0f0f0")
                    fig.update_yaxes(title=y_axis_title_custom, zeroline=True, zerolinecolor="#999", showgrid=True, gridcolor="#f0f0f0")
            
                elif "柱状图" in chart_type:
                    barmode = 'group' if "簇状" in chart_type else 'relative'
                    if is_pct_stack_mode:
                        plot_df['x_label'] = plot_df['公司'] + ' ' + plot_df['报告年份']
                        plot_df['_total'] = plot_df.groupby('x_label')['绘制金额'].transform('sum').replace(0, np.nan)
                        plot_df['绘制占比'] = (plot_df['绘制金额'] / plot_df['_total'] * 100).fillna(0)
                        for item in rename_map.values():
                            d = plot_df[plot_df[color_val] == item]
                            fig.add_trace(go.Bar(x=d['x_label'], y=d['绘制占比'], name=str(item), marker_color=color_map[item],
                                text=d.apply(lambda r: f"{r['绘制占比']:.{decimals}f}%<br>{r['绘制金额']:.{decimals}f}", axis=1) if show_value else None, textposition='inside'))
                        fig.update_yaxes(range=[0, 105]); barmode = 'relative'
                    else:
                        for item in rename_map.values():
                            d = plot_df[plot_df[color_val] == item]
                            fig.add_trace(go.Bar(x=d["公司" if color_val == "报告年份" else "报告年份"], y=d["绘制金额"], name=str(item), marker_color=color_map[item],
                                text=d["绘制金额"] if show_value else None, texttemplate=label_fmt if show_value else None, textposition='outside'))
                    fig.update_layout(barmode=barmode)
            
                elif "折线对比图" in chart_type:
                    for item in rename_map.values():
                        d = plot_df[plot_df[color_val] == item]
                        fig.add_trace(go.Scatter(x=d["公司" if color_val == "报告年份" else "报告年份"], y=d["绘制金额"], name=str(item),
                            mode='lines+markers+text' if show_value else 'lines+markers', marker_color=color_map[item],
                            text=d["绘制金额"], texttemplate=label_fmt, textposition="top center"))
            
                elif chart_type == "饼图":
                    d = plot_df[plot_df['报告年份'] == plot_df['报告年份'].max()]
                    fig = px.pie(d, values='绘制金额', names=color_val, hole=0.4, color=color_val, color_discrete_map=color_map)
                    fig.update_traces(textinfo='percent+label' if show_value else 'percent')
            
                elif chart_type == "内外环结构对比图":
                    years_ring = sorted(plot_df['报告年份'].unique().tolist())
                    if len(years_ring) < 2: st.warning("环形图对比需要至少两年的数据。")
                    else:
                        d_outer, d_inner = plot_df[plot_df['报告年份'] == years_ring[-1]], plot_df[plot_df['报告年份'] == years_ring[0]]
                        fig.add_trace(go.Pie(labels=d_outer[color_val], values=d_outer['绘制金额'], hole=0.7, name=years_ring[-1], marker=dict(colors=[color_map[f] for f in d_outer[color_val]]), textinfo='percent+label' if show_value else 'percent'))
                        fig.add_trace(go.Pie(labels=d_inner[color_val], values=d_inner['绘制金额'], hole=0.4, name=years_ring[0], domain={'x': [0.15, 0.85], 'y': [0.15, 0.85]}, marker=dict(colors=[color_map[f] for f in d_inner[color_val]]), textinfo='percent' if show_value else 'none'))
                        fig.update_layout(annotations=[dict(text=f'内:{years_ring[0]}<br>外:{years_ring[-1]}', x=0.5, y=0.5, font_size=12, showarrow=False)])
            
                bg_color = "rgba(0,0,0,0)" if is_transparent else "white"
                if show_avg and "柱" in chart_type and not is_pct_stack_mode:
                    avg_v = plot_df['绘制金额'].mean()
                    fig.add_hline(y=avg_v, line_dash="dash", line_color=avg_color, annotation_text=f"平均: {avg_v:{fmt}}{suffix}", annotation_font=dict(color=avg_color))
            
                fig.update_layout(font_family="Microsoft YaHei", plot_bgcolor=bg_color, paper_bgcolor=bg_color,
                    margin=dict(t=120, l=10, r=10, b=20),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                    annotations=[] if chart_type == "散点图" else [dict(x=0, y=1.18, xref='paper', yref='paper', text=f"<b>{y_axis_title}</b>", showarrow=False, font=dict(size=14, color="#333333"), xanchor='left')])
                fig.update_xaxes(showgrid=True if chart_type == "散点图" else False)
                fig.update_yaxes(showgrid=True, gridcolor="#f0f0f0")
            
                show_chart(fig, print_mode, m_id)
                with st.expander("📄 查看底层数据明细"):
                    st.dataframe(plot_df, use_container_width=True)
        
            # ==========================================
            # 在页面上调用封装好的交互展示区
            # ==========================================
            render_viz_dashboard(df_clean, df_pivot, val_col, all_fields, all_types, all_co_types, dynamic_year_colors)
        
    # ----- Step 7 公司级对标报告 -----
    with tab7:
        show_step_7_content()
    
    # ----- Step 8 行业分类统计分析 -----
    with tab8:
        show_step_8_content()

else:
    # 普通用户权限
    st.info("💡 普通用户仅可查看 Step 7 公司级对标报告。")
    show_step_7_content()

# ==================== 页脚 ====================
st.markdown("")
